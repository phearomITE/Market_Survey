from __future__ import annotations

from copy import copy
from pathlib import Path
import os
import shutil
import subprocess
import tempfile
import zipfile
from app.core.config import settings


def _find_soffice() -> str | None:
    """Find LibreOffice/soffice on Windows, macOS, or Linux."""
    env_path = settings.libreoffice_path or os.getenv("LIBREOFFICE_PATH") or os.getenv("SOFFICE_PATH")
    if env_path and Path(env_path).exists():
        return env_path

    found = shutil.which("soffice") or shutil.which("libreoffice")
    if found:
        return found

    candidates = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        "/usr/bin/soffice",
        "/usr/bin/libreoffice",
        "/snap/bin/libreoffice",
    ]
    for p in candidates:
        if Path(p).exists():
            return p
    return None


def _contains_khmer(value: object) -> bool:
    return isinstance(value, str) and any("\u1780" <= char <= "\u17ff" for char in value)


def _create_khmer_safe_render_copy(source: Path, destination: Path) -> int:
    """Save a rendering-only workbook with a font that shapes Khmer correctly.

    The user's generated Excel file is not changed. LibreOffice receives this
    temporary copy, preventing fallback fonts from separating Khmer coeng
    sequences such as the letters in "គ្រប់".
    """
    from openpyxl import load_workbook

    font_name = os.getenv("PNG_KHMER_FONT_NAME", "Khmer OS").strip() or "Khmer OS"
    workbook = load_workbook(source)
    changed = 0

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if not _contains_khmer(cell.value):
                    continue
                rendered_font = copy(cell.font)
                rendered_font.name = font_name
                rendered_font.scheme = None
                cell.font = rendered_font
                changed += 1

    workbook.save(destination)
    workbook.close()
    return changed


def excel_to_pdf(xlsx_path: Path) -> Path | None:
    """Convert Excel workbook to PDF using a Khmer-safe temporary render copy."""
    xlsx_path = Path(xlsx_path)
    soffice = _find_soffice()
    if not soffice or not xlsx_path.exists():
        return None

    outdir = xlsx_path.parent
    outdir.mkdir(parents=True, exist_ok=True)
    final_pdf = xlsx_path.with_suffix(".pdf")

    try:
        with tempfile.TemporaryDirectory(prefix="kb-khmer-render-") as temp_dir_name:
            temp_dir = Path(temp_dir_name)
            render_xlsx = temp_dir / xlsx_path.name

            try:
                changed = _create_khmer_safe_render_copy(xlsx_path, render_xlsx)
                print(f"🔤 Khmer-safe PNG render copy: {changed} cells")
            except Exception as exc:
                print(f"⚠️ Khmer render-copy preparation failed; using original workbook: {exc}")
                shutil.copy2(xlsx_path, render_xlsx)

            cmd = [
                soffice,
                "--headless",
                "--nologo",
                "--nofirststartwizard",
                "--convert-to",
                "pdf",
                "--outdir",
                str(temp_dir),
                str(render_xlsx),
            ]
            proc = subprocess.run(
                cmd,
                check=False,
                timeout=180,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                env={
                    **os.environ,
                    "LANG": os.getenv("LANG", "km_KH.UTF-8"),
                    "LC_ALL": os.getenv("LC_ALL", "km_KH.UTF-8"),
                },
            )
            rendered_pdf = render_xlsx.with_suffix(".pdf")
            if proc.returncode != 0 or not rendered_pdf.exists():
                print(f"⚠️ LibreOffice PDF conversion failed: {proc.stderr.strip()}")
                return None

            shutil.copy2(rendered_pdf, final_pdf)
    except Exception as exc:
        print(f"⚠️ Excel-to-PDF conversion failed: {exc}")
        return None

    return final_pdf if final_pdf.exists() and final_pdf.stat().st_size > 0 else None


def _crop_white_border(png_path: Path, padding: int = 35) -> None:
    """Remove big white page margins so Telegram preview is larger/easier to read."""
    try:
        from PIL import Image, ImageChops
    except Exception:
        return

    try:
        img = Image.open(png_path).convert("RGB")
        bg = Image.new("RGB", img.size, (255, 255, 255))
        diff = ImageChops.difference(img, bg)
        bbox = diff.getbbox()
        if not bbox:
            return

        left = max(bbox[0] - padding, 0)
        top = max(bbox[1] - padding, 0)
        right = min(bbox[2] + padding, img.width)
        bottom = min(bbox[3] + padding, img.height)
        cropped = img.crop((left, top, right, bottom))
        cropped.save(png_path, optimize=True)
    except Exception:
        return


def _resize_if_too_wide(png_path: Path, max_width: int = 6000) -> None:
    """Keep PNG readable but avoid very huge Telegram files."""
    try:
        from PIL import Image
    except Exception:
        return

    try:
        img = Image.open(png_path).convert("RGB")
        if img.width <= max_width:
            return
        ratio = max_width / float(img.width)
        new_size = (max_width, int(img.height * ratio))
        img = img.resize(new_size, Image.Resampling.LANCZOS)
        img.save(png_path, optimize=True)
    except Exception:
        return


def pdf_first_page_to_png(pdf_path: Path, png_path: Path | None = None) -> Path | None:
    """Convert first page of a PDF to a large Telegram-readable PNG.

    Fix for small preview:
    - render PDF at high DPI using PyMuPDF zoom
    - crop white margins around the Excel print area
    - send the PNG as a document in Telegram (handled in bot/handlers.py)
    """
    try:
        import fitz  # PyMuPDF
    except Exception:
        return None

    pdf_path = Path(pdf_path)
    if not pdf_path.exists():
        return None

    png_path = png_path or pdf_path.with_suffix(".png")
    try:
        # 4.5 zoom gives a much bigger, clearer report image than 2.2.
        # Override from .env if needed: PNG_RENDER_SCALE=5
        scale = float(os.getenv("PNG_RENDER_SCALE", "4.5"))
        doc = fitz.open(str(pdf_path))
        if len(doc) == 0:
            doc.close()
            return None
        page = doc[0]
        pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
        pix.save(str(png_path))
        doc.close()

        _crop_white_border(png_path, padding=25)
        _resize_if_too_wide(png_path, max_width=int(os.getenv("PNG_MAX_WIDTH", "6000")))
    except Exception:
        return None

    return png_path if png_path.exists() and png_path.stat().st_size > 0 else None


def excel_to_png(xlsx_path: Path) -> Path | None:
    """Create a large PNG preview from the first page/sheet of the Excel report."""
    pdf = excel_to_pdf(Path(xlsx_path))
    if not pdf:
        return None
    return pdf_first_page_to_png(pdf, Path(xlsx_path).with_suffix(".png"))



def excel_workbook_to_png_zip(
    xlsx_path: Path,
    sheet_names: list[str] | None = None,
    zip_path: Path | None = None,
) -> Path | None:
    """Render every PDF page from an Excel workbook into one PNG ZIP.

    Used by /report_today:
      - one Excel workbook contains 65 dealer sheets
      - LibreOffice exports the workbook to a multi-page PDF
      - each PDF page is rendered as a PNG
      - all PNG previews are packed into one ZIP for Telegram

    If LibreOffice/PyMuPDF is unavailable, returns None without breaking Excel output.
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        return None

    pdf = excel_to_pdf(xlsx_path)
    if not pdf:
        return None

    try:
        import fitz  # PyMuPDF
    except Exception:
        return None

    out_dir = xlsx_path.parent / f"{xlsx_path.stem}_png"
    out_dir.mkdir(parents=True, exist_ok=True)
    zip_path = zip_path or xlsx_path.with_name(f"{xlsx_path.stem}_PNG_65_Dealers.zip")

    try:
        # Clean old PNGs for the same workbook to avoid sending stale files.
        for old in out_dir.glob("*.png"):
            try:
                old.unlink()
            except Exception:
                pass

        scale = float(os.getenv("PNG_RENDER_SCALE", "4.5"))
        doc = fitz.open(str(pdf))
        png_files: list[Path] = []

        for i, page in enumerate(doc):
            if sheet_names and i < len(sheet_names):
                base_name = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in str(sheet_names[i]))
                base_name = base_name.strip("_") or f"dealer_{i+1:02d}"
            else:
                base_name = f"dealer_{i+1:02d}"

            png_path = out_dir / f"{i+1:02d}_{base_name}.png"
            pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
            pix.save(str(png_path))
            _crop_white_border(png_path, padding=25)
            _resize_if_too_wide(png_path, max_width=int(os.getenv("PNG_MAX_WIDTH", "6000")))
            if png_path.exists() and png_path.stat().st_size > 0:
                png_files.append(png_path)

        doc.close()

        if not png_files:
            return None

        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for file in png_files:
                zf.write(file, arcname=file.name)

        return zip_path if zip_path.exists() and zip_path.stat().st_size > 0 else None
    except Exception as exc:
        print(f"⚠️ Excel workbook PNG ZIP failed: {exc}")
        return None
