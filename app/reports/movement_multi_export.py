from __future__ import annotations

from copy import copy
from datetime import date, datetime
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SHEET_NAME = "Detail_Movement"
OUTPUT_COLUMNS = (
    "Date",
    "Region",
    "Dealer",
    "Latitude",
    "Longitude",
    "Outlet Name",
    "Outlet Type",
    "Phone Number Outlet",
    "CB LITE NCP",
    "GB SNOW NCP",
    "Hanuman LITE NCP",
    "Krud LITE NCP",
    "Greet LITE NCP",
)
MOVEMENT_PRODUCTS = OUTPUT_COLUMNS[8:]


def _product_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").casefold())


PRODUCT_ALIASES = {
    _product_key("CB LITE NCP"): {
        _product_key("CB LITE NCP"),
        _product_key("CB LITE"),
        _product_key("CBC LITE NCP"),
        _product_key("CBC LITE"),
    },
    _product_key("GB SNOW NCP"): {
        _product_key("GB SNOW NCP"),
        _product_key("GB SNOW"),
    },
    _product_key("Hanuman LITE NCP"): {
        _product_key("Hanuman LITE NCP"),
        _product_key("Hanuman Lite"),
    },
    _product_key("Krud LITE NCP"): {
        _product_key("Krud LITE NCP"),
        _product_key("Krud Lite"),
    },
    _product_key("Greet LITE NCP"): {
        _product_key("Greet LITE NCP"),
        _product_key("Greet Lite"),
        _product_key("Great LITE NCP"),
        _product_key("Great Lite"),
    },
}


def parse_movement_multi_dates(values: list[str] | tuple[str, ...]) -> list[date]:
    """Parse 1–10 distinct ISO dates while preserving command order."""
    if not values:
        raise ValueError(
            "Usage: /export movement_multi 2026-07-04 2026-07-18 2026-07-25"
        )
    if len(values) > 10:
        raise ValueError("Maximum 10 report dates per movement_multi export.")

    result: list[date] = []
    seen: set[date] = set()
    for raw in values:
        try:
            parsed = datetime.strptime(str(raw).strip(), "%Y-%m-%d").date()
        except ValueError as exc:
            raise ValueError(
                f"Invalid date '{raw}'. Dates must use YYYY-MM-DD."
            ) from exc
        if parsed not in seen:
            seen.add(parsed)
            result.append(parsed)
    return result


def _score_map(submission) -> dict[str, int | float]:
    scores: dict[str, int | float] = {}
    for metric in getattr(submission, "product_metrics", ()) or ():
        score = getattr(metric, "movement_score", None)
        if score is None:
            continue
        try:
            numeric = float(score)
        except (TypeError, ValueError):
            continue
        if not 0 <= numeric <= 10:
            continue

        metric_key = _product_key(getattr(metric, "product_name", ""))
        for output_product in MOVEMENT_PRODUCTS:
            aliases = PRODUCT_ALIASES[_product_key(output_product)]
            if metric_key in aliases:
                scores[output_product] = int(numeric) if numeric.is_integer() else numeric
                break
    return scores


def build_movement_rows(submissions: list) -> list[list[object]]:
    """Build one detail row per Kobo outlet submission containing beer movement."""
    rows: list[list[object]] = []
    ordered = sorted(
        submissions,
        key=lambda item: (
            getattr(item, "report_date", None) or date.min,
            str(getattr(item, "region", "") or ""),
            str(getattr(item, "dealer", "") or ""),
            str(getattr(item, "outlet_name", "") or ""),
            getattr(item, "submission_time", None) or datetime.min,
            int(getattr(item, "id", 0) or 0),
        ),
    )
    for submission in ordered:
        scores = _score_map(submission)
        if not scores:
            continue
        rows.append(
            [
                getattr(submission, "report_date", None),
                getattr(submission, "region", None),
                getattr(submission, "dealer", None),
                getattr(submission, "gps_latitude", None),
                getattr(submission, "gps_longitude", None),
                getattr(submission, "outlet_name", None),
                getattr(submission, "outlet_type", None),
                getattr(submission, "phone_number", None),
                *[scores.get(product) for product in MOVEMENT_PRODUCTS],
            ]
        )
    return rows


def create_movement_multi_workbook(
    submissions: list,
    report_dates: list[date],
    *,
    template_path: Path | None = None,
    output_path: Path | None = None,
) -> Path:
    if template_path is None or output_path is None:
        from app.core.config import settings
    template = Path(template_path or settings.movement_multi_template_file)
    if not template.exists():
        raise FileNotFoundError(f"Movement export template not found: {template}")

    if output_path is None:
        settings.export_path.mkdir(parents=True, exist_ok=True)
        date_label = "_".join(item.isoformat() for item in report_dates)
        output_path = settings.export_path / f"Detail_Movement_Beer_{date_label}.xlsx"
    else:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(template)
    worksheet = workbook[SHEET_NAME]
    headers = [worksheet.cell(1, column).value for column in range(1, len(OUTPUT_COLUMNS) + 1)]
    if tuple(headers) != OUTPUT_COLUMNS:
        raise ValueError(
            "Movement template headers do not match the required layout. "
            f"Expected {OUTPUT_COLUMNS}; found {tuple(headers)}."
        )

    template_row = 2
    row_style = []
    for column in range(1, len(OUTPUT_COLUMNS) + 1):
        source = worksheet.cell(template_row, column)
        row_style.append(
            {
                "font": copy(source.font),
                "fill": copy(source.fill),
                "border": copy(source.border),
                "alignment": copy(source.alignment),
                "number_format": source.number_format,
                "protection": copy(source.protection),
            }
        )

    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row - 1)

    rows = build_movement_rows(submissions)
    for row_index, values in enumerate(rows, start=2):
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row_index, column, value)
            style = row_style[column - 1]
            cell.font = copy(style["font"])
            cell.fill = copy(style["fill"])
            cell.border = copy(style["border"])
            cell.alignment = copy(style["alignment"])
            cell.number_format = style["number_format"]
            cell.protection = copy(style["protection"])
        worksheet.cell(row_index, 1).number_format = "yyyy-mm-dd"
        worksheet.cell(row_index, 4).number_format = "0.000000"
        worksheet.cell(row_index, 5).number_format = "0.000000"
        worksheet.cell(row_index, 8).number_format = "@"
        worksheet.cell(row_index, 8).alignment = copy(worksheet.cell(row_index, 6).alignment)

    final_row = max(1, len(rows) + 1)
    for table in worksheet.tables.values():
        table.ref = f"A1:{get_column_letter(len(OUTPUT_COLUMNS))}{final_row}"

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:M{final_row}"
    worksheet.column_dimensions["H"].width = max(
        worksheet.column_dimensions["H"].width or 0,
        21,
    )
    for column in ("I", "J", "K", "L", "M"):
        worksheet.column_dimensions[column].width = max(
            worksheet.column_dimensions[column].width or 0,
            20,
        )
    workbook.save(output_path)
    return output_path
