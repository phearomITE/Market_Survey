from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.core.config import settings


def create_submission_status_export(rows: list[dict], report_date: date) -> Path:
    output = settings.export_path / f"Dealer_Summary_Status_{report_date}.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Summary Status"
    sheet.append(["Date", "Region", "Dealer", "Status"])
    for row in rows:
        sheet.append([
            row["date"].isoformat(),
            row["region"],
            row["dealer"],
            row["status"],
        ])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    missing_fill = PatternFill("solid", fgColor="FFC7CE")
    submitted_fill = PatternFill("solid", fgColor="C6EFCE")
    for cell in sheet["D"][1:]:
        cell.fill = submitted_fill if cell.value == "Summary Submitted" else missing_fill

    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 10
    sheet.column_dimensions["C"].width = 12
    sheet.column_dimensions["D"].width = 22
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    # Use the worksheet AutoFilter only. Combining a generated Table object
    # with a second overlapping filter caused desktop Excel to repair/remove
    # xl/tables/table1.xml. The normal filter arrows remain fully functional.
    workbook.save(output)
    return output
