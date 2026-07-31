from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.config import settings
from app.reports.aggregator import (
    ALL_COMPETITOR_PRODUCTS,
    ALL_OWN_PRODUCTS,
    COMPETITOR_PRODUCTS,
    HORECA_COMPETITOR_PRODUCTS,
    HORECA_OWN_PRODUCTS,
    OWN_PRODUCTS,
    aggregate_submissions,
)


BEER_PRODUCTS = [
    "CB LITE ORD",
    "GB SNOW ORD",
    "HANUMAN LITE ORD",
    "Krud LITE ORD",
    "CBC 4.4 NCP",
    "CB Original NCP",
    "GB Original NCP",
    "Krud NCP",
    "CB LITE NCP",
    "GB SNOW NCP",
    "Hanuman LITE NCP",
    "Krud LITE NCP",
    "Greet LITE NCP",
    "CB BLACK NCP",
    "Hanuman Black NCP",
]

BASE_HEADERS = [
    "Date",
    "Region",
    "Dealer",
    "Latitude",
    "Longitude",
    "Outlet Name",
    "Outlet Type",
    "Phone Number Outlet",
]

SUMMARY_HEADERS = [
    "Region", "Dealer", "Location of Visit Text", "Member", "Total Outlets",
    "Wholesale", "Drink Shop", "Wet Market", "Trolley", "Local Eat",
    "Coffe,Bakery", "Canteen", "Sport Club", "Motor Shop", "Product",
    "WS", "DS", "WM", "TL", "LE", "CB", "MS", "Movement",
]
RAW_HEADERS = ["Date", "Region", "Dealer", "Product", "Movement Rate"]
HORECA_OUTLET_TYPES = {
    "Local Eat", "Coffee,Bakery", "Canteen", "Sport Club", "Motor Shop", "Local Drink",
}


def _clean(value) -> str:
    return str(value or "").strip()


def _metric_scores(submission) -> dict[str, int]:
    scores: dict[str, int] = {}
    for metric in list(getattr(submission, "product_metrics", None) or []):
        value = getattr(metric, "movement_score", None)
        if value is not None:
            scores[_clean(getattr(metric, "product_name", None))] = int(value)
    for metric in list(getattr(submission, "competitor_metrics", None) or []):
        value = getattr(metric, "movement_score", None)
        if value is not None:
            scores[_clean(getattr(metric, "product_name", None))] = int(value)
    return scores


def _canonical_products() -> list[str]:
    return list(dict.fromkeys(ALL_OWN_PRODUCTS + ALL_COMPETITOR_PRODUCTS))


def _products_for_submission(submission) -> list[str]:
    report_type = _clean(getattr(submission, "report_type", None)).upper()
    outlet_type = _clean(getattr(submission, "outlet_type", None))
    if report_type == "HORECA" or outlet_type in HORECA_OUTLET_TYPES:
        return list(dict.fromkeys(HORECA_OWN_PRODUCTS + HORECA_COMPETITOR_PRODUCTS))
    return list(dict.fromkeys(OWN_PRODUCTS + COMPETITOR_PRODUCTS))


def _members_text(submissions: list) -> str:
    values: set[int] = set()
    for item in submissions:
        value = getattr(item, "member_no", None)
        if value not in (None, ""):
            try:
                values.add(int(value))
            except (TypeError, ValueError):
                pass
    return ", ".join(str(value) for value in sorted(values))


def _summary_product_data(agg: dict, product: str) -> dict:
    return (
        (agg.get("products") or {}).get(product)
        or (agg.get("competitors") or {}).get(product)
        or {}
    )


def _apply_header_style(ws, headers: list[str]) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(ws.max_row, 1)}"
    ws.row_dimensions[1].height = 25


def create_daily_export(
    submissions: Iterable,
    report_date: date,
    output_path: Path | None = None,
) -> Path:
    """Create Summary_Data + Location_Outlet for /export DATE.

    Movement comes from aggregate_submissions(), exactly like /report.
    """
    rows = list(submissions)
    grouped: dict[tuple[str, str], list] = defaultdict(list)
    for submission in rows:
        key = (
            _clean(getattr(submission, "region", None)),
            _clean(getattr(submission, "dealer", None)),
        )
        grouped[key].append(submission)

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary_Data"
    summary_ws.append(SUMMARY_HEADERS)
    outlet_keys = [
        "Wholesale", "Drink Shop", "Wet Market", "Trolley", "Local Eat",
        "Coffee,Bakery", "Canteen", "Sport Club", "Motor Shop",
    ]
    availability_keys = [
        "Wholesale", "Drink Shop", "Wet Market", "Trolley",
        "Local Eat", "Coffee,Bakery", "Motor Shop",
    ]

    for key in sorted(grouped, key=lambda value: (value[0], value[1])):
        dealer_rows = grouped[key]
        agg = aggregate_submissions(dealer_rows)
        outlet_types = agg.get("outlet_types") or Counter()
        base = [
            agg.get("region") or key[0],
            agg.get("dealer") or key[1],
            agg.get("location_text") or "",
            _members_text(dealer_rows),
            agg.get("total_outlets") or 0,
            *[int(outlet_types.get(name, 0) or 0) for name in outlet_keys],
        ]
        for product in _canonical_products():
            pdata = _summary_product_data(agg, product)
            availability = pdata.get("availability") or Counter()
            summary_ws.append([
                *base,
                product,
                *[int(availability.get(name, 0) or 0) for name in availability_keys],
                pdata.get("mov"),
            ])

    _apply_header_style(summary_ws, SUMMARY_HEADERS)
    widths = [11, 12, 45, 18, 14, 12, 14, 12, 10, 12, 15, 12, 12, 12, 28, 8, 8, 8, 8, 8, 8, 8, 13]
    for index, width in enumerate(widths, start=1):
        summary_ws.column_dimensions[get_column_letter(index)].width = width

    location_ws = wb.create_sheet("Location_Outlet")
    location_ws.append(BASE_HEADERS)
    for submission in sorted(rows, key=lambda item: (
        getattr(item, "report_date", None) or date.min,
        _clean(getattr(item, "region", None)),
        _clean(getattr(item, "dealer", None)),
        _clean(getattr(item, "outlet_name", None)),
    )):
        location_ws.append([
            getattr(submission, "report_date", None),
            getattr(submission, "region", None),
            getattr(submission, "dealer", None),
            getattr(submission, "gps_latitude", None),
            getattr(submission, "gps_longitude", None),
            getattr(submission, "outlet_name", None),
            getattr(submission, "outlet_type", None),
            _clean(getattr(submission, "phone_number", None)) or "N/A",
        ])
    for cell in location_ws["A"][1:]:
        cell.number_format = "yyyy-mm-dd"
    for cell in location_ws["H"][1:]:
        cell.number_format = "@"
    _apply_header_style(location_ws, BASE_HEADERS)
    for index, width in enumerate([13, 10, 12, 13, 13, 25, 18, 22], start=1):
        location_ws.column_dimensions[get_column_letter(index)].width = width

    settings.export_path.mkdir(parents=True, exist_ok=True)
    output_path = output_path or settings.export_path / f"Market_Survey_Data_{report_date}.xlsx"
    wb.save(output_path)
    return output_path


def create_raw_movement_long_export(
    submissions: Iterable,
    report_date: date,
    output_path: Path | None = None,
) -> Path:
    """Stream the five-column outlet-product detail export."""
    rows = list(submissions)
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Raw_Movement")
    ws.freeze_panes = "A2"
    for column, width in {"A": 13, "B": 10, "C": 12, "D": 28, "E": 16}.items():
        ws.column_dimensions[column].width = width

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_row = []
    for value in RAW_HEADERS:
        cell = WriteOnlyCell(ws, value=value)
        cell.fill = header_fill
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        header_row.append(cell)
    ws.append(header_row)

    written = 1
    for submission in sorted(rows, key=lambda item: (
        _clean(getattr(item, "region", None)),
        _clean(getattr(item, "dealer", None)),
        _clean(getattr(item, "outlet_name", None)),
    )):
        scores = _metric_scores(submission)
        for product in _products_for_submission(submission):
            ws.append([
                getattr(submission, "report_date", None),
                getattr(submission, "region", None),
                getattr(submission, "dealer", None),
                product,
                scores.get(product, 0),
            ])
            written += 1
    ws.auto_filter.ref = f"A1:E{written}"

    settings.export_path.mkdir(parents=True, exist_ok=True)
    output_path = output_path or settings.export_path / f"Raw_Movement_GENERAL_{report_date}.xlsx"
    wb.save(output_path)
    return output_path


def _ordered_products(submissions: Iterable, beer_only: bool) -> list[str]:
    rows = list(submissions)
    present = {
        product
        for submission in rows
        for product in _metric_scores(submission)
        if product
    }
    preferred = BEER_PRODUCTS if beer_only else list(dict.fromkeys(ALL_OWN_PRODUCTS + ALL_COMPETITOR_PRODUCTS))
    ordered = [product for product in preferred if product in present]
    ordered.extend(sorted(present.difference(ordered), key=str.casefold))
    return ordered


def _style_sheet(ws, max_column: int) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max_column)}{max(ws.max_row, 1)}"
    widths = [13, 10, 12, 13, 13, 25, 18, 22]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for index in range(len(BASE_HEADERS) + 1, max_column + 1):
        ws.column_dimensions[get_column_letter(index)].width = 20
    ws.row_dimensions[1].height = 25


def create_movement_export(
    submissions: Iterable,
    report_dates: list[date],
    *,
    beer_only: bool,
    output_path: Path | None = None,
) -> Path:
    """Create a safe, table-free movement workbook.

    One row represents one outlet submission and each product has its own
    movement column. Blank movement remains blank; a genuine zero remains 0.
    """
    rows = sorted(
        list(submissions),
        key=lambda item: (
            getattr(item, "report_date", None) or date.min,
            _clean(getattr(item, "region", None)),
            _clean(getattr(item, "dealer", None)),
            _clean(getattr(item, "outlet_name", None)),
        ),
    )
    products = _ordered_products(rows, beer_only=beer_only)
    headers = BASE_HEADERS + products

    wb = Workbook()
    ws = wb.active
    ws.title = "Movement"
    ws.append(headers)

    for submission in rows:
        scores = _metric_scores(submission)
        ws.append(
            [
                getattr(submission, "report_date", None),
                getattr(submission, "region", None),
                getattr(submission, "dealer", None),
                getattr(submission, "gps_latitude", None),
                getattr(submission, "gps_longitude", None),
                getattr(submission, "outlet_name", None),
                getattr(submission, "outlet_type", None),
                getattr(submission, "phone_number", None),
                *[scores.get(product) for product in products],
            ]
        )

    for cell in ws["A"][1:]:
        cell.number_format = "yyyy-mm-dd"
    _style_sheet(ws, len(headers))

    settings.export_path.mkdir(parents=True, exist_ok=True)
    if output_path is None:
        joined_dates = "_".join(str(item) for item in report_dates)
        prefix = "Detail_Movement_Beer" if beer_only else "Raw_Movement_GENERAL"
        output_path = settings.export_path / f"{prefix}_{joined_dates}.xlsx"
    wb.save(output_path)
    return output_path
