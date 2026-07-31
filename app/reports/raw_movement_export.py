from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.core.config import settings


HEADERS = ("Date", "Region", "Dealer", "Product", "Movement Rate")


def create_raw_movement_export(submissions: list, report_date: date) -> tuple[Path, int]:
    rows: list[tuple] = []
    for submission in submissions:
        metrics = [
            *(getattr(submission, "product_metrics", ()) or ()),
            *(getattr(submission, "competitor_metrics", ()) or ()),
        ]
        for metric in metrics:
            score = getattr(metric, "movement_score", None)
            product = str(getattr(metric, "product_name", "") or "").strip()
            if score is None or not product:
                continue
            try:
                numeric = float(score)
            except (TypeError, ValueError):
                continue
            if not 0 <= numeric <= 10:
                continue
            rows.append((
                report_date.isoformat(),
                getattr(submission, "region", None),
                getattr(submission, "dealer", None),
                product,
                int(numeric) if numeric.is_integer() else numeric,
            ))

    rows.sort(key=lambda row: (str(row[1] or ""), str(row[2] or ""), str(row[3] or ""), row[4]))
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Raw_Movement"
    worksheet.append(HEADERS)
    for row in rows:
        worksheet.append(row)

    header_fill = PatternFill("solid", fgColor="D90B2B")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    worksheet.freeze_panes = "A2"
    widths = (14, 10, 12, 30, 16)
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[chr(64 + index)].width = width
    if rows:
        table = Table(displayName="RawMovementTable", ref=f"A1:E{len(rows) + 1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    settings.export_path.mkdir(parents=True, exist_ok=True)
    output = settings.export_path / f"Raw_Movement_GENERAL_{report_date.isoformat()}.xlsx"
    workbook.save(output)
    return output, len(rows)
