from __future__ import annotations

from collections import defaultdict
from datetime import date
from pathlib import Path
import re
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from app.core.config import BASE_DIR, settings
from app.reports.aggregator import (
    OFFTAKE_COMPARE_GROUPS,
    build_bulk_dealer_aggregates,
    combine_location_visit,
)
from app.reports.member_mode import most_frequent_member


SUMMARY_MARKERS = {
    "បូកសរុបរួម",
    "បូកសរុបរូម",
    "សរុបរួម",
    "បួកសរុបរួម",
}

SUMMARY_SHEET = "Summary_Data"
LOCATION_SHEET = "Location_Outlet"
DATA_EXPORT_TEMPLATE = BASE_DIR / "templates" / "Template_Data_Survey.xlsx"

# These are the expected headers in the user's approved template. The exporter
# reads the workbook's real row-1 headers at runtime and does not replace them.
SUMMARY_HEADERS = [
    "Region",
    "Dealer",
    "Location of Visit Text",
    "Member",
    "Total Outlets",
    "Wholesale",
    "Drink Shop",
    "Wet Market",
    "Trolley",
    "Local Eat",
    "Coffe,Bakery",
    "Canteen",
    "Sport Club",
    "Motor Shop",
    "Product",
    "WS",
    "DS",
    "WM",
    "TL",
    "LE",
    "CB",
    "MS",
    "Movement",
]

LOCATION_HEADERS = [
    "Date",
    "Region",
    "Dealer",
    "Latitude",
    "Longitude",
    "Outlet Name",
    "Outlet Type",
    "Phone Number Outlet",
]

# The comparison groups define the exact report/export product order.
EXPORT_PRODUCTS: tuple[str, ...] = tuple(
    product
    for group in OFFTAKE_COMPARE_GROUPS
    for product in group
)

PRODUCT_DISPLAY_ALIASES = {
    "CAMBODIA Sport 300ml": "CAMBODIA Sport 300mL",
}

PRODUCT_LOOKUP_ALIASES = {
    "CAMBODIA Sport 300ml": (
        "CAMBODIA Sport 300mL",
        "CAMBODIA Sport 300ml",
    ),
    "EXPREZ Can 330ml": (
        "EXPREZ Can 330ml",
    ),
}

OUTLET_TYPE_KEYS = {
    "wholesale": {"wholesale"},
    "drinkshop": {"drinkshop"},
    "wetmarket": {"wetmarket"},
    "trolley": {"trolley"},
    "localeat": {"localeat"},
    # Support both the correct spelling and the approved template's "Coffe".
    "coffeebakery": {"coffeebakery", "coffebakery"},
    "canteen": {"canteen"},
    "sportclub": {"sportclub"},
    "motorshop": {"motorshop"},
}

# General movement must match the GENERAL Market Improvement Report exactly.
# These outlet types are therefore excluded only from movement calculation;
# they are still included in total outlet counts and product availability.
CHANNEL_SPECIALIST_TYPE_KEYS = set().union(
    OUTLET_TYPE_KEYS["localeat"],
    OUTLET_TYPE_KEYS["coffeebakery"],
    OUTLET_TYPE_KEYS["canteen"],
    OUTLET_TYPE_KEYS["sportclub"],
    OUTLET_TYPE_KEYS["motorshop"],
)

SUMMARY_REQUIRED_HEADER_KEYS = {
    "region",
    "dealer",
    "locationofvisittext",
    "member",
    "totaloutlets",
    "product",
    "ws",
    "ds",
    "wm",
    "tl",
    "le",
    "cb",
    "ms",
    "movement",
}
LOCATION_REQUIRED_HEADER_KEYS = {
    "date",
    "region",
    "dealer",
    "latitude",
    "longitude",
    "outletname",
    "outlettype",
    "phonenumberoutlet",
}


def _clean(value: Any) -> str:
    return str(value or "").strip()


def _header_key(value: Any) -> str:
    """Normalize a template header while preserving the user's column order."""
    return re.sub(r"[^a-z0-9]+", "", _clean(value).lower())


def _is_summary_submission(submission: Any) -> bool:
    return _clean(getattr(submission, "outlet_name", None)) in SUMMARY_MARKERS


def _region_sort(value: Any) -> tuple[int, str]:
    text = _clean(value).upper()
    match = re.fullmatch(r"R(\d+)", text)
    return (int(match.group(1)), text) if match else (999, text)


def _numeric_text_sort(value: Any) -> tuple[int, Any]:
    if value in (None, ""):
        return (2, "")
    try:
        return (0, int(value))
    except (TypeError, ValueError):
        return (1, _clean(value).lower())


def _group_sort_key(item: tuple[tuple[str, str], list[Any]]):
    (region, dealer), _rows = item
    return (_region_sort(region), _clean(dealer).upper())


def _location_sort_key(submission: Any):
    return (
        _region_sort(getattr(submission, "region", None)),
        _clean(getattr(submission, "dealer", None)).upper(),
        _numeric_text_sort(getattr(submission, "member_no", None)),
        _clean(getattr(submission, "outlet_name", None)).lower(),
    )


def _normalize_outlet_type(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", _clean(value).lower())


def _count_type(counter: Any, target_keys: set[str]) -> int:
    if not counter:
        return 0
    items = counter.items() if hasattr(counter, "items") else []
    return sum(
        int(count or 0)
        for label, count in items
        if _normalize_outlet_type(label) in target_keys
    )


def _joined_unique(rows: Iterable[Any], attribute: str) -> str:
    values = {
        getattr(row, attribute, None)
        for row in rows
        if getattr(row, attribute, None) not in (None, "")
    }
    return ", ".join(_clean(value) for value in sorted(values, key=_numeric_text_sort))


def _joined_locations(rows: Iterable[Any]) -> str:
    """Combine real submitted location text without duplicate place variants.

    The shared aggregator helper performs case-insensitive normalization, Khmer
    alias matching and spelling-variant matching. Examples such as Psar prek
    pnov, psar prek pov, Praek Pnov, ព្រែកព្នៅ and Pnov become one label:
    Prek Pnov. Saroang and Samroang become Samroang.
    """
    values = [getattr(row, "location_text", None) for row in rows]
    return combine_location_visit(values)


def _product_metric(aggregate: dict[str, Any], product: str) -> dict[str, Any] | None:
    display_name = PRODUCT_DISPLAY_ALIASES.get(product, product)
    candidates: list[str] = [product, display_name]
    candidates.extend(PRODUCT_LOOKUP_ALIASES.get(product, ()))

    # Own-product values take priority for cross-over products such as
    # CB Original NCP, CAMBODIA Sport 300mL and EXPREZ Can 330ml.
    for bucket_name in ("products", "competitors"):
        bucket = aggregate.get(bucket_name) or {}
        for candidate in dict.fromkeys(candidates):
            metric = bucket.get(candidate)
            if isinstance(metric, dict):
                return metric
    return None


def _movement_for_product(aggregate: dict[str, Any], product: str):
    metric = _product_metric(aggregate, product)
    return metric.get("mov") if metric else None


def _availability_for_product(aggregate: dict[str, Any], product: str):
    metric = _product_metric(aggregate, product)
    return (metric.get("availability") or {}) if metric else {}


def _coordinates(submission: Any) -> tuple[float | None, float | None]:
    lat = getattr(submission, "gps_latitude", None)
    lon = getattr(submission, "gps_longitude", None)
    if lat is not None or lon is not None:
        return lat, lon

    gps_text = _clean(getattr(submission, "gps_text", None))
    if not gps_text:
        return None, None

    numbers = re.findall(r"-?\d+(?:\.\d+)?", gps_text)
    if len(numbers) < 2:
        return None, None
    try:
        return float(numbers[0]), float(numbers[1])
    except ValueError:
        return None, None


def _template_headers(ws) -> list[str]:
    """Read the real template headers without rewriting the user's workbook."""
    last_column = 0
    for column in range(1, ws.max_column + 1):
        if _clean(ws.cell(1, column).value):
            last_column = column
    if last_column == 0:
        raise ValueError(f"Sheet {ws.title!r} has no headers in row 1")
    return [_clean(ws.cell(1, column).value) for column in range(1, last_column + 1)]


def _validate_headers(sheet_name: str, headers: list[str], required_keys: set[str]) -> None:
    present = {_header_key(header) for header in headers if _clean(header)}
    missing = sorted(required_keys - present)
    if missing:
        raise ValueError(
            f"{sheet_name} template is missing required column(s): " + ", ".join(missing)
        )


def _clear_data_rows(ws) -> None:
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def _resize_sheet_tables(ws, last_row: int, last_column: int) -> None:
    """Expand existing Excel tables while keeping the user's table style."""
    end_row = max(2, last_row)
    end_column = get_column_letter(last_column)
    for table in ws.tables.values():
        table.ref = f"A1:{end_column}{end_row}"
    ws.auto_filter.ref = f"A1:{end_column}{end_row}"
    if not ws.freeze_panes:
        ws.freeze_panes = "A2"


def _summary_row_values(headers: list[str], values: dict[str, Any]) -> list[Any]:
    return [values.get(_header_key(header), "") for header in headers]


def _write_summary_data(
    ws,
    submissions: Iterable[Any],
    headers: list[str],
    *,
    dealer_aggregates: dict[str, dict[str, Any]] | None = None,
) -> tuple[int, int]:
    """Write one row per Dealer + Product using the uploaded template columns.

    Dealer-level outlet totals use every real outlet submission. Product-level
    WS/DS/WM/TL/LE/CB/MS counts show how many outlets of each type sell the
    product. Movement is calculated from GENERAL outlet types only so it is
    exactly the same final normalized value shown by /report and /summary.
    The exporter never replaces or reorders the user's row-1 headers.
    """
    groups: dict[tuple[str, str], list[Any]] = defaultdict(list)
    for submission in submissions:
        if _is_summary_submission(submission):
            continue
        key = (
            _clean(getattr(submission, "region", None)).upper(),
            _clean(getattr(submission, "dealer", None)).upper(),
        )
        groups[key].append(submission)

    if dealer_aggregates is None:
        dealer_aggregates, _cache_hit = build_bulk_dealer_aggregates(list(submissions))

    output_rows: list[list[Any]] = []
    for (region, dealer), rows in sorted(groups.items(), key=_group_sort_key):
        # One cached aggregate already contains all-outlet availability and
        # GENERAL-only final movement. No second dealer aggregation is needed.
        aggregate = (dealer_aggregates or {}).get(dealer, {})
        outlet_counts = aggregate.get("outlet_types") or {}

        combined_location = (
            _joined_locations(rows)
            or _clean(aggregate.get("location_text"))
        )

        dealer_values = {
            "region": region,
            "dealer": dealer,
            # Exact approved V57 header plus compatibility with older custom
            # templates that used Location_Visit or Location Visit Text.
            "locationofvisittext": combined_location,
            "locationvisittext": combined_location,
            "locationvisit": combined_location,
            "member": most_frequent_member(rows),
            "group": _joined_unique(rows, "group_no"),
            "totaloutlets": len(rows),
            "wholesale": _count_type(outlet_counts, OUTLET_TYPE_KEYS["wholesale"]),
            "drinkshop": _count_type(outlet_counts, OUTLET_TYPE_KEYS["drinkshop"]),
            "wetmarket": _count_type(outlet_counts, OUTLET_TYPE_KEYS["wetmarket"]),
            "trolley": _count_type(outlet_counts, OUTLET_TYPE_KEYS["trolley"]),
            "localeat": _count_type(outlet_counts, OUTLET_TYPE_KEYS["localeat"]),
            "coffebakery": _count_type(outlet_counts, OUTLET_TYPE_KEYS["coffeebakery"]),
            "coffeebakery": _count_type(outlet_counts, OUTLET_TYPE_KEYS["coffeebakery"]),
            "canteen": _count_type(outlet_counts, OUTLET_TYPE_KEYS["canteen"]),
            "sportclub": _count_type(outlet_counts, OUTLET_TYPE_KEYS["sportclub"]),
            "motorshop": _count_type(outlet_counts, OUTLET_TYPE_KEYS["motorshop"]),
        }

        for product in EXPORT_PRODUCTS:
            availability = _availability_for_product(aggregate, product)
            product_values = dict(dealer_values)
            product_values.update(
                {
                    "product": PRODUCT_DISPLAY_ALIASES.get(product, product),
                    "movement": _movement_for_product(aggregate, product),
                    "mov": _movement_for_product(aggregate, product),
                    # Optional compatibility when an older/custom template has
                    # product-specific outlet columns.
                    "ws": _count_type(availability, OUTLET_TYPE_KEYS["wholesale"]),
                    "ds": _count_type(availability, OUTLET_TYPE_KEYS["drinkshop"]),
                    "wm": _count_type(availability, OUTLET_TYPE_KEYS["wetmarket"]),
                    "tl": _count_type(availability, OUTLET_TYPE_KEYS["trolley"]),
                    "le": _count_type(availability, OUTLET_TYPE_KEYS["localeat"]),
                    "cb": _count_type(availability, OUTLET_TYPE_KEYS["coffeebakery"]),
                    "ms": _count_type(availability, OUTLET_TYPE_KEYS["motorshop"]),
                }
            )
            output_rows.append(_summary_row_values(headers, product_values))

    for row_number, values in enumerate(output_rows, start=2):
        for column_number, value in enumerate(values, start=1):
            cell = ws.cell(row_number, column_number)
            cell.value = value
            header_key = _header_key(headers[column_number - 1])
            if header_key in {
                "totaloutlets", "wholesale", "drinkshop", "wetmarket", "trolley",
                "localeat", "coffebakery", "coffeebakery", "canteen",
                "sportclub", "motorshop", "movement", "mov", "ws", "ds", "wm", "tl",
                "le", "cb", "ms",
            }:
                cell.number_format = "0"
            elif header_key == "member":
                cell.number_format = "@"
            elif header_key in {
                "locationofvisittext",
                "locationvisittext",
                "locationvisit",
            }:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Keep the full combined location readable instead of visually cutting it.
    for column_number, header in enumerate(headers, start=1):
        if _header_key(header) in {
            "locationofvisittext",
            "locationvisittext",
            "locationvisit",
        }:
            column_letter = get_column_letter(column_number)
            current_width = ws.column_dimensions[column_letter].width or 0
            ws.column_dimensions[column_letter].width = max(current_width, 48)

    last_row = len(output_rows) + 1
    _resize_sheet_tables(ws, last_row, len(headers))
    return len(groups), len(output_rows)


def _location_row_values(headers: list[str], submission: Any, report_date: date) -> list[Any]:
    lat, lon = _coordinates(submission)
    values = {
        "date": getattr(submission, "report_date", None) or report_date,
        "region": _clean(getattr(submission, "region", None)).upper(),
        "dealer": _clean(getattr(submission, "dealer", None)).upper(),
        "latitude": lat,
        "longitude": lon,
        "outletname": _clean(getattr(submission, "outlet_name", None)),
        "outlettype": _clean(getattr(submission, "outlet_type", None)),
        "phonenumberoutlet": _clean(getattr(submission, "phone_number", None)),
    }
    return [values.get(_header_key(header), "") for header in headers]


def _write_location_data(
    ws,
    submissions: Iterable[Any],
    report_date: date,
    headers: list[str],
) -> int:
    rows = [s for s in submissions if not _is_summary_submission(s)]
    rows.sort(key=_location_sort_key)

    for row_number, submission in enumerate(rows, start=2):
        values = _location_row_values(headers, submission, report_date)
        for column_number, value in enumerate(values, start=1):
            cell = ws.cell(row_number, column_number)
            cell.value = value
            header_key = _header_key(headers[column_number - 1])
            if header_key == "date":
                cell.number_format = "dd/mm/yyyy"
            elif header_key in {"latitude", "longitude"}:
                cell.number_format = "0.0000000"
            elif header_key == "phonenumberoutlet":
                cell.number_format = "@"

    last_row = len(rows) + 1
    _resize_sheet_tables(ws, last_row, len(headers))
    return len(rows)


def create_data_export(
    submissions: Iterable[Any],
    report_date: date,
    template_path: Path | str | None = None,
    output_path: Path | str | None = None,
    *,
    dealer_aggregates: dict[str, dict[str, Any]] | None = None,
) -> tuple[Path, dict[str, int]]:
    template = Path(template_path or DATA_EXPORT_TEMPLATE)
    if not template.exists():
        raise FileNotFoundError(f"Data export template not found: {template}")

    workbook = load_workbook(template)
    missing_sheets = [
        name for name in (SUMMARY_SHEET, LOCATION_SHEET)
        if name not in workbook.sheetnames
    ]
    if missing_sheets:
        raise ValueError(
            "Data export template is missing sheet(s): " + ", ".join(missing_sheets)
        )

    summary_ws = workbook[SUMMARY_SHEET]
    location_ws = workbook[LOCATION_SHEET]
    summary_headers = _template_headers(summary_ws)
    location_headers = _template_headers(location_ws)
    _validate_headers(SUMMARY_SHEET, summary_headers, SUMMARY_REQUIRED_HEADER_KEYS)
    _validate_headers(LOCATION_SHEET, location_headers, LOCATION_REQUIRED_HEADER_KEYS)

    # Keep the user's row-1 labels, order, widths, colors and table styles.
    _clear_data_rows(summary_ws)
    _clear_data_rows(location_ws)

    submission_list = list(submissions or [])
    if dealer_aggregates is None:
        dealer_aggregates, _cache_hit = build_bulk_dealer_aggregates(submission_list)

    dealer_groups, summary_rows = _write_summary_data(
        summary_ws,
        submission_list,
        summary_headers,
        dealer_aggregates=dealer_aggregates,
    )
    location_rows = _write_location_data(
        location_ws,
        submission_list,
        report_date,
        location_headers,
    )

    destination = Path(output_path) if output_path else (
        settings.export_path / f"Market_Survey_Data_{report_date:%Y-%m-%d}.xlsx"
    )
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)

    return destination, {
        "dealer_groups": dealer_groups,
        # Retained for compatibility with older callers.
        "member_groups": dealer_groups,
        "summary_rows": summary_rows,
        "location_rows": location_rows,
        "products": len(EXPORT_PRODUCTS),
    }
