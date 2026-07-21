# app/reports/excel_report.py

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from copy import copy

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

from app.core.config import settings
from app.data.dealers import ALL_DEALERS
from app.reports.aggregator import OWN_PRODUCTS, COMPETITOR_PRODUCTS, RING_PRODUCTS

# Exact cell layout from template_by_dealer.xlsx.
# General Trade report uses 4 outlet-type columns.
GENERAL_OUTLET_COLS = {"Wholesale": 8, "Drink Shop": 12, "Wet Market": 16, "Trolley": 20}
# Channel Specialist report uses 5 outlet-type columns.
CHANNEL_SPECIALIST_OUTLET_COLS = {
    "Local Eat": 8,
    "Coffee,Bakery": 12,
    "Canteen": 16,
    "Sport Club": 20,
    "Motor Shop": 23,
}
CHANNEL_SPECIALIST_TYPES = set(CHANNEL_SPECIALIST_OUTLET_COLS)
OUTLET_COLS = GENERAL_OUTLET_COLS
FRESHNESS_START_ROW = 7
FRESHNESS_END_ROW = 24
MOVEMENT_HEADER_ROW = 26
MOVEMENT_START_ROW = 27
MOVEMENT_END_ROW = 42
RING_PULL_START_ROW = 45
ISSUE_START_ROW = 45
SUGGESTION_START_ROW = 45

# Report logo rendering. Increase these values if the logo still appears small in the PNG preview.
REPORT_LOGO_WIDTH_PX = 110
REPORT_LOGO_HEIGHT_PX = 110
REPORT_LOGO_ANCHOR = "A1"

# Summary text display settings.
# Do NOT manually insert line breaks. Excel/LibreOffice should use the full
# merged summary cell width first, then wrap naturally only when the text reaches
# the cell border. Row height is estimated only to make the PNG preview clean.
SUMMARY_CHARS_PER_LINE = 145

# Keep each numbered item (1-4) on a clearly separated visual line.
# 32 points is about 43 screen pixels, matching the requested preview spacing.
SUMMARY_MIN_ROW_HEIGHT = 32
SUMMARY_LINE_HEIGHT = 22
SUMMARY_MAX_ROW_HEIGHT = 140

# Noto Sans Khmer is installed by the Railway Dockerfile. It prevents Khmer
# glyphs and diacritics from colliding when LibreOffice renders Excel to PNG.
SUMMARY_FONT_NAME = "Noto Sans Khmer"
SUMMARY_FONT_SIZE = 17

# Template label differences -> aggregation product names.
PRODUCT_NAME_MAP = {
    "CBC LITE ORD": "CB LITE ORD",
    "CB LITE ORD": "CB LITE ORD",
    "CBC 4.4": "CBC 4.4 NCP",
    "CBC 4.4 NCP": "CBC 4.4 NCP",
    "CB Original": "CB Original NCP",
    "CB Original NCP": "CB Original NCP",
    "CB LITE": "CB LITE NCP",
    "CB LITE NCP": "CB LITE NCP",
    "CB BLACK": "CB BLACK NCP",
    "CB BLACK NCP": "CB BLACK NCP",
    "CAMBODIA COLA 330ml": "CAMBODIA COLA",
    "CAMBODIA COLA 330ML": "CAMBODIA COLA",
    "CAMBODIA ED": "CAMBODIA ED",
    "ភេសជ្ជៈប៉ូវកម្លាំង​កម្ពុជា": "CAMBODIA ED",
    "IZE PET 300ml All SKUs": "IZE PET 300ml Flavour",
    "IZE PET 300ML ALL SKUS": "IZE PET 300ml Flavour",
    "EXPREZ ត្រសក់ផ្អែម": "EXPREZ Melon",
    "EXPREZ Can 330ml": "EXPREZ Can 330ml",
    "EXPREZ Can 330mL": "EXPREZ Can 330ml",
    "CAMBODIA Sport 500ml": "CAMBODIA Sport 500mL",
    "CAMBODIA Sport 500ML": "CAMBODIA Sport 500mL",
    "CAMBODIA Sport 300mL": "CAMBODIA Sport 300mL",
    "CAMBODIA Sport 300ml": "CAMBODIA Sport 300ml",
    "CAMBODIA Sport 300ML": "CAMBODIA Sport 300ml",
    "GB Original": "GB Original NCP",
    "GB  Original": "GB Original NCP",
    "GB Original NCP": "GB Original NCP",
    "GB SNOW": "GB SNOW NCP",
    "Hanuman Lite": "Hanuman LITE NCP",
    "Krud": "Krud NCP",
    "Krud Lite": "Krud LITE NCP",
    "Greet Lite": "Greet LITE NCP",
    "Great Lite": "Greet LITE NCP",
    "Hanuman Black": "Hanuman Black NCP",
    "Ganzberg  500ml": "Ganzberg 500ml",
    "CBC, CBL Can and CBB Can": "CBL NCP 6 Can",
    "Wurkz NCP 5 USD": "CBL NCP 5 USD",
}


def _clean(v) -> str:
    return " ".join(str(v or "").split()).strip()


def _product_key(name: str) -> str:
    s = _clean(name)
    return PRODUCT_NAME_MAP.get(s) or PRODUCT_NAME_MAP.get(s.upper()) or s



def _norm_lookup_key(name: str) -> str:
    """Normalize product names for safe lookup across template/code aliases.

    Fixes cases like:
    - GB Original vs GB  Original
    - non-breaking spaces / hidden whitespace
    - upper/lower case differences
    """
    text = _product_key(name)
    return "".join(ch.lower() for ch in str(text) if ch.isalnum())


def _metric_value(metrics: dict | None, *names: str):
    """Return the first non-empty metric value from possible field names."""
    if not isinstance(metrics, dict):
        return None
    for name in names:
        value = metrics.get(name)
        if value not in (None, "", "nan"):
            return value
    return None


def _numeric_mov(value) -> float:
    try:
        if value in (None, "", "nan"):
            return -1
        return float(str(value).replace(",", "").strip())
    except Exception:
        return -1


def _metric_mov_score(metrics: dict | None) -> float:
    """Movement value used when choosing between alias candidates."""
    return _numeric_mov(
        _metric_value(
            metrics,
            "mov",
            "movement",
            "movement_score",
            "final_mov",
            "final_movement",
            "offtake_movement",
        )
    )




def _final_movement_from_metrics(metrics: dict | None):
    """Return the final movement calculated by the aggregator.

    The comparison-row rule is applied in ``aggregator.py`` after raw averages
    are calculated. Excel must trust that final ``mov`` value and must not
    independently promote rounded 8/9 values to 10, otherwise the comparison
    ranking and the one-10-per-row rule would be overwritten.
    """
    if not isinstance(metrics, dict):
        return None

    final_value = _metric_value(
        metrics,
        "mov",
        "final_mov",
        "final_movement",
        "movement",
        "movement_score",
    )
    if final_value not in (None, "", "nan"):
        return final_value

    # Backward-compatible fallback for old metric dictionaries that contain
    # only raw movement values and no aggregated final movement.
    raw_values = metrics.get("_movement_values") or metrics.get("movement_values")
    if isinstance(raw_values, (list, tuple)):
        nums = []
        for value in raw_values:
            try:
                if value not in (None, "", "nan"):
                    nums.append(float(str(value).replace(",", "").strip()))
            except Exception:
                pass
        if nums:
            rounded = int((sum(nums) / len(nums)) + 0.5)
            return max(0, min(10, rounded))

    return None

def _lookup_metrics(agg: dict, template_name: str) -> dict | None:
    """Return product/competitor metrics using robust alias matching.

    The old report could show GB Original = 2 because Excel lookup was matching
    a stale/alias entry instead of the final aggregated GB Original value.
    This function first tries exact canonical names, then normalized aliases, and
    if duplicate aliases exist it prefers the candidate with the highest movement
    value. That makes renamed products like GB Original stable after template/form
    updates.
    """
    canonical = _product_key(template_name)
    buckets = [agg.get("competitors") or {}, agg.get("products") or {}]

    # Exact canonical lookup first.
    exact_candidates = []
    for bucket in buckets:
        val = bucket.get(canonical)
        if isinstance(val, dict):
            exact_candidates.append(val)
    if len(exact_candidates) == 1:
        return exact_candidates[0]
    if len(exact_candidates) > 1:
        return max(exact_candidates, key=lambda d: _numeric_mov(d.get("mov")))

    # Normalized alias lookup.
    target = _norm_lookup_key(canonical)
    candidates = []
    for bucket in buckets:
        for key, val in bucket.items():
            if _norm_lookup_key(key) == target and isinstance(val, dict):
                candidates.append(val)

    if not candidates:
        return None
    return max(candidates, key=lambda d: _numeric_mov(d.get("mov")))


def _lookup_competitor_metrics(agg: dict, template_name: str) -> dict | None:
    """Lookup competitor metrics using strict + fuzzy aliases.

    This function never returns the first exact match immediately. It collects
    all possible matching keys, including hidden spaces and old aliases, then
    returns the candidate with the strongest final movement.

    Supported examples:
      - GB Original
      - GB  Original
      - GBOriginal
      - gb_original
      - comp_mov_gb_original
    """
    canonical = _product_key(template_name)
    competitors = agg.get("competitors") or {}
    products = agg.get("products") or {}

    # Comparison columns may contain cross-over own products. V46 adds
    # EXPREZ Can 330ml to that group, so search both result buckets.
    buckets = [competitors, products]
    if not competitors and not products:
        return None

    target = _norm_lookup_key(canonical)
    raw_template = str(template_name or "")
    clean_template = " ".join(raw_template.split())

    alias_map = {
        "gboriginal": {
            "GB Original",
            "GB  Original",
            "GBOriginal",
            "gb_original",
            "comp_mov_gb_original",
            "comp_gb_original_mov",
            "gb_original_movement",
            "gb_original_movement_score",
            "gb_original_movement_score_0_10",
        },
        "gbsnow": {"GB SNOW", "GB Snow", "gb_snow", "gb_snow_movement", "gb_snow_movement_score"},
        "greetlite": {"Greet Lite", "Great Lite", "greet_lite", "great_lite"},
        "greatlite": {"Greet Lite", "Great Lite", "greet_lite", "great_lite"},
    }

    candidate_names = {
        canonical,
        clean_template,
        raw_template,
        " ".join(str(canonical).split()),
    }
    candidate_names.update(alias_map.get(target, set()))

    candidates: list[dict] = []

    # 1) Direct alias matches across competitor and cross-over own products.
    for bucket in buckets:
        for key in candidate_names:
            val = bucket.get(key)
            if isinstance(val, dict):
                candidates.append(val)

    # 2) Normalized full-name matches.
    for bucket in buckets:
        for key, val in bucket.items():
            if not isinstance(val, dict):
                continue
            if _norm_lookup_key(key) == target:
                candidates.append(val)

    # 3) Fuzzy contains match for renamed/legacy GB Original keys.
    if target in {"gboriginal", "gboriginalncp"}:
        for bucket in buckets:
            for key, val in bucket.items():
                if not isinstance(val, dict):
                    continue
                nk = _norm_lookup_key(key)
                if "gb" in nk and "original" in nk:
                    candidates.append(val)

    if not candidates:
        return None

    # Remove duplicate dict objects.
    unique: list[dict] = []
    seen: set[int] = set()
    for item in candidates:
        ident = id(item)
        if ident not in seen:
            seen.add(ident)
            unique.append(item)

    best = max(unique, key=_metric_mov_score)

    if target in {"gboriginal", "gboriginalncp"}:
        print(
            "✅ Excel GB Original lookup candidates:",
            [_metric_value(c, "mov", "movement_score", "final_mov", "final_movement") for c in unique],
            "selected:",
            _metric_value(best, "mov", "movement_score", "final_mov", "final_movement"),
        )

    return best


def final_report_movement_value(
    agg: dict,
    product_name: str,
    *,
    competitor: bool = False,
) -> int | None:
    """Return the exact movement value that the final Excel report writes.

    Summary and export features must use this helper instead of reading raw
    aggregate buckets directly. The final report has robust alias handling for
    renamed competitor products, so sharing the same lookup guarantees that a
    dealer summary cannot name Greet LITE NCP as the movement leader when the
    final report actually shows Hanuman LITE NCP = 10.
    """
    if competitor:
        metrics = _lookup_competitor_metrics(agg, product_name)
    else:
        canonical = _product_key(product_name)
        metrics = (agg.get("products") or {}).get(canonical)
        if not isinstance(metrics, dict):
            metrics = _lookup_metrics(agg, product_name)

    value = _final_movement_from_metrics(metrics)
    if value in (None, "", "nan"):
        return None
    try:
        return max(0, min(10, int(float(str(value).replace(",", "").strip()))))
    except Exception:
        return None


def _write_metrics_cells(ws: Worksheet, row: int, product_col: int, metrics: dict | None) -> None:
    """Write movement/stock/buy/sell cells safely."""
    if not metrics:
        for c in range(product_col + 1, product_col + 5):
            ws.cell(row, c).value = ""
        return

    ws.cell(row, product_col + 1).value = _blank_if_none(_final_movement_from_metrics(metrics))
    ws.cell(row, product_col + 2).value = _blank_if_none(_metric_value(metrics, "stock", "stock_status"))
    _set_number_cell(ws, row, product_col + 3, _metric_value(metrics, "buy_in", "buy_in_price"))
    _set_number_cell(ws, row, product_col + 4, _metric_value(metrics, "sell_out", "sell_out_price"))


def _force_rewrite_competitor_blocks(ws: Worksheet, agg: dict) -> None:
    """Final safety pass for competitor movement blocks.

    This fixes cases where a copied template or earlier lookup left an old value
    such as GB Original = 2. The final pass scans the product-name cells and
    rewrites the metric cells from the aggregated result dictionary.
    """
    for row in range(MOVEMENT_START_ROW, MOVEMENT_END_ROW + 1):
        for product_col in [8, 13, 18, 23]:
            product_label = ws.cell(row, product_col).value
            if not product_label:
                continue
            comp_name = _product_key(product_label)
            # Always clear old/stale template values first, then rewrite.
            for c in range(product_col + 1, product_col + 5):
                ws.cell(row, c).value = ""
            metrics = _lookup_competitor_metrics(agg, comp_name)
            if metrics:
                _write_metrics_cells(ws, row, product_col, metrics)




def _force_specific_competitor_value(ws: Worksheet, agg: dict, product_name: str) -> None:
    """Force a competitor value into every matching template label cell.

    This is intentionally stronger than the normal block writer. It scans the
    whole movement area for the product label and writes the final movement into
    the next cell, so copied template values cannot survive.
    """
    metrics = _lookup_competitor_metrics(agg, product_name)
    final_mov = _final_movement_from_metrics(metrics)
    if final_mov in (None, ""):
        return

    target = _norm_lookup_key(product_name)
    for row in range(MOVEMENT_START_ROW, MOVEMENT_END_ROW + 1):
        for col in range(1, 28):
            label = ws.cell(row, col).value
            if _norm_lookup_key(label) == target:
                for c in range(col + 1, min(col + 5, 28)):
                    ws.cell(row, c).value = ""
                ws.cell(row, col + 1).value = _blank_if_none(final_mov)
                if isinstance(metrics, dict):
                    ws.cell(row, col + 2).value = _blank_if_none(_metric_value(metrics, "stock", "stock_status"))
                    _set_number_cell(ws, row, col + 3, _metric_value(metrics, "buy_in", "buy_in_price"))
                    _set_number_cell(ws, row, col + 4, _metric_value(metrics, "sell_out", "sell_out_price"))
                print(f"✅ FINAL FORCE {product_name} Excel {ws.cell(row, col + 1).coordinate} = {final_mov}")


def _safe_title(title: str, used: set[str]) -> str:
    base = "".join(ch for ch in title if ch not in "[]:*?/\\")[:31] or "Sheet"
    candidate = base
    i = 1
    while candidate in used:
        suffix = f"_{i}"
        candidate = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(candidate)
    return candidate


def _blank_if_none(value):
    return "" if value is None else value

def _is_channel_specialist_report(agg: dict) -> bool:
    """Detect Channel Specialist report by outlet-type data or explicit channel field.

    This keeps one template but changes the top outlet-type headers dynamically.
    """
    channel = str(agg.get("channel") or agg.get("channel_type") or "").strip().lower()
    if "channel" in channel or "specialist" in channel:
        return True
    outlet_types = agg.get("outlet_types") or {}
    return any(k in CHANNEL_SPECIALIST_TYPES for k in outlet_types.keys())


def _outlet_cols_for_report(agg: dict) -> dict[str, int]:
    return CHANNEL_SPECIALIST_OUTLET_COLS if _is_channel_specialist_report(agg) else GENERAL_OUTLET_COLS


def _new_purchase_col_for_report(agg: dict) -> int:
    # General: W. Channel Specialist: Z because W is used by Motor Shop.
    return 26 if _is_channel_specialist_report(agg) else 23


def _volume_col_for_report(agg: dict) -> int:
    # General: Z. Channel Specialist: AA.
    return 27 if _is_channel_specialist_report(agg) else 26




def _copy_cell_style(src, dst) -> None:
    """Copy style pieces safely between openpyxl cells."""
    if src is None or dst is None:
        return
    try:
        dst._style = copy(src._style)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
    except Exception:
        pass


def _unmerge_if_exists(ws: Worksheet, range_text: str) -> None:
    for rng in list(ws.merged_cells.ranges):
        if str(rng) == range_text:
            ws.unmerge_cells(range_text)
            return


def _prepare_channel_specialist_layout(ws: Worksheet, agg: dict) -> None:
    """Prepare Section 1 columns for Channel Specialist.

    General template:
      H:K Wholesale | L:O Drink Shop | P:S Wet Market | T:V Trolley
      W:Y New Outlet Purchase | Z:AA Volume

    Channel Specialist template:
      H:K Local Eat | L:O Coffee,Bakery | P:S Canteen | T:V Sport Club
      W:Y Motor Shop | Z New Outlet Purchase | AA Volume

    The critical fix is to split Z:AA and copy formatting/borders to AA so
    "Volume" stays inside the table instead of appearing outside the border.
    """
    if not _is_channel_specialist_report(agg):
        return

    for r in range(6, FRESHNESS_END_ROW + 1):
        # Keep W:Y merged for Motor Shop, but split Z:AA for New Outlet + Volume.
        _unmerge_if_exists(ws, f"Z{r}:AA{r}")

        # After unmerge, AA may be a plain cell with no borders/style. Copy from Z.
        _copy_cell_style(ws.cell(r, 26), ws.cell(r, 27))

        # Make sure both cells have centered text and visible borders.
        ws.cell(r, 26).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(r, 27).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.column_dimensions["Z"].width = max(ws.column_dimensions["Z"].width or 10, 20)
    ws.column_dimensions["AA"].width = max(ws.column_dimensions["AA"].width or 10, 14)

def _set_outlet_type_headers(ws: Worksheet, agg: dict) -> None:
    """Write the correct header row for General Trade or Channel Specialist.

    Preserves the existing template size/layout; only the labels and target columns change.
    """
    # Clear all possible outlet/new-purchase/volume header cells first.
    for col in [8, 12, 16, 20, 23, 26, 27]:
        ws.cell(6, col).value = ""

    outlet_types = agg.get("outlet_types") or {}
    for label, col in _outlet_cols_for_report(agg).items():
        ws.cell(6, col).value = f"{label}: {int(outlet_types.get(label, 0) or 0)}"

    new_col = _new_purchase_col_for_report(agg)
    vol_col = _volume_col_for_report(agg)
    ws.cell(6, new_col).value = "New Outlet Purchase"
    ws.cell(6, vol_col).value = "Volume"
    ws.cell(6, new_col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(6, vol_col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)



def _set_movement_headers(ws: Worksheet) -> None:
    """Restore Product Movement VS Competitor column titles.

    Some uploaded templates or previous report logic can lose the blue header
    row because data was written starting on the header row. This function
    always rewrites row 25 before data is filled, for General and Channel
    Specialist reports.
    """
    headers = {
        1: "#",
        2: "Product",
        3: "Mov",
        4: "Stock",
        5: "Buy In",
        6: "Sell Out",
        7: "Ring Pull",
        8: "Product",
        9: "Mov",
        10: "Stock",
        11: "Buy In",
        12: "Sell Out",
        13: "Product",
        14: "Mov",
        15: "Stock",
        16: "Buy In",
        17: "Sell Out",
        18: "Product",
        19: "Mov",
        20: "Stock",
        21: "Buy In",
        22: "Sell Out",
        23: "Product",
        24: "Mov",
        25: "Stock",
        26: "Buy In",
        27: "Sell Out",
    }

    header_fill = PatternFill("solid", fgColor="003B6F")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[MOVEMENT_HEADER_ROW].height = 18
    for col, title in headers.items():
        cell = ws.cell(MOVEMENT_HEADER_ROW, col)
        cell.value = title
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _excel_number(value):
    """Write clean numbers to Excel: 52000.0 -> 52000, 6.25 -> 6.25."""
    if value in (None, "", "nan"):
        return ""
    try:
        f = float(str(value).replace(",", "").strip())
        return int(f) if f.is_integer() else round(f, 2)
    except Exception:
        return value


def _set_number_cell(ws: Worksheet, row: int, col: int, value) -> None:
    cleaned = _excel_number(value)
    cell = ws.cell(row, col)
    cell.value = cleaned
    if isinstance(cleaned, int):
        cell.number_format = "0"
    elif isinstance(cleaned, float):
        cell.number_format = "0.##"


def _find_report_logo() -> Path | None:
    """Return the first available KB logo path.

    openpyxl does not copy images when copy_worksheet() is used, so generated
    sheets must insert the logo from an image file every time. The project now
    ships the logo in templates/kb_logo.png; fallbacks keep older folders working.
    """
    candidates = [
        settings.template_file.parent / "kb_logo.png",
        settings.template_file.parent.parent / "media" / "kb_logo.png",
        settings.template_file.parent.parent / "app" / "Logo" / "khmer_beverages_co_ltd_logo.jpg",
    ]
    for path in candidates:
        if path.exists():
            return path
    return None


def _add_report_logo(ws: Worksheet) -> None:
    """Add a visible Khmer Beverages logo to generated report sheets."""
    logo_path = _find_report_logo()
    if not logo_path:
        print("⚠️ KB logo not found. Expected templates/kb_logo.png")
        return

    try:
        # copy_worksheet does not reliably copy images. Clearing prevents
        # duplicate/old small logos, then we insert one clean logo.
        ws._images = []

        img = XLImage(str(logo_path))
        img.width = REPORT_LOGO_WIDTH_PX
        img.height = REPORT_LOGO_HEIGHT_PX
        ws.add_image(img, REPORT_LOGO_ANCHOR)

        # Make enough room so the logo is visible in Excel and Telegram PNG.
        ws.row_dimensions[1].height = max(ws.row_dimensions[1].height or 15, 82)
        ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 7, 15)
    except Exception as exc:
        print(f"⚠️ Could not add KB logo to report sheet: {exc}")


def _summary_visual_length(text: str) -> int:
    """Approximate rendered width for Khmer/English mixed text.

    We only use this to estimate row height. We never insert manual line breaks,
    because manual wrapping can split Khmer/English words badly in the PNG.
    """
    total = 0
    for ch in str(text or ""):
        if ch == "\n":
            # Treat manual user/AI newline as a full line break marker.
            total += SUMMARY_CHARS_PER_LINE
        elif ord(ch) > 127:
            # Khmer/CJK characters are visually wider than Latin letters.
            total += 2
        else:
            total += 1
    return total


def _estimate_summary_lines(text: str) -> int:
    """Estimate how many display lines the summary cell needs."""
    text = (text or "").strip()
    if not text:
        return 1

    lines = 0
    for part in text.splitlines() or [text]:
        part = part.strip()
        if not part:
            lines += 1
        else:
            visual_len = _summary_visual_length(part)
            lines += max(1, (visual_len + SUMMARY_CHARS_PER_LINE - 1) // SUMMARY_CHARS_PER_LINE)
    return max(1, lines)


def _clean_summary_text(text: str) -> str:
    """Clean summary text without forcing artificial line breaks."""
    text = (text or "").strip()
    # Keep explicit user/AI paragraph breaks, but remove excessive spacing.
    cleaned_lines: list[str] = []
    for line in text.replace("\r", "\n").splitlines():
        line = " ".join(line.split())
        if line:
            cleaned_lines.append(line)
    return "\n".join(cleaned_lines)


def _set_row_text(ws: Worksheet, row: int, col: int, prefix_no: int, text: str) -> int:
    """Write numbered Key Issue/Suggestion text professionally.

    Important:
    - No textwrap.wrap()
    - No manual line breaking
    - Excel/LibreOffice uses full cell width first
    - Long text wraps naturally in the generated Excel/PNG
    """
    cleaned = _clean_summary_text(text)
    value = f"{prefix_no}. {cleaned}" if cleaned else f"{prefix_no}."

    cell = ws.cell(row, col)
    cell.value = value

    # Use a Khmer-safe font and center the text inside a taller row.
    # The four summary items remain on rows 45-48; only their visual spacing
    # changes, so no template coordinates or business calculations are affected.
    cell.font = Font(
        name=SUMMARY_FONT_NAME,
        size=SUMMARY_FONT_SIZE,
        bold=False,
        color="000000",
    )
    cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
        shrink_to_fit=False,
    )

    return _estimate_summary_lines(value)


def _fit_summary_row_height(ws: Worksheet, row: int, issue_lines: int, suggestion_lines: int) -> None:
    """Give each of summary rows 1-4 clear vertical breathing room.

    Empty rows also keep the same minimum height, so 1, 2, 3 and 4 remain
    evenly separated in Excel and in Railway's LibreOffice PNG output.
    """
    max_lines = max(issue_lines, suggestion_lines, 1)
    height = SUMMARY_MIN_ROW_HEIGHT + ((max_lines - 1) * SUMMARY_LINE_HEIGHT)
    ws.row_dimensions[row].height = min(SUMMARY_MAX_ROW_HEIGHT, height)

def fill_template_sheet(ws: Worksheet, agg: dict) -> None:
    dealer = agg.get("dealer") or ""
    total = int(agg.get("total_outlets") or 0)
    rdate = agg.get("report_date")
    if isinstance(rdate, (date, datetime)):
        rdate_txt = rdate.strftime("%d/%m/%Y")
    else:
        rdate_txt = str(rdate or "")

    _add_report_logo(ws)
    _prepare_channel_specialist_layout(ws, agg)
    # New template layout: Dealer and Report Date are on the same row/cell.
    # Example: "Dealer : CA1                              Report Date: 02/07/2026 14:55:10"
    ws["A1"] = ""
    ws["B1"] = ""
    report_date_time = f"{rdate_txt} {datetime.now().strftime('%H:%M:%S')}"
    if _is_channel_specialist_report(agg):
        ws["A3"] = f"Dealer : {dealer}    CHANNEL SPECIALIST                         Report Date: {report_date_time}"
    else:
        ws["A3"] = f"Dealer : {dealer}                              Report Date: {report_date_time}"
    ws["A3"].font = ws["A3"].font.copy(bold=True)
    ws["A4"] = f"Group : {agg.get('group_no') or 2}"
    member_no = agg.get("member_no") or (max(1, min(10, total // 3 or 1)) if total else 0)
    ws["C4"] = f"Member : {member_no}"
    location_text = str(agg.get("location_text") or "").strip()
    ws["H4"] = f"Location of Visit: {location_text}"
    # H4:V4 is a merged range in the report template. Keep the complete
    # combined location text and allow it to wrap instead of being visually
    # cut off in Excel/PDF/PNG. Noto Sans Khmer preserves Khmer shaping.
    ws["H4"].font = Font(name=SUMMARY_FONT_NAME, size=10, bold=False)
    ws["H4"].alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
        shrink_to_fit=False,
    )
    if len(location_text) > 150:
        ws.row_dimensions[4].height = max(ws.row_dimensions[4].height or 15, 45)
    elif len(location_text) > 80:
        ws.row_dimensions[4].height = max(ws.row_dimensions[4].height or 15, 32)

    # Total outlet visit cell in template is around W4, but search is safer.
    for row in ws.iter_rows(min_row=1, max_row=6):
        for cell in row:
            if isinstance(cell.value, str) and "Total outlet Visit" in cell.value:
                cell.value = f"Total outlet Visit : {total}"

    # Section 1: outlet type totals in header and availability counts per product.
    _set_outlet_type_headers(ws, agg)
    outlet_cols = _outlet_cols_for_report(agg)
    new_purchase_col = _new_purchase_col_for_report(agg)
    volume_col = _volume_col_for_report(agg)

    for row in range(FRESHNESS_START_ROW, FRESHNESS_END_ROW + 1):
        product_name = _product_key(ws.cell(row, 2).value)
        pdata = (agg.get("products") or {}).get(product_name, {})
        ws.cell(row, 3).value = _blank_if_none(pdata.get("bbe"))
        counts = pdata.get("availability") or {}

        # Clear all possible count/output columns first so copied template data cannot remain.
        for col in [8, 12, 16, 20, 23, 26, 27]:
            ws.cell(row, col).value = ""

        for label, col in outlet_cols.items():
            ws.cell(row, col).value = int(counts.get(label, 0) or 0)

        ws.cell(row, new_purchase_col).value = int(pdata.get("new_purchase", 0) or 0)
        vol = pdata.get("volume")
        ws.cell(row, volume_col).value = "" if vol in (None, 0, 0.0) else _excel_number(vol)

    # Section 2/3: product movement vs competitors.
    # Always restore the header row first. Product data starts on row 26.
    _set_movement_headers(ws)

    for row in range(MOVEMENT_START_ROW, MOVEMENT_END_ROW + 1):
        # Own products: Product, Mov, Stock, Buy In, Sell Out, Ring Pull = B:G
        own_name = _product_key(ws.cell(row, 2).value)
        own = (agg.get("products") or {}).get(own_name, {})
        ws.cell(row, 3).value = _blank_if_none(own.get("mov"))
        ws.cell(row, 4).value = _blank_if_none(own.get("stock"))
        _set_number_cell(ws, row, 5, own.get("buy_in"))
        _set_number_cell(ws, row, 6, own.get("sell_out"))
        _set_number_cell(ws, row, 7, own.get("ring_pull"))

        # Competitor blocks: H:L, M:Q, R:V, W:AA
        for product_col in [8, 13, 18, 23]:
            comp_name = _product_key(ws.cell(row, product_col).value)
            # Always clear old/stale template values first, then rewrite.
            for c in range(product_col + 1, product_col + 5):
                ws.cell(row, c).value = ""
            metrics = _lookup_competitor_metrics(agg, comp_name)
            if metrics:
                _write_metrics_cells(ws, row, product_col, metrics)
            else:
                # Preserve product names; clear stale metric cells for empty or unknown competitors.
                if ws.cell(row, product_col).value:
                    for c in range(product_col + 1, product_col + 5):
                        ws.cell(row, c).value = ""

    # Final safety pass: remove stale values from competitor blocks.
    _force_rewrite_competitor_blocks(ws, agg)
    _force_specific_competitor_value(ws, agg, "GB Original NCP")

    # Hard production guard for GB Original. If aggregator contains any GB Original
    # alias with final movement 10, force the Excel movement cell to that value.
    # This prevents copied template cells from keeping a stale value like 2.
    gb_metrics = _lookup_competitor_metrics(agg, "GB Original NCP")
    if gb_metrics:
        gb_mov = _final_movement_from_metrics(gb_metrics)
        for rr in range(MOVEMENT_START_ROW, MOVEMENT_END_ROW + 1):
            for cc in [8, 13, 18, 23]:
                if _norm_lookup_key(ws.cell(rr, cc).value) in {"gboriginal", "gboriginalncp"}:
                    ws.cell(rr, cc + 1).value = _blank_if_none(gb_mov)
                    print("✅ FORCE WRITE GB Original to Excel:", gb_mov, "cell", ws.cell(rr, cc + 1).coordinate)

    # Section 4: Ring Pull fixed products.
    for i, product in enumerate(RING_PRODUCTS, start=RING_PULL_START_ROW):
        rp = (agg.get("ring_pull") or {}).get(product, {})
        ws.cell(i, 2).value = product
        ws.cell(i, 4).value = int(rp.get("total_outlets", 0) or 0)
        ws.cell(i, 6).value = int(rp.get("qty", 0) or 0)

    # Bottom summary text. Only this area wraps long lines.
    key_issues = list((agg.get("key_issues") or [])[:4])
    suggestions = list((agg.get("suggestions") or [])[:4])
    while len(key_issues) < 4:
        key_issues.append("")
    while len(suggestions) < 4:
        suggestions.append("")

    for i in range(4):
        row = ISSUE_START_ROW + i
        issue_lines = _set_row_text(ws, row, 9, i + 1, key_issues[i])
        suggestion_lines = _set_row_text(ws, row, 19, i + 1, suggestions[i])
        _fit_summary_row_height(ws, row, issue_lines, suggestion_lines)

    # Page setup helps LibreOffice render a single clean preview image/PDF.
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = "A1:AA48"



def _template_for_aggs(aggs: list[dict]) -> Path:
    """Choose General or Channel Specialist template if split templates exist.

    Backward compatible: if split templates are missing, falls back to
    TEMPLATE_PATH from .env.
    """
    base = settings.template_file
    template_dir = base.parent
    if aggs and all(_is_channel_specialist_report(agg) for agg in aggs):
        candidate = template_dir / "template_channel_specialist.xlsx"
        if candidate.exists():
            return candidate

    candidate = template_dir / "template_general.xlsx"
    if candidate.exists():
        return candidate
    return base


def create_report_workbook(aggs: list[dict], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(_template_for_aggs(aggs))
    template = wb.worksheets[0]

    if not aggs:
        template.title = "No Data"
        wb.save(output_path)
        return output_path

    used = set(ws.title for ws in wb.worksheets)
    used.discard(template.title)
    created_sheets = []
    for agg in aggs:
        ws = wb.copy_worksheet(template)
        ws.title = _safe_title(str(agg.get("dealer") or "Dealer"), used)
        fill_template_sheet(ws, agg)
        created_sheets.append(ws)

    # Remove original template sheet after all copies are created.
    wb.remove(template)

    # Remove any extra sheets from uploaded template except generated sheets.
    for ws in list(wb.worksheets):
        if ws not in created_sheets:
            wb.remove(ws)

    wb.active = 0
    wb.save(output_path)
    return output_path


def create_single_report(agg: dict) -> Path:
    name = f"Market_Improvement_{agg['dealer']}_{agg.get('report_date')}.xlsx".replace(":", "-")
    return create_report_workbook([agg], settings.export_path / name)


def _blank_agg(dealer: str, report_date) -> dict:
    return {
        "dealer": dealer,
        "report_date": report_date,
        "total_outlets": 0,
        "outlet_types": {},
        "location_text": "",
        "products": {p: {"availability": {}} for p in OWN_PRODUCTS},
        "competitors": {p: {} for p in COMPETITOR_PRODUCTS},
        "ring_pull": {p: {"total_outlets": 0, "qty": 0} for p in RING_PRODUCTS},
        "key_issues": ["", "", "", ""],
        "suggestions": ["", "", "", ""],
    }


def create_all_dealer_report(aggs_by_dealer: dict[str, dict], report_date) -> Path:
    # User requested 65 sheets: one sheet per dealer, even if no submission.
    aggs = [aggs_by_dealer.get(dealer) or _blank_agg(dealer, report_date) for dealer in ALL_DEALERS]
    out = settings.export_path / f"Market_Improvement_All_Dealers_{report_date}.xlsx"
    return create_report_workbook(aggs, out)


def create_selected_dealer_report(
    aggs_by_dealer: dict[str, dict],
    dealers: list[str] | tuple[str, ...],
    report_date,
) -> Path:
    """Create one workbook containing only the requested dealer sheets."""
    ordered = [str(dealer).strip().upper() for dealer in dealers if str(dealer).strip()]
    aggs = [aggs_by_dealer.get(dealer) or _blank_agg(dealer, report_date) for dealer in ordered]
    dealer_part = "_".join(ordered)
    # Keep Windows paths manageable while preserving a useful file name.
    if len(dealer_part) > 90:
        dealer_part = f"{len(ordered)}_Dealers"
    out = settings.export_path / f"Market_Improvement_{dealer_part}_{report_date}.xlsx"
    return create_report_workbook(aggs, out)

