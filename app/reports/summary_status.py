from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.data.dealers import DEALER_REGION, ALL_DEALERS
from app.reports.aggregator import is_final_summary_outlet_name


STATUS_SUBMITTED = "បានធ្វើបូកសរុបរួម"
STATUS_MISSING = "មិនបានធ្វើបូកសរុបរួម"


def build_summary_status_rows(
    submissions: Iterable[Any], report_date: date
) -> list[dict[str, Any]]:
    """Build one status row per dealer using the report's shared matcher."""
    submitted = {
        str(getattr(row, "dealer", "") or "").strip().upper()
        for row in submissions
        if is_final_summary_outlet_name(getattr(row, "outlet_name", None))
    }
    return [
        {
            "date": report_date,
            "region": DEALER_REGION[dealer],
            "dealer": dealer,
            "status": STATUS_SUBMITTED if dealer in submitted else STATUS_MISSING,
        }
        for dealer in ALL_DEALERS
    ]


def create_summary_status_export(
    submissions: Iterable[Any], report_date: date, output_path: Path
) -> Path:
    """Create the localized dealer summary-completion workbook."""
    rows = build_summary_status_rows(submissions, report_date)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Summary Status"
    sheet.append(["Date", "Region", "Dealer", "Status"])
    for item in rows:
        sheet.append(
            [item["date"], item["region"], item["dealer"], item["status"]]
        )

    navy = "1F4E78"
    green = "C6EFCE"
    red = "FFC7CE"
    thin = Side(style="thin", color="D9E2F3")
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)
    sheet.row_dimensions[1].height = 24

    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 1).number_format = "yyyy-mm-dd"
        sheet.cell(row, 4).fill = PatternFill(
            "solid",
            fgColor=(
                green
                if sheet.cell(row, 4).value == STATUS_SUBMITTED
                else red
            ),
        )
        for col in range(1, 5):
            sheet.cell(row, col).alignment = Alignment(vertical="center")
            sheet.cell(row, col).border = Border(bottom=thin)

    sheet.column_dimensions["A"].width = 16
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 14
    sheet.column_dimensions["D"].width = 32
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:D{sheet.max_row}"
    table = Table(displayName="SummaryStatusTable", ref=f"A1:D{sheet.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.sheet_view.showGridLines = False
    sheet.auto_filter.ref = f"A1:D{sheet.max_row}"
    workbook.save(output_path)
    return output_path
