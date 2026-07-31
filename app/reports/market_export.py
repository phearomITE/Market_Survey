from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
from statistics import mean, median
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.config import settings
from app.reports.aggregator import (
    ALL_COMPETITOR_PRODUCTS,
    ALL_OWN_PRODUCTS,
    is_final_summary_outlet_name,
)


SUMMARY_HEADERS = [
    "Region",
    "Dealer",
    "Location of Visit Text",
    "Member",
    "Total Outlets",
    "Wholesale",
    "Drink Shop",
    "Wet Market",
    "Trolley",
    "Local Eat",
    "Coffe,Bakery",
    "Canteen",
    "Sport Club",
    "Motor Shop",
    "Product",
    "WS",
    "DS",
    "WM",
    "TL",
    "LE",
    "CB",
    "MS",
    "Movement",
]

LOCATION_HEADERS = [
    "Date",
    "Region",
    "Dealer",
    "Latitude",
    "Longitude",
    "Outlet Name",
    "Outlet Type",
    "Phone Number Outlet",
]

OUTLET_TYPES = [
    "Wholesale",
    "Drink Shop",
    "Wet Market",
    "Trolley",
    "Local Eat",
    "Coffee,Bakery",
    "Canteen",
    "Sport Club",
    "Motor Shop",
]

PRODUCT_TYPE_COLUMNS = {
    "Wholesale": "WS",
    "Drink Shop": "DS",
    "Wet Market": "WM",
    "Trolley": "TL",
    "Local Eat": "LE",
    "Coffee,Bakery": "CB",
    "Motor Shop": "MS",
}


def _clean(value) -> str:
    return " ".join(str(value or "").strip().split())


def _member_number(value) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _outlet_type(value) -> str:
    normalized = _clean(value).casefold().replace(" ", "")
    aliases = {
        "wholesale": "Wholesale",
        "drinkshop": "Drink Shop",
        "wetmarket": "Wet Market",
        "trolley": "Trolley",
        "localeat": "Local Eat",
        "coffee,bakery": "Coffee,Bakery",
        "coffe,bakery": "Coffee,Bakery",
        "coffeebakery": "Coffee,Bakery",
        "canteen": "Canteen",
        "sportclub": "Sport Club",
        "motorshop": "Motor Shop",
    }
    return aliases.get(normalized, _clean(value))


def _metric_rows(submission) -> list:
    return [
        *list(getattr(submission, "product_metrics", None) or []),
        *list(getattr(submission, "competitor_metrics", None) or []),
    ]


def _metric_is_present(metric) -> bool:
    return any(
        getattr(metric, name, None) not in (None, "")
        for name in ("status", "movement_score", "stock_status", "buy_in_price", "sell_out_price")
    ) or bool(getattr(metric, "available", False))


def _score(metric) -> int | None:
    value = getattr(metric, "movement_score", None)
    if value in (None, ""):
        return None
    try:
        return max(0, min(10, int(float(value))))
    except (TypeError, ValueError):
        return None


def _final_movement(scores: list[int]) -> int | None:
    """Use the approved absolute rule: 70% mean + 30% median."""
    if not scores:
        return None
    adjusted = (0.7 * mean(scores)) + (0.3 * median(scores))
    return max(0, min(10, int(adjusted + 0.5)))


def _style_sheet(ws, widths: list[int]) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{max(ws.max_row, 1)}"
    ws.row_dimensions[1].height = 25
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def create_market_export(
    submissions: Iterable,
    report_date: date,
    output_path: Path | None = None,
) -> Path:
    """Create the exact two-sheet workbook used by /export YYYY-MM-DD."""
    rows = [
        row
        for row in submissions
        if not is_final_summary_outlet_name(getattr(row, "outlet_name", None))
    ]
    rows.sort(
        key=lambda item: (
            _clean(getattr(item, "region", None)),
            _clean(getattr(item, "dealer", None)),
            _clean(getattr(item, "outlet_name", None)),
        )
    )

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary_Data"
    location_ws = wb.create_sheet("Location_Outlet")
    summary_ws.append(SUMMARY_HEADERS)
    location_ws.append(LOCATION_HEADERS)

    grouped: dict[tuple[str, str], list] = defaultdict(list)
    for submission in rows:
        grouped[
            (
                _clean(getattr(submission, "region", None)),
                _clean(getattr(submission, "dealer", None)),
            )
        ].append(submission)
        location_ws.append(
            [
                getattr(submission, "report_date", None),
                getattr(submission, "region", None),
                getattr(submission, "dealer", None),
                getattr(submission, "gps_latitude", None),
                getattr(submission, "gps_longitude", None),
                getattr(submission, "outlet_name", None),
                getattr(submission, "outlet_type", None),
                getattr(submission, "phone_number", None),
            ]
        )

    preferred_products = list(dict.fromkeys(ALL_OWN_PRODUCTS + ALL_COMPETITOR_PRODUCTS))
    for (region, dealer), dealer_rows in sorted(grouped.items()):
        type_counts = Counter(_outlet_type(getattr(row, "outlet_type", None)) for row in dealer_rows)
        locations = list(
            dict.fromkeys(
                _clean(getattr(row, "location_text", None))
                for row in dealer_rows
                if _clean(getattr(row, "location_text", None))
            )
        )
        members = sorted(
            {
                number
                for row in dealer_rows
                if (number := _member_number(getattr(row, "member_no", None))) is not None
            }
        )
        product_counts: dict[str, Counter] = defaultdict(Counter)
        product_scores: dict[str, list[int]] = defaultdict(list)

        for submission in dealer_rows:
            outlet_type = _outlet_type(getattr(submission, "outlet_type", None))
            for metric in _metric_rows(submission):
                product = _clean(getattr(metric, "product_name", None))
                if not product or not _metric_is_present(metric):
                    continue
                product_counts[product][outlet_type] += 1
                score = _score(metric)
                if score is not None:
                    product_scores[product].append(score)

        present = set(product_counts) | set(product_scores)
        products = [product for product in preferred_products if product in present]
        products.extend(sorted(present.difference(products), key=str.casefold))
        for product in products:
            counts = product_counts[product]
            summary_ws.append(
                [
                    region,
                    dealer,
                    ", ".join(locations),
                    ", ".join(str(value) for value in members),
                    len(dealer_rows),
                    *[type_counts[name] for name in OUTLET_TYPES],
                    product,
                    *[counts[name] for name in PRODUCT_TYPE_COLUMNS],
                    _final_movement(product_scores[product]),
                ]
            )

    for cell in location_ws["A"][1:]:
        cell.number_format = "yyyy-mm-dd"
    for cell in location_ws["H"][1:]:
        cell.number_format = "@"
    _style_sheet(
        summary_ws,
        [10, 12, 42, 12, 14, 12, 13, 12, 10, 11, 14, 11, 12, 12, 24, 8, 8, 8, 8, 8, 8, 8, 12],
    )
    _style_sheet(location_ws, [13, 10, 12, 13, 13, 26, 18, 22])

    settings.export_path.mkdir(parents=True, exist_ok=True)
    output_path = output_path or settings.export_path / f"Market_Survey_Data_{report_date}.xlsx"
    wb.save(output_path)
    return output_path
