
# app/reports/summary_report.py


from __future__ import annotations

from collections import Counter, defaultdict
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.config import settings
from app.data.dealers import REGION_DEALERS
from app.reports.aggregator import (
    aggregate_submissions,
    is_final_summary_outlet_name,
    load_wide_payloads,
)


HEADER_FILL = "1F4E78"
REGION_FILL = "D9EAF7"
ZERO_FILL = "FCE4D6"
PARTIAL_FILL = "FFF2CC"
OK_FILL = "E2F0D9"
BORDER_COLOR = "D9E2F3"
SUMMARY_HEADERS = [
    "Region",
    "Dealer",
    "Member",
    "Total Submissions",
    "Total Outlets",
    "Status",
    "<5",
    "5 to 8",
    "9 to 10",
    "Product Competitor",
    "Movement Lead",
]
SUMMARY_COMPETITORS = ("GB SNOW NCP", "Hanuman LITE NCP", "Greet LITE NCP")


def _clean(value) -> str:
    return str(value or "").strip()


def _status(total_submissions: int, total_outlets: int, target: int | None) -> str:
    if total_submissions <= 0:
        return "❌ No Submit"
    if target and total_outlets < target:
        return "⚠ Partial"
    return "✅"


def _safe_int(value) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(str(value).replace(",", "").strip()))
    except Exception:
        return None


def _product_key(value) -> str:
    return " ".join(_clean(value).upper().split())


def _metric_map(submission, attribute: str) -> dict[str, object]:
    return {
        _product_key(getattr(metric, "product_name", "")): metric
        for metric in list(getattr(submission, attribute, None) or [])
        if _product_key(getattr(metric, "product_name", ""))
    }


def _score(metric) -> int | None:
    if metric is None:
        return None
    value = getattr(metric, "movement_score", None)
    if value in (None, ""):
        return None
    try:
        return max(0, min(10, int(float(value))))
    except (TypeError, ValueError):
        return None


def _adjusted_score(values: Iterable[int | None]) -> float | None:
    """Absolute movement score: 70% mean + 30% median.

    Blank values are excluded. Genuine zeroes remain valid. The result is kept
    unrounded for ranking and rounded only when displayed.
    """
    valid = [int(value) for value in values if value is not None]
    if not valid:
        return None
    return (0.7 * mean(valid)) + (0.3 * median(valid))


def _normal_round(value: float | None) -> int | None:
    if value is None:
        return None
    return max(0, min(10, int(value + 0.5)))


def _dealer_movement(
    submissions: Iterable,
    wide_map: dict | None = None,
) -> dict:
    submission_rows = list(submissions)
    members: set[int] = set()

    for submission in submission_rows:
        member = _safe_int(getattr(submission, "member_no", None))
        if member is not None:
            members.add(member)

    agg = aggregate_submissions(submission_rows, wide_map=wide_map)
    own_data = (agg.get("products") or {}).get("CB LITE NCP") or {}
    own_display = own_data.get("mov")
    ranked_competitors = [
        (product, (agg.get("competitors") or {}).get(product, {}).get("mov"))
        for product in SUMMARY_COMPETITORS
        if (agg.get("competitors") or {}).get(product, {}).get("mov") is not None
    ]
    ranked_competitors.sort(key=lambda item: (-item[1], item[0].casefold()))
    leader_name, leader_display = ranked_competitors[0] if ranked_competitors else (None, None)

    return {
        "member_count": len(members) or None,
        "own_adjusted": own_display,
        "own_display": own_display,
        "competitor": leader_name,
        "competitor_adjusted": leader_display,
        "competitor_display": leader_display,
    }


def _copy_row_style(ws, source_row: int, target_row: int, max_column: int) -> None:
    for column in range(1, max_column + 1):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.protection = copy(source.protection)
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def _set_summary_movement_cells(ws, row_number: int, movement: dict) -> None:
    for column in range(7, 12):
        ws.cell(row_number, column).value = None
    own_score = movement["own_display"]
    if own_score is not None:
        if own_score < 5:
            ws.cell(row_number, 7).value = own_score
        elif own_score <= 8:
            ws.cell(row_number, 8).value = own_score
        else:
            ws.cell(row_number, 9).value = own_score
    ws.cell(row_number, 10).value = movement["competitor"]
    ws.cell(row_number, 11).value = movement["competitor_display"]


def _create_gt_template_report(
    rows: list[dict],
    submissions: Iterable,
    report_date: date,
    output_path: Path,
) -> Path:
    template = settings.gt_summary_template_file
    if not template.exists():
        raise FileNotFoundError(
            f"GT summary template not found: {template}. "
            "Add templates/template_gt_summary.xlsx to the deployment."
        )

    wb = load_workbook(template)
    if "Summary" not in wb.sheetnames:
        raise ValueError("GT summary template must contain a Summary sheet.")
    ws = wb["Summary"]
    # GT management summary must contain only one worksheet. Removing Detail
    # immediately also avoids building thousands of unnecessary detail rows.
    for sheet in list(wb.worksheets):
        if sheet.title != "Summary":
            wb.remove(sheet)

    submission_rows = list(submissions)
    grouped: dict[str, list] = defaultdict(list)
    for submission in submission_rows:
        dealer = _clean(getattr(submission, "dealer", "")).upper()
        if dealer:
            grouped[dealer].append(submission)

    wide_map = load_wide_payloads(submission_rows)
    movements = {
        dealer: _dealer_movement(dealer_rows, wide_map=wide_map)
        for dealer, dealer_rows in grouped.items()
    }

    total_dealers = len(rows)
    submitted_dealers = sum(1 for row in rows if row["total_submissions"] > 0)
    total_submissions = sum(row["total_submissions"] for row in rows)
    total_outlets = sum(row["total_outlets"] for row in rows)

    ws["A1"] = "KB Market Survey - GT Region & Dealer Submission Summary"
    ws["A2"] = f"Report Date: {report_date} | Generated: {datetime.now():%d/%m/%Y %H:%M:%S}"
    ws["A4"] = "Total Regions"
    ws["B4"] = "Total Dealers"
    ws["C4"] = "Submitted Dealers"
    ws["D4"] = "No Submit Dealers"
    ws["E4"] = "Total Submissions"
    ws["F4"] = "<5"
    ws["G4"] = "5 to 8"
    ws["H4"] = "9 to 10"
    ws["I4"] = "GB SNOW NCP"
    ws["J4"] = "Hanuman LITE NCP"
    ws["K4"] = "Greet LITE NCP"

    own_display_scores = [
        movement["own_display"]
        for movement in movements.values()
        if movement["own_display"] is not None
    ]
    competitor_counts = Counter(
        movement["competitor"]
        for movement in movements.values()
        if movement["competitor"]
    )
    ws["A5"] = len(REGION_DEALERS)
    ws["B5"] = total_dealers
    ws["C5"] = submitted_dealers
    ws["D5"] = total_dealers - submitted_dealers
    ws["E5"] = total_submissions
    ws["F5"] = sum(score < 5 for score in own_display_scores)
    ws["G5"] = sum(5 <= score <= 8 for score in own_display_scores)
    ws["H5"] = sum(score >= 9 for score in own_display_scores)
    ws["I5"] = competitor_counts["GB SNOW NCP"]
    ws["J5"] = competitor_counts["Hanuman LITE NCP"]
    ws["K5"] = competitor_counts["Greet LITE NCP"]
    ws["G7"] = "Movement CB LITE NCP compared with competitors"

    for column, value in enumerate(SUMMARY_HEADERS, start=1):
        ws.cell(8, column).value = value

    row_by_dealer = {row["dealer"]: row for row in rows}
    current_row = 9
    for region, dealers in REGION_DEALERS.items():
        region_summary_rows = []
        for dealer in dealers:
            summary_row = row_by_dealer.get(
                dealer,
                {
                    "region": region,
                    "dealer": dealer,
                    "total_submissions": 0,
                    "total_outlets": 0,
                    "status": "❌ No Submit",
                },
            )
            region_summary_rows.append(summary_row)
            movement = movements.get(
                dealer,
                {
                    "member_count": None,
                    "own_display": None,
                    "competitor": None,
                    "competitor_display": None,
                },
            )
            ws.cell(current_row, 1).value = region
            ws.cell(current_row, 2).value = dealer
            ws.cell(current_row, 3).value = movement["member_count"]
            ws.cell(current_row, 4).value = summary_row["total_submissions"]
            ws.cell(current_row, 5).value = summary_row["total_outlets"]
            ws.cell(current_row, 6).value = summary_row["status"]
            _set_summary_movement_cells(ws, current_row, movement)
            status_fill = (
                ZERO_FILL
                if "No Submit" in summary_row["status"]
                else PARTIAL_FILL
                if "Partial" in summary_row["status"]
                else OK_FILL
            )
            for column in range(1, 12):
                ws.cell(current_row, column).fill = PatternFill("solid", fgColor=status_fill)
            current_row += 1

        ws.cell(current_row, 1).value = region
        ws.cell(current_row, 2).value = "Region Total"
        ws.cell(current_row, 3).value = None
        ws.cell(current_row, 4).value = sum(row["total_submissions"] for row in region_summary_rows)
        ws.cell(current_row, 5).value = sum(row["total_outlets"] for row in region_summary_rows)
        submitted = sum(row["total_submissions"] > 0 for row in region_summary_rows)
        ws.cell(current_row, 6).value = f"{submitted}/{len(dealers)} dealers submitted"
        for column in range(7, 12):
            ws.cell(current_row, column).value = None
        for column in range(1, 12):
            ws.cell(current_row, column).fill = PatternFill("solid", fgColor=REGION_FILL)
            ws.cell(current_row, column).font = copy(ws.cell(current_row, column).font)
            ws.cell(current_row, column).font = Font(
                name=ws.cell(current_row, column).font.name or "Calibri",
                size=ws.cell(current_row, column).font.sz or 11,
                bold=True,
                color=ws.cell(current_row, column).font.color,
            )
        current_row += 1

    # Keep the Summary table range valid after replacing old template rows.
    for table in ws.tables.values():
        table.ref = f"A1:K{max(current_row - 1, 8)}"

    wb.save(output_path)
    return output_path


def build_summary_rows(submissions: Iterable) -> list[dict]:
    """Return one row for every configured dealer, including zero-submit dealers.

    Total Submissions = every submitted record.
    Total Outlets = Total Submissions minus exact summary-marker records.
    Status = No Submit / Partial / OK.
    """
    grouped: dict[str, list] = defaultdict(list)
    for s in submissions:
        dealer = _clean(getattr(s, "dealer", "")).upper()
        if dealer:
            grouped[dealer].append(s)

    rows: list[dict] = []
    for region, dealers in REGION_DEALERS.items():
        for dealer in dealers:
            dealer_rows = grouped.get(dealer, [])
            outlet_rows = [
                s for s in dealer_rows
                if not is_final_summary_outlet_name(getattr(s, "outlet_name", None))
            ]
            total_submissions = len(dealer_rows)
            summary_control_count = total_submissions - len(outlet_rows)
            total_outlets = max(0, total_submissions - summary_control_count)

            targets = [_safe_int(getattr(s, "total_outlet_visit_target", None)) for s in outlet_rows]
            targets = [x for x in targets if x is not None]
            target = max(targets) if targets else None

            rows.append(
                {
                    "region": region,
                    "dealer": dealer,
                    "total_submissions": total_submissions,
                    "total_outlets": total_outlets,
                    "target": target,
                    "status": _status(total_submissions, total_outlets, target),
                }
            )
    return rows


def _style_summary_sheet(ws) -> None:
    thin = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", horizontal="center")
            cell.font = Font(name="Calibri", size=11)

    # Title rows
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=HEADER_FILL)
    ws["A2"].font = Font(name="Calibri", size=11, italic=True, color="666666")

    # KPI block
    for row in range(4, 7):
        for col in range(1, 9):
            c = ws.cell(row, col)
            c.fill = PatternFill("solid", fgColor="F8FBFD")
            c.font = Font(name="Calibri", size=11, bold=(row == 4))

    # Header
    header_row = 8
    for col in range(1, 6):
        c = ws.cell(header_row, col)
        c.fill = PatternFill("solid", fgColor=HEADER_FILL)
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for row in range(9, ws.max_row + 1):
        status = str(ws.cell(row, 5).value or "")
        fill = None
        if "No Submit" in status:
            fill = PatternFill("solid", fgColor=ZERO_FILL)
        elif "Partial" in status:
            fill = PatternFill("solid", fgColor=PARTIAL_FILL)
        elif "✅" in status:
            fill = PatternFill("solid", fgColor=OK_FILL)
        if fill:
            for col in range(1, 6):
                ws.cell(row, col).fill = fill

        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).font = Font(bold=True)

    ws.freeze_panes = "A9"
    widths = {"A": 12, "B": 14, "C": 20, "D": 16, "E": 18, "G": 22, "H": 14}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 22


def create_summary_report(
    rows: list[dict],
    report_date: date,
    output_path: Path | None = None,
    report_type: str = "GT",
    submissions: Iterable | None = None,
) -> Path:
    settings.export_path.mkdir(parents=True, exist_ok=True)
    report_type = str(report_type or "GT").upper()
    output_path = output_path or settings.export_path / f"Market_Survey_Summary_{report_type}_{report_date}.xlsx"

    if report_type == "GT":
        return _create_gt_template_report(
            rows,
            list(submissions or []),
            report_date,
            output_path,
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    total_dealers = len(rows)
    submitted_dealers = sum(1 for r in rows if r["total_submissions"] > 0)
    no_submit = total_dealers - submitted_dealers
    total_submissions = sum(r["total_submissions"] for r in rows)
    total_outlets = sum(r["total_outlets"] for r in rows)
    completion = submitted_dealers / total_dealers if total_dealers else 0

    ws.merge_cells("A1:H1")
    ws["A1"] = f"KB Market Survey - {report_type} Region & Dealer Submission Summary"
    ws.merge_cells("A2:H2")
    ws["A2"] = f"Report Date: {report_date} | Generated: {datetime.now():%d/%m/%Y %H:%M:%S}"

    kpis = [
        ("Total Regions", len(set(r["region"] for r in rows))),
        ("Total Dealers", total_dealers),
        ("Submitted Dealers", submitted_dealers),
        ("No Submit Dealers", no_submit),
        ("Total Submissions", total_submissions),
        ("Total Outlets", total_outlets),
        ("Completion", f"{completion:.1%}"),
    ]
    for idx, (label, value) in enumerate(kpis, start=1):
        ws.cell(4, idx).value = label
        ws.cell(5, idx).value = value

    header = ["Region", "Dealer", "Total Submissions", "Total Outlets", "Status"]
    for col, value in enumerate(header, start=1):
        ws.cell(8, col).value = value

    current_row = 9
    for region, dealers in REGION_DEALERS.items():
        region_rows = [r for r in rows if r["region"] == region]
        for r in region_rows:
            ws.cell(current_row, 1).value = r["region"]
            ws.cell(current_row, 2).value = r["dealer"]
            ws.cell(current_row, 3).value = r["total_submissions"]
            ws.cell(current_row, 4).value = r["total_outlets"]
            ws.cell(current_row, 5).value = r["status"]
            current_row += 1

        # Region subtotal row
        ws.cell(current_row, 1).value = region
        ws.cell(current_row, 2).value = "Region Total"
        ws.cell(current_row, 3).value = sum(r["total_submissions"] for r in region_rows)
        ws.cell(current_row, 4).value = sum(r["total_outlets"] for r in region_rows)
        submitted = sum(1 for r in region_rows if r["total_submissions"] > 0)
        ws.cell(current_row, 5).value = f"{submitted}/{len(region_rows)} dealers submitted"
        for col in range(1, 6):
            ws.cell(current_row, col).fill = PatternFill("solid", fgColor=REGION_FILL)
            ws.cell(current_row, col).font = Font(bold=True)
        current_row += 1

    _style_summary_sheet(ws)
    wb.save(output_path)
    return output_path
