from __future__ import annotations

from datetime import date
from pathlib import Path
import unicodedata

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.data.dealers import REGION_DEALERS


SUMMARY_OUTLET_TEXT = "បូកសរុបរួម"


def _clean_text(value: object) -> str:
    text = unicodedata.normalize("NFC", str(value or ""))
    return text.replace("\u200b", "").replace("\ufeff", "").strip()


def is_summary_submission(outlet_name: object) -> bool:
    """Only the official summary outlet name counts as dealer submission."""
    return SUMMARY_OUTLET_TEXT in _clean_text(outlet_name)


def create_summary_status_export(
    submissions: list[object],
    report_date: date,
    output_path: Path | None = None,
) -> Path:
    """Create status rows for every official dealer, including missing dates."""
    official = {dealer for dealers in REGION_DEALERS.values() for dealer in dealers}
    submitted = {
        _clean_text(getattr(row, "dealer", "")).upper()
        for row in submissions
        if is_summary_submission(getattr(row, "outlet_name", ""))
    } & official

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Summary Status"
    sheet.append(["Date", "Region", "Dealer", "Status"])

    for region, dealers in REGION_DEALERS.items():
        for dealer in dealers:
            status = "Submitted Summary" if dealer in submitted else "Missing Summary"
            sheet.append([report_date, region, dealer, status])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    submitted_fill = PatternFill("solid", fgColor="C6EFCE")
    missing_fill = PatternFill("solid", fgColor="FFC7CE")
    thin = Side(style="thin", color="D9E2F3")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row_number in range(2, sheet.max_row + 1):
        sheet.cell(row_number, 1).number_format = "yyyy-mm-dd"
        status_cell = sheet.cell(row_number, 4)
        status_cell.fill = submitted_fill if status_cell.value == "Submitted Summary" else missing_fill
        for cell in sheet[row_number]:
            cell.border = Border(bottom=thin)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:D{sheet.max_row}"
    sheet.column_dimensions["A"].width = 15
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 14
    sheet.column_dimensions["D"].width = 24
    sheet.row_dimensions[1].height = 25

    if output_path is None:
        from app.core.config import settings

        destination = settings.export_path / f"Dealer_Summary_Status_{report_date}.xlsx"
    else:
        destination = output_path
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination
