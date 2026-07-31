"""Dedicated simple GT Region/Dealer submission summary.

This module intentionally contains no movement, product, competitor, or detail
report logic.  The Telegram command ``/summary GT YYYY-MM-DD`` calls this
generator directly so it cannot fall back to the legacy management summary.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.core.config import settings
from app.data.dealers import REGION_DEALERS


TITLE_FILL = "1F4E78"
REGION_TOTAL_FILL = "D9EAF7"
SUBMITTED_FILL = "E2F0D9"
NO_SUBMIT_FILL = "FCE4D6"
GRID_COLOR = "D9E2F3"


def create_gt_submission_summary(
    rows: list[dict],
    report_date: date,
    output_path: Path | None = None,
) -> Path:
    """Create the requested one-sheet GT submission summary workbook only."""
    settings.export_path.mkdir(parents=True, exist_ok=True)
    output_path = output_path or (
        settings.export_path / f"GT_Submission_Summary_{report_date}.xlsx"
    )

    # Never reuse an old workbook containing legacy movement/detail sheets.
    output_path.unlink(missing_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Summary"
    sheet.sheet_view.showGridLines = False

    total_dealers = len(rows)
    submitted_dealers = sum(
        1 for row in rows if int(row.get("total_submissions", 0) or 0) > 0
    )
    total_submissions = sum(
        int(row.get("total_submissions", 0) or 0) for row in rows
    )
    total_outlets = sum(int(row.get("total_outlets", 0) or 0) for row in rows)
    completion = submitted_dealers / total_dealers if total_dealers else 0

    sheet.merge_cells("A1:H1")
    sheet["A1"] = "KB Market Survey - GT Region & Dealer Submission Summary"
    sheet.merge_cells("A2:H2")
    sheet["A2"] = (
        f"Report Date: {report_date:%Y-%m-%d} | "
        f"Generated: {datetime.now():%d/%m/%Y %H:%M:%S}"
    )

    kpis = [
        ("Total Regions", len(REGION_DEALERS)),
        ("Total Dealers", total_dealers),
        ("Submitted Dealers", submitted_dealers),
        ("No Submit Dealers", total_dealers - submitted_dealers),
        ("Total Submissions", total_submissions),
        ("Total Outlets", total_outlets),
        ("Completion", completion),
    ]
    for column, (label, value) in enumerate(kpis, start=1):
        sheet.cell(4, column, label)
        sheet.cell(5, column, value)
    sheet["G5"].number_format = "0.0%"

    headers = [
        "Region",
        "Dealer",
        "Total Submissions",
        "Total Outlets",
        "Status",
    ]
    for column, label in enumerate(headers, start=1):
        sheet.cell(8, column, label)

    row_number = 9
    rows_by_dealer = {
        str(row.get("dealer", "")).strip().upper(): row for row in rows
    }
    for region, dealers in REGION_DEALERS.items():
        region_rows: list[dict] = []
        for dealer in dealers:
            row = rows_by_dealer.get(
                dealer,
                {
                    "region": region,
                    "dealer": dealer,
                    "total_submissions": 0,
                    "total_outlets": 0,
                    "status": "❌ No Submit",
                },
            )
            region_rows.append(row)
            values = [
                region,
                dealer,
                int(row.get("total_submissions", 0) or 0),
                int(row.get("total_outlets", 0) or 0),
                "✅"
                if int(row.get("total_submissions", 0) or 0) > 0
                else "❌ No Submit",
            ]
            for column, value in enumerate(values, start=1):
                sheet.cell(row_number, column, value)
            row_number += 1

        submitted = sum(
            1
            for row in region_rows
            if int(row.get("total_submissions", 0) or 0) > 0
        )
        totals = [
            region,
            "Region Total",
            sum(int(row.get("total_submissions", 0) or 0) for row in region_rows),
            sum(int(row.get("total_outlets", 0) or 0) for row in region_rows),
            f"{submitted}/{len(dealers)} dealers submitted",
        ]
        for column, value in enumerate(totals, start=1):
            sheet.cell(row_number, column, value)
        row_number += 1

    _apply_simple_style(sheet)
    workbook.save(output_path)
    _validate_simple_workbook(output_path)
    return output_path


def _apply_simple_style(sheet) -> None:
    thin = Side(style="thin", color=GRID_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet["A1"].fill = PatternFill("solid", fgColor=TITLE_FILL)
    sheet["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["A2"].font = Font(
        name="Calibri", size=11, italic=True, color="666666"
    )
    sheet["A2"].alignment = Alignment(horizontal="center", vertical="center")

    for row in range(4, 6):
        for column in range(1, 8):
            cell = sheet.cell(row, column)
            cell.fill = PatternFill("solid", fgColor="F8FBFD")
            cell.font = Font(name="Calibri", size=11, bold=(row == 4))

    for column in range(1, 6):
        cell = sheet.cell(8, column)
        cell.fill = PatternFill("solid", fgColor=TITLE_FILL)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    region_total_rows = set()
    for row in range(9, sheet.max_row + 1):
        if sheet.cell(row, 2).value == "Region Total":
            region_total_rows.add(row)
            fill = PatternFill("solid", fgColor=REGION_TOTAL_FILL)
        elif sheet.cell(row, 5).value == "✅":
            fill = PatternFill("solid", fgColor=SUBMITTED_FILL)
        else:
            fill = PatternFill("solid", fgColor=NO_SUBMIT_FILL)
        for column in range(1, 6):
            sheet.cell(row, column).fill = fill

    for row in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 22
        for column in range(1, 9):
            cell = sheet.cell(row, column)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if row >= 9 and column in (1, 2):
                cell.font = Font(
                    name="Calibri",
                    size=11,
                    bold=True,
                )
            elif row not in (1, 2, 4, 8) and row not in region_total_rows:
                cell.font = Font(name="Calibri", size=11)

    sheet.column_dimensions["A"].width = 12
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 20
    sheet.column_dimensions["D"].width = 16
    sheet.column_dimensions["E"].width = 24
    sheet.column_dimensions["F"].width = 12
    sheet.column_dimensions["G"].width = 18
    sheet.column_dimensions["H"].width = 12
    sheet.freeze_panes = "A9"


def _validate_simple_workbook(path: Path) -> None:
    """Fail before Telegram upload if legacy content appears in the file."""
    workbook = load_workbook(path, read_only=True, data_only=False)
    if workbook.sheetnames != ["Summary"]:
        raise RuntimeError(
            "GT summary validation failed: workbook must contain only Summary."
        )
    sheet = workbook["Summary"]
    expected_headers = [
        "Region",
        "Dealer",
        "Total Submissions",
        "Total Outlets",
        "Status",
    ]
    actual_headers = [sheet.cell(8, column).value for column in range(1, 6)]
    if actual_headers != expected_headers:
        raise RuntimeError("GT summary validation failed: incorrect simple layout.")
    forbidden = {
        "movement",
        "competitor",
        "product",
        "detail",
        "member",
        "<5",
        "5 to 8",
        "9 to 10",
    }
    for row in sheet.iter_rows():
        for cell in row:
            text = str(cell.value or "").strip().lower()
            if text in forbidden or "compare to competitors" in text:
                raise RuntimeError(
                    f"GT summary validation failed: legacy field {cell.value!r}."
                )
