from datetime import date
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from app.reports.summary_report import build_summary_rows, create_summary_report


def _submission(outlet: str, cb_lite: int, gb_snow: int):
    product = SimpleNamespace(
        product_name="CB LITE NCP",
        movement_score=cb_lite,
        stock_status="full",
        bbe_date="2026-12-01",
        available=True,
        status="yes",
        volume_ctn=None,
        buy_in_price=None,
        sell_out_price=None,
        ring_pull_value=None,
        new_outlet_purchase=False,
    )
    competitor = SimpleNamespace(
        product_name="GB SNOW NCP",
        movement_score=gb_snow,
        stock_status="full",
        status="yes",
        buy_in_price=None,
        sell_out_price=None,
    )
    return SimpleNamespace(
        dealer="CA1",
        region="R1",
        outlet_name=outlet,
        outlet_type="Drink Shop",
        phone_number="012345678",
        report_date=date(2026, 7, 25),
        member_no=12,
        group_no=1,
        total_outlet_visit_target=2,
        gps_latitude=11.55,
        gps_longitude=104.91,
        location_text="",
        key_issue_text="",
        suggestion_text="",
        product_metrics=[product],
        competitor_metrics=[competitor],
        ring_pull_metrics=[],
    )


def test_gt_summary_uses_approved_summary_and_detail_template(tmp_path: Path):
    rows = build_summary_rows(
        [
            _submission("Outlet A", 6, 10),
            _submission("Outlet B", 7, 9),
        ]
    )
    ca1 = next(row for row in rows if row["dealer"] == "CA1")

    assert ca1["member"] == 12
    assert ca1["movement"] == 7
    assert ca1["lead_product"] == "GB SNOW NCP"
    assert ca1["lead_score"] == 10
    assert len(ca1["detail_rows"]) == 2

    output = create_summary_report(
        rows,
        date(2026, 7, 25),
        tmp_path / "summary.xlsx",
        report_type="GT",
    )
    workbook = load_workbook(output)

    assert workbook.sheetnames == ["Summary", "Detail"]
    assert workbook["Summary"]["A1"].value == (
        "KB Market Survey - Region & Dealer Submission Summary"
    )
    assert workbook["Summary"]["B5"].value == 65
    assert workbook["Detail"]["D2"].value == "Outlet A"
    assert workbook["Detail"]["J2"].value == "GB SNOW NCP"
    assert workbook["Detail"]["L2"].hyperlink is not None

