from datetime import date
from types import SimpleNamespace

from openpyxl import load_workbook

from app.reports.movement_exports import (
    BASE_HEADERS,
    MOVEMENT_MULTI_PRODUCTS,
    RAW_MOVEMENT_HEADERS,
    create_movement_multi_export,
    create_raw_movement_export,
)


def _metric(name: str, score: int | None):
    return SimpleNamespace(product_name=name, movement_score=score)


def _submission():
    return SimpleNamespace(
        report_date=date(2026, 7, 25),
        region="R1",
        dealer="CA1",
        gps_latitude=11.55,
        gps_longitude=104.91,
        outlet_name="Test Outlet",
        outlet_type="Drink Shop",
        phone_number="012345678",
        product_metrics=[
            _metric("CB LITE NCP", 10),
            _metric("Unrelated Own Product", 7),
        ],
        competitor_metrics=[
            _metric("GB SNOW NCP", 0),
            _metric("Hanuman LITE NCP", None),
            _metric("Krud LITE NCP", 5),
            _metric("Greet LITE NCP", 4),
            _metric("Other Competitor", 3),
        ],
    )


def test_movement_multi_has_exact_fixed_columns(tmp_path):
    output = create_movement_multi_export(
        [_submission()],
        [date(2026, 7, 25)],
        output_path=tmp_path / "multi.xlsx",
    )
    ws = load_workbook(output, data_only=True).active
    headers = [cell.value for cell in ws[1]]
    assert headers == BASE_HEADERS + MOVEMENT_MULTI_PRODUCTS
    values = dict(zip(headers, [cell.value for cell in ws[2]]))
    assert values["CB LITE NCP"] == 10
    assert values["GB SNOW NCP"] == 0
    assert values["Hanuman LITE NCP"] is None
    assert values["Krud LITE NCP"] == 5
    assert values["Greet LITE NCP"] == 4


def test_raw_movement_is_normalized_and_preserves_zero(tmp_path):
    output = create_raw_movement_export(
        [_submission()],
        date(2026, 7, 25),
        output_path=tmp_path / "raw.xlsx",
    )
    ws = load_workbook(output, data_only=True).active
    assert [cell.value for cell in ws[1]] == RAW_MOVEMENT_HEADERS
    rows = [tuple(cell.value for cell in row) for row in ws.iter_rows(min_row=2)]
    assert len(rows) == 6
    by_product = {row[3]: row[4] for row in rows}
    assert by_product["GB SNOW NCP"] == 0
    assert "Hanuman LITE NCP" not in by_product
    assert by_product["Other Competitor"] == 3
