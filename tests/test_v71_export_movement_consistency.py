from __future__ import annotations

from datetime import date
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from app.reports.aggregator import (
    _apply_offtake_comparison_goal,
    _movement_from_payload,
    final_offtake_movement,
    movement_average,
)
from app.reports.data_export import create_data_export
from app.reports.summary_report import movement_summary_from_aggregate


def test_average_uses_only_positive_submitted_scores():
    values = [10, 10, 10, 10, 10, 10, 9, 10]
    assert movement_average(values) == 79 / 8
    assert final_offtake_movement(values) == 10
    assert _movement_from_payload(
        {"fresh_movement_score_cb_lite_ncp": 0},
        "CB LITE NCP",
        False,
    ) is None


def test_same_increase_is_applied_to_every_scored_product():
    result = {
        "products": {
            "CB LITE NCP": {"mov": 8, "_mov_avg": 8.4},
        },
        "competitors": {
            "GB SNOW NCP": {"mov": 6, "_mov_avg": 6.2},
        },
    }
    _apply_offtake_comparison_goal(result)
    assert result["products"]["CB LITE NCP"]["mov"] == 10
    assert result["competitors"]["GB SNOW NCP"]["mov"] == 8


def test_only_one_final_ten_in_comparison_group():
    result = {
        "products": {
            "CB LITE NCP": {"mov": 10, "_mov_avg": 9.875},
        },
        "competitors": {
            "GB SNOW NCP": {"mov": 8, "_mov_avg": 7.75},
            "Hanuman LITE NCP": {"mov": 8, "_mov_avg": 7.6},
            "Krud LITE NCP": {"mov": 7, "_mov_avg": 6.8},
            "Greet LITE NCP": {"mov": 6, "_mov_avg": 5.9},
        },
    }
    _apply_offtake_comparison_goal(result)
    scores = [
        result["products"]["CB LITE NCP"]["mov"],
        result["competitors"]["GB SNOW NCP"]["mov"],
        result["competitors"]["Hanuman LITE NCP"]["mov"],
        result["competitors"]["Krud LITE NCP"]["mov"],
        result["competitors"]["Greet LITE NCP"]["mov"],
    ]
    assert scores.count(10) == 1
    assert len(scores) == len(set(scores))


def test_summary_and_export_read_same_final_movement(tmp_path: Path):
    aggregate = {
        "dealer": "CA1",
        "region": "R1",
        "report_date": date(2026, 7, 25),
        "total_outlets": 1,
        "outlet_types": {"Drink Shop": 1},
        "location_text": "Phnom Penh",
        "products": {
            "CB LITE NCP": {
                "mov": 10,
                "availability": {"Drink Shop": 1},
            },
        },
        "competitors": {
            "GB SNOW NCP": {"mov": 8, "availability": {"Drink Shop": 1}},
            "Hanuman LITE NCP": {"mov": 7, "availability": {}},
            "Krud LITE NCP": {"mov": 6, "availability": {}},
            "Greet LITE NCP": {"mov": 5, "availability": {}},
        },
    }
    submission = SimpleNamespace(
        id=1,
        submission_id="1",
        submission_time=None,
        report_date=date(2026, 7, 25),
        region="R1",
        dealer="CA1",
        group_no=1,
        member_no=7,
        total_outlet_visit_target=None,
        outlet_name="Outlet 1",
        outlet_type="Drink Shop",
        phone_number="+855123",
        location_text="Phnom Penh",
        gps_text="",
        gps_latitude=11.5,
        gps_longitude=104.9,
        updated_at=None,
        product_metrics=(),
        competitor_metrics=(),
        ring_pull_metrics=(),
        key_issue_text="",
        suggestion_text="",
    )

    summary = movement_summary_from_aggregate(aggregate)
    assert summary["movement_9_10"] == 10
    assert summary["product_competitor"] == ""

    output = tmp_path / "export.xlsx"
    path, _stats = create_data_export(
        [submission],
        date(2026, 7, 25),
        output_path=output,
        dealer_aggregates={"CA1": aggregate},
    )
    wb = load_workbook(path, data_only=True)
    ws = wb["Summary_Data"]
    headers = {cell.value: cell.column for cell in ws[1]}
    product_col = headers["Product"]
    movement_col = headers["Movement"]
    cb_row = next(
        row for row in range(2, ws.max_row + 1)
        if ws.cell(row, product_col).value == "CB LITE NCP"
    )
    assert ws.cell(cb_row, movement_col).value == 10
