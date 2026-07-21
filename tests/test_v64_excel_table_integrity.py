from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo

from app.reports.data_export import _resize_sheet_tables


def _invalid_template(path: Path) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary_Data"
    headers = [
        "Region", "Dealer", "Location of Visit Text", "Member", "Total Outlets",
        "Wholesale", "Drink Shop", "Wet Market", "Trolley", "Local Eat",
        "Coffe,Bakery", "Canteen", "Sport Club", "Motor Shop", "Product",
        "WS", "DS", "WM", "TL", "LE", "CB", "MS", "Movement",
    ]
    summary.append(headers)

    # Reproduce the bad template state: A:W is 23 columns but the table XML
    # carries a stale 24th tableColumn named Column2.
    table = Table(displayName="Table1", ref="A1:W1048576")
    table.tableColumns = [
        TableColumn(id=index, name=name)
        for index, name in enumerate(headers + ["Column2"], start=1)
    ]
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium7", showRowStripes=True
    )
    summary.add_table(table)

    location = wb.create_sheet("Location_Outlet")
    location_headers = [
        "Date", "Region", "Dealer", "Latitude", "Longitude",
        "Outlet Name", "Outlet Type", "Phone Number Outlet",
    ]
    location.append(location_headers)
    location_table = Table(displayName="Table3", ref="A1:H1048576")
    location_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium7", showRowStripes=True
    )
    location.add_table(location_table)
    wb.save(path)


def test_rebuilds_table_columns_to_match_real_output_range(tmp_path):
    source = tmp_path / "invalid.xlsx"
    output = tmp_path / "fixed.xlsx"
    _invalid_template(source)

    wb = load_workbook(source)
    summary = wb["Summary_Data"]
    for row in range(2, 7):
        summary.cell(row, 1).value = "R1"
    _resize_sheet_tables(summary, last_row=6, last_column=23)

    location = wb["Location_Outlet"]
    location.cell(2, 1).value = "18/07/2026"
    _resize_sheet_tables(location, last_row=2, last_column=8)
    wb.save(output)

    fixed = load_workbook(output)
    summary_table = next(iter(fixed["Summary_Data"].tables.values()))
    location_table = next(iter(fixed["Location_Outlet"].tables.values()))

    assert summary_table.ref == "A1:W6"
    assert len(summary_table.tableColumns) == 23
    assert [column.name for column in summary_table.tableColumns][-1] == "Movement"
    assert location_table.ref == "A1:H2"
    assert len(location_table.tableColumns) == 8
