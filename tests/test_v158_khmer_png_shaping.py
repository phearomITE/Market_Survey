from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from app.reports.excel_report import (
    SUMMARY_FONT_NAME,
    _normalize_khmer_cells,
    _normalize_khmer_text,
)


CORRECT_KHMER_WORD = "គ្រប់"
CORRECT_CODEPOINTS = [0x1782, 0x17D2, 0x179A, 0x1794, 0x17CB]


def test_exact_khmer_coeng_sequence_is_preserved():
    normalized = _normalize_khmer_text(CORRECT_KHMER_WORD)
    assert normalized == CORRECT_KHMER_WORD
    assert [ord(char) for char in normalized] == CORRECT_CODEPOINTS


def test_hidden_and_whitespace_cluster_breaks_are_repaired():
    broken_values = (
        "គ្\u200bរប់",
        "គ្\u200cរប់",
        "គ្\u200dរប់",
        "គ្\u2060រប់",
        "គ្\ufeffរប់",
        "គ្ រប់",
        "គ្\u00a0រប់",
        "គ្\u202fរប់",
    )
    for broken in broken_values:
        assert _normalize_khmer_text(broken) == CORRECT_KHMER_WORD


def test_normal_word_spaces_are_not_removed():
    phrase = "សួស្តី ពិភពលោក"
    assert _normalize_khmer_text(phrase) == phrase


def test_saved_workbook_keeps_one_unicode_run_and_explicit_font(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "គ្\u200bរប់"
    sheet["A1"].font = Font(name="Calibri", scheme="minor")
    _normalize_khmer_cells(sheet)

    output = tmp_path / "khmer-shaping.xlsx"
    workbook.save(output)
    reloaded = load_workbook(output)
    cell = reloaded.active["A1"]

    assert cell.value == CORRECT_KHMER_WORD
    assert [ord(char) for char in cell.value] == CORRECT_CODEPOINTS
    assert cell.font.name == SUMMARY_FONT_NAME
    assert cell.font.scheme is None
    assert cell.font.charset is None


def test_docker_image_installs_complex_text_shaping_dependencies():
    dockerfile = Path("Dockerfile").read_text(encoding="utf-8")
    assert "fonts-noto-core" in dockerfile
    assert "libharfbuzz0b" in dockerfile
    assert "libfribidi0" in dockerfile
    assert 'fc-match -f' in dockerfile

