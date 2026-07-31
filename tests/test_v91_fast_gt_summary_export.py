from datetime import date
from types import SimpleNamespace

from openpyxl import load_workbook

from app.reports.summary_report import build_summary_rows, create_summary_report


def _row(dealer: str, outlet: str):
    return SimpleNamespace(
        dealer=dealer,
        outlet_name=outlet,
        total_outlet_visit_target=None,
    )


def test_gt_summary_matches_approved_simple_layout(tmp_path):
    rows = build_summary_rows(
        [_row("CA1", "Outlet A"), _row("CA1", "Outlet B"), _row("CA1", "Outlet B")]
    )
    path = create_summary_report(
        rows,
        date(2026, 7, 25),
        output_path=tmp_path / "summary.xlsx",
        report_type="GT",
    )

    workbook = load_workbook(path, data_only=False)
    assert workbook.sheetnames == ["Summary"]
    sheet = workbook["Summary"]
    assert sheet["A1"].value == "KB Market Survey - GT Region & Dealer Submission Summary"
    assert [sheet.cell(4, col).value for col in range(1, 8)] == [
        "Total Regions",
        "Total Dealers",
        "Submitted Dealers",
        "No Submit Dealers",
        "Total Submissions",
        "Total Outlets",
        "Completion",
    ]
    assert [sheet.cell(8, col).value for col in range(1, 6)] == [
        "Region",
        "Dealer",
        "Total Submissions",
        "Total Outlets",
        "Status",
    ]
    assert sheet["B9"].value == "CA1"
    assert sheet["C9"].value == 3
    assert sheet["D9"].value == 2

