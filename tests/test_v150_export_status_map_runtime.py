from datetime import date
from pathlib import Path
from types import SimpleNamespace
import tempfile
import unittest

from openpyxl import load_workbook

from app.data.dealers import ALL_DEALERS
from app.reports.status_export import create_summary_status_export, is_summary_submission


class ExportStatusTests(unittest.TestCase):
    def test_only_summary_outlet_counts_and_all_65_are_written(self):
        rows = [
            SimpleNamespace(dealer="CA1", outlet_name="Shop A"),
            SimpleNamespace(dealer="CA3", outlet_name="  បូកសរុបរួម  "),
            SimpleNamespace(dealer="UNKNOWN", outlet_name="បូកសរុបរួម"),
        ]
        with tempfile.TemporaryDirectory() as folder:
            path = create_summary_status_export(
                rows, date(2026, 8, 1), Path(folder) / "status.xlsx"
            )
            sheet = load_workbook(path).active
            self.assertEqual(sheet.max_row, 66)
            statuses = {sheet.cell(r, 3).value: sheet.cell(r, 4).value for r in range(2, 67)}
            self.assertEqual(set(statuses), set(ALL_DEALERS))
            self.assertEqual(statuses["CA3"], "Submitted Summary")
            self.assertEqual(statuses["CA1"], "Missing Summary")
            self.assertEqual(sum(v == "Submitted Summary" for v in statuses.values()), 1)

    def test_empty_date_makes_every_dealer_missing(self):
        with tempfile.TemporaryDirectory() as folder:
            path = create_summary_status_export(
                [], date(2026, 8, 2), Path(folder) / "empty.xlsx"
            )
            sheet = load_workbook(path).active
            self.assertTrue(all(sheet.cell(r, 4).value == "Missing Summary" for r in range(2, 67)))

    def test_contains_match(self):
        self.assertTrue(is_summary_submission("Dealer - បូកសរុបរួម"))
        self.assertFalse(is_summary_submission("Normal outlet"))


if __name__ == "__main__":
    unittest.main()
