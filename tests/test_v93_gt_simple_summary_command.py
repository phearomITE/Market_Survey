from datetime import date
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from app.reports.gt_submission_summary import create_gt_submission_summary


def test_gt_summary_contains_only_the_approved_simple_layout(tmp_path):
    rows = [
        {
            "region": "R1",
            "dealer": "CA1",
            "total_submissions": 25,
            "total_outlets": 25,
            "status": "✅",
        },
        {
            "region": "R1",
            "dealer": "CA8",
            "total_submissions": 0,
            "total_outlets": 0,
            "status": "❌ No Submit",
        },
    ]
    path = create_gt_submission_summary(
        rows,
        date(2026, 7, 25),
        tmp_path / "GT_Submission_Summary_2026-07-25.xlsx",
    )
    workbook = load_workbook(path, data_only=False)
    assert workbook.sheetnames == ["Summary"]
    sheet = workbook["Summary"]
    assert sheet["A1"].value == (
        "KB Market Survey - GT Region & Dealer Submission Summary"
    )
    assert [sheet.cell(8, column).value for column in range(1, 6)] == [
        "Region",
        "Dealer",
        "Total Submissions",
        "Total Outlets",
        "Status",
    ]
    all_text = " ".join(
        str(cell.value or "")
        for row in sheet.iter_rows()
        for cell in row
    ).lower()
    assert "movement" not in all_text
    assert "competitor" not in all_text
    assert "product" not in all_text
    assert "detail" not in all_text


def test_gt_summary_replaces_old_file_instead_of_reusing_legacy_sheets(tmp_path):
    output = tmp_path / "GT_Submission_Summary_2026-07-25.xlsx"
    output.write_bytes(b"old invalid workbook")
    path = create_gt_submission_summary(
        [],
        date(2026, 7, 25),
        output,
    )
    workbook = load_workbook(path)
    assert workbook.sheetnames == ["Summary"]


def test_summary_handler_routes_gt_to_dedicated_generator():
    source = Path("app/bot/handlers.py").read_text(encoding="utf-8")
    assert "generate_gt_submission_summary" in source
    assert 'if report_type == "GT":' in source
