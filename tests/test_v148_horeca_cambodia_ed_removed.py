from openpyxl import load_workbook

from app.reports.excel_report import _layout_rows


def test_horeca_template_has_no_cambodia_ed():
    wb = load_workbook("templates/template_horeca_products.xlsx", data_only=False)
    ws = wb.worksheets[0]
    values = {
        str(cell.value).strip()
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    }
    assert "CAMBODIA ED" not in values


def test_horeca_layout_matches_product_removed_template():
    wb = load_workbook("templates/template_horeca_products.xlsx", data_only=False)
    ws = wb.worksheets[0]
    layout = _layout_rows(ws, {"report_type": "HORECA", "channel": "HORECA"})

    assert ws.cell(layout["freshness_end"], 2).value == "CAMBODIA WATER 500mL"
    assert ws.cell(layout["movement_header"], 1).value == "#"
    assert ws.cell(layout["movement_start"], 2).value == "CB Pint"
    assert ws.cell(layout["movement_end"], 2).value == "CAMBODIA WATER 500mL"
    assert ws.cell(layout["ring_start"], 2).value == "CBL NCP 6 Can"


def test_kobo_form_hides_cambodia_ed_for_horeca_but_keeps_gt():
    wb = load_workbook("templates/KB_Market_Improvement_XLSForm_GT_HORECA.xlsx", data_only=False)
    ws = wb["survey"]

    row = next(
        row
        for row in range(2, ws.max_row + 1)
        if ws.cell(row, 2).value == "fresh_cambodia_ed_group"
    )
    assert ws.cell(row, 8).value == "${final_summary_report_type} = 'gt'"
