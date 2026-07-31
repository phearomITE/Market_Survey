from datetime import date
from types import SimpleNamespace

from openpyxl import load_workbook

from app.reports.market_export import (
    LOCATION_HEADERS,
    SUMMARY_HEADERS,
    create_market_export,
)


def _metric(name, score, status="sale"):
    return SimpleNamespace(
        product_name=name,
        movement_score=score,
        status=status,
        stock_status="full",
        available=True,
        buy_in_price=None,
        sell_out_price=None,
    )


def test_export_has_exact_two_sheets_and_columns(tmp_path):
    submission = SimpleNamespace(
        report_date=date(2026, 7, 25),
        region="R1",
        dealer="CA1",
        member_no=7,
        location_text="Prek Kdam",
        gps_latitude=11.55,
        gps_longitude=104.91,
        outlet_name="Test Outlet",
        outlet_type="Drink Shop",
        phone_number="012345678",
        product_metrics=[_metric("CB LITE NCP", 8)],
        competitor_metrics=[_metric("GB SNOW NCP", 7)],
    )
    output = create_market_export(
        [submission],
        date(2026, 7, 25),
        output_path=tmp_path / "export.xlsx",
    )
    workbook = load_workbook(output, data_only=True)
    assert workbook.sheetnames == ["Summary_Data", "Location_Outlet"]
    assert [cell.value for cell in workbook["Summary_Data"][1]] == SUMMARY_HEADERS
    assert [cell.value for cell in workbook["Location_Outlet"][1]] == LOCATION_HEADERS
    assert workbook["Summary_Data"]["A2"].value == "R1"
    assert workbook["Summary_Data"]["B2"].value == "CA1"
    assert workbook["Summary_Data"]["O2"].value == "CB LITE NCP"
    assert workbook["Summary_Data"]["Q2"].value == 1
    assert workbook["Summary_Data"]["W2"].value == 8
    assert workbook["Location_Outlet"]["F2"].value == "Test Outlet"


def test_export_keeps_genuine_zero_movement(tmp_path):
    submission = SimpleNamespace(
        report_date=date(2026, 7, 25),
        region="R1",
        dealer="CA1",
        member_no=2,
        location_text="Test",
        gps_latitude=11.5,
        gps_longitude=104.9,
        outlet_name="Zero Outlet",
        outlet_type="Wholesale",
        phone_number="010000000",
        product_metrics=[_metric("CB LITE NCP", 0)],
        competitor_metrics=[],
    )
    output = create_market_export(
        [submission],
        date(2026, 7, 25),
        output_path=tmp_path / "zero.xlsx",
    )
    sheet = load_workbook(output, data_only=True)["Summary_Data"]
    assert sheet["P2"].value == 1
    assert sheet["W2"].value == 0
