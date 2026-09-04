from __future__ import annotations

from datetime import date
import importlib.util
from pathlib import Path
import sys
import tempfile
from types import ModuleType, SimpleNamespace
import unittest

from openpyxl import load_workbook


RAW_AGGREGATES = {
    "AIRA": (0, 0),
    "BACCHUSE": (18, 171),
    "BACCHUSE Sugar Free": (2, 9),
    "Big Cola 3L": (0, 0),
    "Boostrong": (17, 132),
    "CAMBODIA COLA": (2, 15),
    "CAMBODIA ED": (11, 50),
    "CAMBODIA Sport 300mL": (0, 0),
    "CAMBODIA Sport 300ml": (0, 0),
    "CAMBODIA Sport 500mL": (16, 160),
    "CAMBODIA WATER 1500mL": (12, 107),
    "CAMBODIA WATER 500mL": (9, 81),
    "CB BLACK NCP": (0, 0),
    "CB LITE NCP": (14, 112),
    "CB LITE ORD": (10, 96),
    "CB Original NCP": (0, 0),
    "CBC 4.4 NCP": (0, 0),
    "Champion": (16, 158),
    "Coca 1.5L": (2, 20),
    "Coca Cola 330ml": (16, 151),
    "DAZZ": (18, 144),
    "DAZZ Zero Sugar": (1, 3),
    "Dragon": (8, 56),
    "EXPREZ Can 330ml": (19, 185),
    "EXPREZ Melon": (7, 68),
    "GB Original NCP": (1, 5),
    "GB SNOW NCP": (12, 115),
    "GB SNOW ORD": (0, 0),
    "Ganzberg 1500ml": (0, 0),
    "Ganzberg 500ml": (0, 0),
    "Greet LITE NCP": (0, 0),
    "HANUMAN LITE ORD": (5, 37),
    "Hanuman Black NCP": (4, 35),
    "Hanuman LITE NCP": (10, 69),
    "Hitech 1500mL": (2, 20),
    "Hitech 500mL": (4, 35),
    "IZE COLA PET 1.5L All SKUs": (1, 10),
    "IZE PET 300ml Flavour": (2, 20),
    "Idol Can 330ml": (11, 85),
    "King Kong": (6, 56),
    "King Kong Ice": (1, 8),
    "Krud ED": (9, 59),
    "Krud Ice": (0, 0),
    "Krud LITE NCP": (1, 5),
    "Krud LITE ORD": (0, 0),
    "Krud NCP": (0, 0),
    "POP Z Flavour": (0, 0),
    "Pocari Sweat": (1, 7),
    "Provida 1500mL": (2, 20),
    "Provida 500mL": (2, 20),
    "Sting Can 330ml": (8, 46),
    "Super Boostrong": (0, 0),
    "V Cola 350ml": (0, 0),
    "V-Active Sport": (0, 0),
    "Vital 1500mL": (7, 70),
    "Vital 500mL": (7, 68),
    "WURKZ": (21, 206),
    "Wurkz Ice": (2, 6),
}


EXPECTED_FINAL = {
    "CB LITE ORD": 10,
    "GB SNOW ORD": 0,
    "HANUMAN LITE ORD": 5,
    "Krud LITE ORD": 0,
    "CBC 4.4 NCP": 0,
    "CB Original NCP": 0,
    "GB Original NCP": 10,
    "Krud NCP": 0,
    "CB LITE NCP": 9,
    "GB SNOW NCP": 10,
    "Hanuman LITE NCP": 6,
    "Krud LITE NCP": 2,
    "Greet LITE NCP": 0,
    "CB BLACK NCP": 0,
    "Hanuman Black NCP": 10,
    "CAMBODIA COLA": 2,
    "Coca Cola 330ml": 10,
    "WURKZ": 10,
    "Boostrong": 7,
    "Krud ED": 3,
    "Wurkz Ice": 2,
    "Champion": 10,
    "King Kong Ice": 2,
    "Krud Ice": 0,
    "CAMBODIA ED": 8,
    "Super Boostrong": 0,
    "King Kong": 10,
    "AIRA": 0,
    "DAZZ": 8,
    "BACCHUSE": 10,
    "Dragon": 4,
    "DAZZ Zero Sugar": 8,
    "BACCHUSE Sugar Free": 10,
    "IZE PET 300ml Flavour": 10,
    "POP Z Flavour": 0,
    "V Cola 350ml": 0,
    "IZE COLA PET 1.5L All SKUs": 8,
    "Coca 1.5L": 10,
    "Big Cola 3L": 0,
    "EXPREZ Melon": 4,
    "EXPREZ Can 330ml": 10,
    "Sting Can 330ml": 3,
    "Idol Can 330ml": 5,
    "CAMBODIA Sport 500mL": 10,
    "CAMBODIA Sport 300ml": 0,
    "Pocari Sweat": 2,
    "V-Active Sport": 0,
    "CAMBODIA WATER 500mL": 10,
    "Vital 500mL": 9,
    "Provida 500mL": 4,
    "Ganzberg 500ml": 0,
    "Hitech 500mL": 6,
    "CAMBODIA WATER 1500mL": 10,
    "Vital 1500mL": 7,
    "Provida 1500mL": 3,
    "Ganzberg 1500ml": 0,
    "Hitech 1500mL": 3,
}


def _load_report_modules():
    sqlalchemy = ModuleType("sqlalchemy")
    sqlalchemy.text = lambda statement: statement
    database = ModuleType("app.db.database")
    database.SessionLocal = None
    config = ModuleType("app.core.config")
    config.settings = SimpleNamespace(export_path=Path(tempfile.gettempdir()))
    sys.modules["sqlalchemy"] = sqlalchemy
    sys.modules["app.db.database"] = database
    sys.modules["app.core.config"] = config

    aggregator_path = Path("app/reports/aggregator.py")
    aggregator_spec = importlib.util.spec_from_file_location(
        "app.reports.aggregator", aggregator_path
    )
    aggregator = importlib.util.module_from_spec(aggregator_spec)
    sys.modules["app.reports.aggregator"] = aggregator
    assert aggregator_spec.loader is not None
    aggregator_spec.loader.exec_module(aggregator)

    export_path = Path("app/reports/movement_exports.py")
    export_spec = importlib.util.spec_from_file_location(
        "v156_movement_exports", export_path
    )
    movement_exports = importlib.util.module_from_spec(export_spec)
    assert export_spec.loader is not None
    export_spec.loader.exec_module(movement_exports)
    return aggregator, movement_exports


def _scores(positive_count: int, total_points: int) -> list[int]:
    if positive_count <= 0:
        return [0] * 21
    base, remainder = divmod(total_points, positive_count)
    positives = [base + 1] * remainder + [base] * (positive_count - remainder)
    return positives + [0] * (21 - positive_count)


def _survey_submissions(aggregator) -> list[SimpleNamespace]:
    product_scores = {
        product: _scores(positive_count, total_points)
        for product, (positive_count, total_points) in RAW_AGGREGATES.items()
    }
    rows = []
    for index in range(21):
        own_metrics = []
        for product in aggregator.OWN_PRODUCTS:
            score = product_scores.get(product, [0] * 21)[index]
            own_metrics.append(
                SimpleNamespace(
                    product_name=product,
                    movement_score=score,
                    available=score > 0,
                    status="sale" if score > 0 else "no_sale",
                )
            )
        competitor_metrics = []
        for product in aggregator.COMPETITOR_PRODUCTS:
            score = product_scores.get(product, [0] * 21)[index]
            competitor_metrics.append(
                SimpleNamespace(
                    product_name=product,
                    movement_score=score,
                    status="sale" if score > 0 else "no_sale",
                )
            )
        rows.append(
            SimpleNamespace(
                id=index + 1,
                submission_id=f"ksv5-{index + 1}",
                report_date=date(2026, 8, 29),
                submission_time=None,
                report_type="GT",
                region="R2",
                dealer="KSV5",
                group_no=1,
                member_no=1 if index < 11 else 4,
                location_text="ភូមិដូនស, ដូនស",
                outlet_name=f"KSV5 outlet {index + 1}",
                outlet_type="Drink Shop",
                phone_number=None,
                gps_latitude=None,
                gps_longitude=None,
                product_metrics=own_metrics,
                competitor_metrics=competitor_metrics,
                ring_pull_metrics=[],
                key_issue_text=None,
                suggestion_text=None,
                submitter_name=None,
            )
        )
    return rows


class V156AvailabilityMovementTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.aggregator, cls.movement_exports = _load_report_modules()
        cls.rows = _survey_submissions(cls.aggregator)
        cls.result = cls.aggregator.aggregate_submissions(cls.rows, wide_map={})

    def test_report_final_movement_matches_all_kvs5_targets(self):
        actual = {
            product: self.aggregator.final_movement_product_data(
                self.result, product
            ).get("mov")
            for product in EXPECTED_FINAL
        }
        self.assertEqual(actual, EXPECTED_FINAL)

    def test_zero_availability_never_receives_winner_ten(self):
        cb_black = self.aggregator.final_movement_product_data(
            self.result, "CB BLACK NCP"
        )
        hanuman_black = self.aggregator.final_movement_product_data(
            self.result, "Hanuman Black NCP"
        )
        self.assertEqual(cb_black["availability_total"], 0)
        self.assertEqual(cb_black["mov"], 0)
        self.assertEqual(hanuman_black["availability_total"], 4)
        self.assertEqual(hanuman_black["mov"], 10)

    def test_daily_export_uses_the_same_final_movement(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "Market_Survey_Data_2026-08-29.xlsx"
            self.movement_exports.create_daily_export(
                self.rows, date(2026, 8, 29), output_path=output
            )
            workbook = load_workbook(output, read_only=True, data_only=True)
            sheet = workbook["Summary_Data"]
            exported = {
                str(row[14].value): row[22].value
                for row in sheet.iter_rows(min_row=2)
            }
            for product, expected in EXPECTED_FINAL.items():
                if product == "CAMBODIA Sport 300ml":
                    product = "CAMBODIA Sport 300mL"
                self.assertEqual(exported.get(product), expected, product)

            rows_by_product = {
                str(row[14].value): row for row in sheet.iter_rows(min_row=2)
            }
            self.assertEqual(rows_by_product["CB BLACK NCP"][16].value, 0)
            self.assertEqual(rows_by_product["Hanuman Black NCP"][16].value, 4)
            workbook.close()


if __name__ == "__main__":
    unittest.main()
