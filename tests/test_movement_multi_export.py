from datetime import date, datetime
from types import SimpleNamespace

from openpyxl import load_workbook

from app.reports.movement_multi_export import (
    OUTPUT_COLUMNS,
    build_movement_rows,
    create_movement_multi_workbook,
    parse_movement_multi_dates,
)


def _metric(name, score):
    return SimpleNamespace(product_name=name, movement_score=score)


def _submission(report_date, outlet, metrics):
    return SimpleNamespace(
        id=1,
        report_date=report_date,
        submission_time=datetime.combine(report_date, datetime.min.time()),
        region="R1",
        dealer="CA1",
        gps_latitude=11.5564,
        gps_longitude=104.9282,
        outlet_name=outlet,
        outlet_type="Drink Shop",
        phone_number="012345678",
        product_metrics=metrics,
        competitor_metrics=[],
    )


def test_parse_movement_multi_dates_preserves_order_and_removes_duplicates():
    result = parse_movement_multi_dates(
        ["2026-07-04", "2026-07-18", "2026-07-04", "2026-07-25"]
    )
    assert result == [
        date(2026, 7, 4),
        date(2026, 7, 18),
        date(2026, 7, 25),
    ]


def test_build_rows_preserves_real_zero_and_leaves_blank_unanswered():
    rows = build_movement_rows(
        [
            _submission(
                date(2026, 7, 4),
                "Outlet A",
                [
                    _metric("CBC LITE", 0),
                    _metric("GB SNOW", 8),
                    _metric("Unrelated Product", 10),
                ],
            )
        ]
    )
    assert len(rows) == 1
    assert rows[0][8:] == [0, 8, None, None, None]


def test_build_rows_reads_own_and_competitor_metric_tables():
    submission = _submission(
        date(2026, 7, 25),
        "Outlet All Beer",
        [_metric("CB LITE NCP", 6)],
    )
    submission.competitor_metrics = [
        _metric("GB SNOW NCP", 7),
        _metric("Hanuman Lite", 8),
        _metric("Krud Lite", 9),
        _metric("Great Lite", 10),
    ]
    rows = build_movement_rows([submission])
    assert rows[0][8:] == [6, 7, 8, 9, 10]


def test_create_workbook_uses_supplied_template(tmp_path):
    template = (
        __import__("pathlib").Path(__file__).resolve().parents[1]
        / "templates"
        / "detail_movement_beer.xlsx"
    )
    output = tmp_path / "movement.xlsx"
    submissions = [
        _submission(date(2026, 7, 4), "Outlet A", [_metric("CB LITE NCP", 6)]),
        _submission(date(2026, 7, 18), "Outlet B", [_metric("Great Lite", 9)]),
        _submission(date(2026, 7, 25), "No Beer", [_metric("CAMBODIA ED", 10)]),
    ]

    result = create_movement_multi_workbook(
        submissions,
        [date(2026, 7, 4), date(2026, 7, 18), date(2026, 7, 25)],
        template_path=template,
        output_path=output,
    )

    workbook = load_workbook(result, data_only=False)
    worksheet = workbook["Detail_Movement"]
    assert tuple(worksheet.cell(1, column).value for column in range(1, 14)) == OUTPUT_COLUMNS
    assert worksheet.max_row == 3
    assert worksheet["A2"].value.date() == date(2026, 7, 4)
    assert worksheet["I2"].value == 6
    assert worksheet["M3"].value == 9
    table = next(iter(worksheet.tables.values()))
    assert table.ref == "A1:M3"
    assert table.autoFilter.ref == "A1:M3"
    assert worksheet.auto_filter.ref is None
    assert [column.id for column in table.tableColumns] == list(range(1, 14))
