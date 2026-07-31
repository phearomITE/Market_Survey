from datetime import date
from types import SimpleNamespace

from openpyxl import load_workbook

from app.reports.movement_exports import create_movement_export
from app.reports.summary_report import create_summary_report


def _metric(name: str, score: int | None):
    return SimpleNamespace(product_name=name, movement_score=score)


def test_movement_export_writes_every_product_and_preserves_zero(tmp_path):
    row = SimpleNamespace(
        report_date=date(2026, 7, 25),
        region="R1",
        dealer="CA1",
        gps_latitude=11.55,
        gps_longitude=104.91,
        outlet_name="Test Outlet",
        outlet_type="Drink Shop",
        phone_number="012345678",
        product_metrics=[
            _metric("CB LITE NCP", 10),
            _metric("CBC 4.4 NCP", 0),
        ],
        competitor_metrics=[_metric("GB SNOW NCP", 8)],
    )
    output = create_movement_export(
        [row],
        [date(2026, 7, 25)],
        beer_only=False,
        output_path=tmp_path / "raw.xlsx",
    )
    ws = load_workbook(output, data_only=True).active
    headers = [cell.value for cell in ws[1]]
    assert "CB LITE NCP" in headers
    assert "CBC 4.4 NCP" in headers
    assert "GB SNOW NCP" in headers
    values = {headers[index]: ws.cell(2, index + 1).value for index in range(len(headers))}
    assert values["CB LITE NCP"] == 10
    assert values["CBC 4.4 NCP"] == 0
    assert values["GB SNOW NCP"] == 8


def test_gt_summary_uses_uploaded_template_layout(tmp_path):
    rows = [
        {
            "region": "R1",
            "dealer": "CA1",
            "total_submissions": 25,
            "total_outlets": 25,
            "target": None,
            "status": "✅",
        },
        {
            "region": "R1",
            "dealer": "CA8",
            "total_submissions": 0,
            "total_outlets": 0,
            "target": None,
            "status": "❌ No Submit",
        },
    ]
    submission = SimpleNamespace(
        report_date=date(2026, 7, 25),
        region="R1",
        dealer="CA1",
        member_no=7,
        outlet_name="Template Test Outlet",
        phone_number="012345678",
        outlet_type="Drink Shop",
        gps_latitude=11.55,
        gps_longitude=104.91,
        product_metrics=[
            SimpleNamespace(
                product_name="CB LITE NCP",
                movement_score=8,
                stock_status="full",
                bbe_date="07/2027",
            )
        ],
        competitor_metrics=[
            SimpleNamespace(product_name="GB SNOW NCP", movement_score=9)
        ],
    )
    output = create_summary_report(
        rows,
        date(2026, 7, 25),
        output_path=tmp_path / "summary.xlsx",
        report_type="GT",
        submissions=[submission],
    )
    workbook = load_workbook(output, data_only=True)
    assert workbook.sheetnames == ["Summary", "Detail"]
    ws = workbook["Summary"]
    assert ws["A1"].value == "KB Market Survey - GT Region & Dealer Submission Summary"
    assert [ws.cell(8, column).value for column in range(1, 12)] == [
        "Region",
        "Dealer",
        "Member",
        "Total Submissions",
        "Total Outlets",
        "Status",
        "<5",
        "5 to 8",
        "9 to 10",
        "Product Competitor",
        "Movement Lead",
    ]
    assert ws["A4"].value == "Total Regions"
    assert ws["F4"].value == "<5"
    assert ws["I4"].value == "GB SNOW NCP"
    assert ws["C9"].value == 1
    assert ws["H9"].value == 8
    assert ws["J9"].value == "GB SNOW NCP"
    assert ws["K9"].value == 9
    assert ws["A1"].fill.fgColor.rgb.endswith("1F4E78")

    detail = workbook["Detail"]
    assert [detail.cell(1, column).value for column in range(1, 13)] == [
        "Date",
        "Region",
        "Dealer",
        "Outlet Name",
        "Phone Number Outlet",
        "Outlet Type",
        "Stock Status",
        "Freshness Date",
        "0 to 8",
        "Product Competitor",
        "Movement Lead",
        "Link Map",
    ]
    assert detail["D2"].value == "Template Test Outlet"
    assert detail["I2"].value == 8
    assert detail["J2"].value == "GB SNOW NCP"
    assert detail["K2"].value == 9
