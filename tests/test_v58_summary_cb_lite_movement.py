from datetime import date

from openpyxl import load_workbook

from app.reports.summary_report import (
    movement_comparison_from_aggregate,
    create_summary_report,
)


def test_cb_lite_movement_and_competitor_lead_from_final_aggregate():
    aggregate = {
        "products": {"CB LITE NCP": {"mov": 7}},
        "competitors": {
            "GB SNOW NCP": {"mov": 10},
            "Hanuman LITE NCP": {"mov": 8},
            "Krud LITE NCP": {"mov": 4},
            "Greet LITE NCP": {"mov": 2},
        },
    }

    result = movement_comparison_from_aggregate(aggregate)

    assert result["movement_under_5"] is None
    assert result["movement_5_to_8"] == 7
    assert result["movement_9_to_10"] is None
    assert result["competitor_product"] == "GB SNOW NCP"
    assert result["competitor_movement_lead"] == 10


def test_competitor_columns_blank_when_no_competitor_has_10():
    aggregate = {
        "products": {"CB LITE NCP": {"mov": 10}},
        "competitors": {
            "GB SNOW NCP": {"mov": 9},
            "Hanuman LITE NCP": {"mov": 8},
        },
    }

    result = movement_comparison_from_aggregate(aggregate)

    assert result["movement_9_to_10"] == 10
    assert result["competitor_product"] == ""
    assert result["competitor_movement_lead"] is None


def test_summary_workbook_has_new_headers_and_colors(tmp_path):
    path = tmp_path / "summary.xlsx"
    rows = [
        {
            "region": "R1",
            "dealer": "CA3",
            "member": "13",
            "total_submissions": 43,
            "total_outlets": 41,
            "status": "✅",
            "movement_under_5": None,
            "movement_5_to_8": 7,
            "movement_9_to_10": None,
            "competitor_product": "GB SNOW NCP",
            "competitor_movement_lead": 10,
        }
    ]

    create_summary_report(rows, date(2026, 7, 18), path)
    workbook = load_workbook(path)
    sheet = workbook["Summary"]

    assert [sheet.cell(8, col).value for col in range(1, 12)] == [
        "Region",
        "Dealer",
        "Member",
        "Total Submissions",
        "Total Outlets",
        "Status",
        "<5",
        "5 to 8",
        "9 to 10",
        "Product Competitor",
        "Movement Lead",
    ]
    assert sheet["H9"].value == 7
    assert sheet["J9"].value == "GB SNOW NCP"
    assert sheet["K9"].value == 10
    assert sheet["G8"].fill.fgColor.rgb.endswith("C00000")
    assert sheet["H8"].fill.fgColor.rgb.endswith("FFC000")
    assert sheet["I8"].fill.fgColor.rgb.endswith("00B050")
