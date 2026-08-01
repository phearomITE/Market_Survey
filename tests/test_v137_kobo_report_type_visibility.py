from __future__ import annotations

from pathlib import Path
import unittest

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
FORM_PATH = ROOT / "templates" / "KB_Market_Improvement_XLSForm_GT_HORECA.xlsx"


class V137KoboReportTypeVisibilityTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.workbook = load_workbook(FORM_PATH, read_only=False, data_only=False)
        cls.sheet = cls.workbook["survey"]
        cls.rows = {
            cls.sheet.cell(row, 2).value: row
            for row in range(2, cls.sheet.max_row + 1)
            if cls.sheet.cell(row, 2).value
        }

    @classmethod
    def tearDownClass(cls):
        cls.workbook.close()

    def relevant(self, name: str) -> str:
        return str(self.sheet.cell(self.rows[name], 8).value or "")

    def test_initial_screen_contains_only_region_dealer_and_report_type(self):
        for name in ("region", "dealer", "final_summary_report_type"):
            self.assertEqual(self.relevant(name), "")
            self.assertEqual(self.sheet.cell(self.rows[name], 4).value, "yes")

        for name in (
            "report_date",
            "group_no",
            "member_no",
            "outlet_name",
            "gps_location",
            "location_of_visit_text",
            "phone_number",
            "outlet_type",
            "is_new_outlet",
            "submitter_name",
            "product_check_by_type",
            "ring_pull_group",
            "key_issues_suggestion_group",
        ):
            route = self.relevant(name)
            self.assertIn("${final_summary_report_type}", route, name)
            self.assertIn("= 'gt'", route, name)
            self.assertIn("= 'horeca'", route, name)

    def test_required_fields_remain_required_after_selection(self):
        for name in ("region", "dealer", "final_summary_report_type", "report_date", "outlet_type"):
            self.assertEqual(self.sheet.cell(self.rows[name], 4).value, "yes")

    def test_gt_horeca_and_shared_products_are_routed(self):
        self.assertEqual(
            self.relevant("fresh_cb_pint_group"),
            "${final_summary_report_type} = 'horeca'",
        )
        self.assertEqual(
            self.relevant("fresh_wurkz_ice_group"),
            "${final_summary_report_type} = 'gt'",
        )
        for name in (
            "fresh_cb_lite_ord_group",
            "fresh_cb_lite_ncp_group",
            "comp_super_boostrong_group",
            "comp_king_kong_group",
            "comp_aira_group",
            "fresh_cambodia_water_500_group",
        ):
            route = self.relevant(name)
            self.assertIn("= 'gt'", route, name)
            self.assertIn("= 'horeca'", route, name)

    def test_pint_own_products_keep_only_four_approved_fields(self):
        prefixes = (
            "fresh_cb_pint",
            "fresh_cbl_pint",
            "fresh_cb_supeeme_pint",
            "fresh_cb_black_pint",
        )
        names = [name for name in self.rows if name.startswith(prefixes)]
        self.assertFalse(any("buy_in" in name for name in names))
        self.assertFalse(any("sell_out" in name for name in names))
        self.assertFalse(any("ring_pull" in name for name in names))


if __name__ == "__main__":
    unittest.main()
