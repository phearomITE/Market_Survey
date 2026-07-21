from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from types import SimpleNamespace as NS

from openpyxl import load_workbook

from app.reports.data_export import create_data_export
from app.reports.summary_report import build_summary_rows


def _product_metric(name: str, movement: int, available: bool = True):
    return NS(
        product_name=name,
        status="sale" if available else "no_sale",
        available=available,
        movement_score=movement,
        stock_status=None,
        bbe_date=None,
        buy_in_price=None,
        sell_out_price=None,
        ring_pull_value=None,
        new_outlet_purchase=False,
        volume_ctn=None,
    )


def _competitor_metric(name: str, movement: int, available: bool = True):
    return NS(
        product_name=name,
        status="sale" if available else "no_sale",
        movement_score=movement,
        stock_status=None,
        buy_in_price=None,
        sell_out_price=None,
    )


def _submission(index: int, outlet_type: str):
    return NS(
        submission_id=f"v62-{index}",
        submission_time=datetime(2026, 7, 18, 8, index),
        report_date=date(2026, 7, 18),
        region="R1",
        dealer="CA1",
        group_no=2,
        member_no=7,
        total_outlet_visit_target=None,
        outlet_name=f"Outlet {index}",
        outlet_type=outlet_type,
        is_new_outlet=False,
        submitter_name="Tester",
        phone_number="0",
        location_text="Phnom Penh",
        gps_text=None,
        gps_latitude=11.5,
        gps_longitude=104.9,
        key_issue_text=None,
        suggestion_text=None,
        product_metrics=[
            _product_metric("CBC 4.4 NCP", 8, True),
            _product_metric("CB LITE NCP", 7, True),
        ],
        competitor_metrics=[
            _competitor_metric("GB Original NCP", 10, True),
            _competitor_metric("GB SNOW NCP", 10, True),
            _competitor_metric("Hanuman LITE NCP", 8, True),
            _competitor_metric("Krud LITE NCP", 4, True),
            _competitor_metric("Greet LITE NCP", 2, True),
        ],
        ring_pull_metrics=[],
    )


def test_export_product_counts_and_summary_movement_match(tmp_path: Path):
    outlet_types = [
        "Wholesale",
        "Drink Shop",
        "Wet Market",
        "Trolley",
        "Local Eat",
        "Coffee,Bakery",
        "Motor Shop",
    ]
    submissions = [_submission(i + 1, outlet_type) for i, outlet_type in enumerate(outlet_types)]

    template = Path(__file__).parents[1] / "templates" / "Template_Data_Survey.xlsx"
    output = tmp_path / "export.xlsx"
    create_data_export(
        submissions,
        date(2026, 7, 18),
        template_path=template,
        output_path=output,
    )

    wb = load_workbook(output, data_only=True)
    ws = wb["Summary_Data"]
    headers = {cell.value: cell.column for cell in ws[1] if cell.value}
    cbc_row = next(row for row in range(2, ws.max_row + 1) if ws.cell(row, headers["Product"]).value == "CBC 4.4 NCP")

    assert ws.cell(cbc_row, headers["Total Outlets"]).value == 7
    assert ws.cell(cbc_row, headers["WS"]).value == 1
    assert ws.cell(cbc_row, headers["DS"]).value == 1
    assert ws.cell(cbc_row, headers["WM"]).value == 1
    assert ws.cell(cbc_row, headers["TL"]).value == 1
    assert ws.cell(cbc_row, headers["LE"]).value == 1
    assert ws.cell(cbc_row, headers["CB"]).value == 1
    assert ws.cell(cbc_row, headers["MS"]).value == 1
    assert ws.cell(cbc_row, headers["Movement"]).value == 8

    summary_rows = build_summary_rows(submissions)
    ca1 = next(row for row in summary_rows if row["dealer"] == "CA1")
    assert ca1["member"] == "7"
    assert ca1["movement_5_to_8"] == 7
    assert ca1["competitor_product"] == "GB SNOW NCP"
    assert ca1["competitor_movement_lead"] == 10
