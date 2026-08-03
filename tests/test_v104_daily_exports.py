from datetime import date
from pathlib import Path
from types import ModuleType, SimpleNamespace
import importlib.util
import sys
import tempfile
import unittest

from openpyxl import load_workbook


def _load_exports_module():
    config = ModuleType("app.core.config")
    config.settings = SimpleNamespace(export_path=Path(tempfile.gettempdir()))
    aggregator = ModuleType("app.reports.aggregator")
    aggregator.OWN_PRODUCTS = ["OWN"]
    aggregator.COMPETITOR_PRODUCTS = ["COMP"]
    aggregator.HORECA_OWN_PRODUCTS = ["HOWN"]
    aggregator.HORECA_COMPETITOR_PRODUCTS = ["HCOMP"]
    aggregator.ALL_OWN_PRODUCTS = ["OWN", "HOWN"]
    aggregator.ALL_COMPETITOR_PRODUCTS = ["COMP", "HCOMP"]
    aggregator.is_final_summary_outlet_name = (
        lambda value: str(value or "").replace(" ", "").strip() == "បូកសរុបរួម"
    )
    aggregator.load_wide_payloads = lambda rows: {}
    aggregator.aggregate_submissions = lambda rows, wide_map=None: {
        "region": "R1",
        "dealer": "D1",
        "location_text": "Phnom Penh",
        "total_outlets": len(rows),
        "outlet_types": {"Drink Shop": len(rows)},
        "products": {"OWN": {"availability": {"Drink Shop": len(rows)}, "mov": 10}},
        "competitors": {"COMP": {"mov": 8}},
    }
    sys.modules["app.core.config"] = config
    sys.modules["app.reports.aggregator"] = aggregator
    path = Path("app/reports/movement_exports.py")
    spec = importlib.util.spec_from_file_location("v104_movement_exports", path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


class V104DailyExportTests(unittest.TestCase):
    def setUp(self):
        self.exports = _load_exports_module()
        self.submission = SimpleNamespace(
            report_date=date(2026, 7, 25),
            region="R1",
            dealer="D1",
            member_no=1,
            location_text="Phnom Penh",
            gps_latitude=11.56,
            gps_longitude=104.92,
            outlet_name="Outlet 1",
            outlet_type="Drink Shop",
            phone_number="012345678",
            report_type="GT",
            product_metrics=[SimpleNamespace(product_name="OWN", movement_score=7)],
            competitor_metrics=[SimpleNamespace(product_name="COMP", movement_score=6)],
        )

    def test_export_has_exact_two_sheets_and_headers(self):
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "daily.xlsx"
            self.exports.create_daily_export([self.submission], date(2026, 7, 25), output)
            wb = load_workbook(output, read_only=True)
            self.assertEqual(wb.sheetnames, ["Summary_Data", "Location_Outlet"])
            self.assertEqual(
                [cell.value for cell in next(wb["Summary_Data"].iter_rows(max_row=1))],
                self.exports.SUMMARY_HEADERS,
            )
            self.assertEqual(
                [cell.value for cell in next(wb["Location_Outlet"].iter_rows(max_row=1))],
                self.exports.BASE_HEADERS,
            )
            summary = list(wb["Summary_Data"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
            self.assertEqual(summary[-1], 10)
            wb.close()

    def test_daily_export_uses_member_mode_and_excludes_summary_location(self):
        member_values = [4, 6, 7, 5, 4, 4, 4, 4, 5, 4, 4, 4, 5, 5, 6]
        submissions = []
        for index, member_value in enumerate(member_values, start=1):
            submission = SimpleNamespace(**vars(self.submission))
            submission.member_no = member_value
            submission.outlet_name = f"Outlet {index}"
            submissions.append(submission)

        final_summary = SimpleNamespace(**vars(self.submission))
        final_summary.member_no = 9
        final_summary.outlet_name = "បូកសរុបរួម"
        final_summary.outlet_type = None
        final_summary.phone_number = None
        final_summary.gps_latitude = None
        final_summary.gps_longitude = None
        submissions.append(final_summary)

        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "daily_clean_location.xlsx"
            self.exports.create_daily_export(
                submissions,
                date(2026, 8, 1),
                output,
            )
            wb = load_workbook(output, read_only=True, data_only=True)

            member_values_in_export = {
                row[3]
                for row in wb["Summary_Data"].iter_rows(
                    min_row=2,
                    values_only=True,
                )
            }
            self.assertEqual(member_values_in_export, {4})

            exported_outlet_names = {
                row[5]
                for row in wb["Location_Outlet"].iter_rows(
                    min_row=2,
                    values_only=True,
                )
            }
            self.assertNotIn("បូកសរុបរួម", exported_outlet_names)
            self.assertEqual(len(exported_outlet_names), len(member_values))
            wb.close()

    def test_raw_export_has_six_columns_including_outlet_type(self):
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "raw.xlsx"
            self.exports.create_raw_movement_long_export([self.submission], date(2026, 7, 25), output)
            wb = load_workbook(output, read_only=True)
            ws = wb["Raw_Movement"]
            first_row = [cell.value for cell in next(ws.iter_rows(max_row=1))]
            self.assertEqual(first_row, self.exports.RAW_HEADERS)
            self.assertEqual(len(first_row), 6)
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            own_row = next(row for row in data_rows if row[4] == "OWN")
            self.assertEqual(own_row[3], "Drink Shop")
            self.assertEqual(own_row[5], 7)
            wb.close()

    def test_raw_export_preserves_zero_and_skips_summary(self):
        zero_submission = SimpleNamespace(**vars(self.submission))
        zero_submission.outlet_name = "Outlet Zero"
        zero_submission.product_metrics = [
            SimpleNamespace(product_name="OWN", movement_score=0)
        ]
        zero_submission.competitor_metrics = []

        summary_submission = SimpleNamespace(**vars(self.submission))
        summary_submission.outlet_name = "បូកសរុបរួម"
        summary_submission.product_metrics = [
            SimpleNamespace(product_name="OWN", movement_score=10)
        ]

        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "raw_baseline.xlsx"
            self.exports.create_raw_movement_long_export(
                [zero_submission, summary_submission],
                date(2026, 7, 25),
                output,
            )
            wb = load_workbook(output, read_only=True)
            rows = list(
                wb["Raw_Movement"].iter_rows(min_row=2, values_only=True)
            )
            self.assertTrue(rows)
            self.assertTrue(all(row[5] == 0 for row in rows))
            self.assertTrue(all(row[5] != 10 for row in rows))
            wb.close()

    def test_horeca_raw_export_contains_only_horeca_product_set(self):
        horeca = SimpleNamespace(**vars(self.submission))
        horeca.report_type = "HORECA"
        horeca.outlet_type = "Local Eat"
        horeca.product_metrics = [
            SimpleNamespace(product_name="HOWN", movement_score=4)
        ]
        horeca.competitor_metrics = [
            SimpleNamespace(product_name="HCOMP", movement_score=6)
        ]

        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "raw_horeca.xlsx"
            self.exports.create_raw_movement_long_export(
                [horeca], date(2026, 7, 25), output
            )
            wb = load_workbook(output, read_only=True)
            products = {
                row[4]
                for row in wb["Raw_Movement"].iter_rows(
                    min_row=2, values_only=True
                )
            }
            self.assertEqual(products, {"HOWN", "HCOMP"})
            self.assertTrue(all(
                row[3] == "Local Eat"
                for row in wb["Raw_Movement"].iter_rows(
                    min_row=2, values_only=True
                )
            ))
            wb.close()


if __name__ == "__main__":
    unittest.main()
