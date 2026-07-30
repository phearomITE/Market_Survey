from __future__ import annotations

from datetime import date
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from app.kobo.parser import normalize_submission
from app.kobo.sync import (
    SYNC_SCHEMA_VERSION,
    _competitor_metrics_from_flat,
    _product_metrics_from_flat,
    _source_hash,
)
from app.reports.aggregator import (
    GT_OWN_PRODUCTS,
    HORECA_COMPETITOR_PRODUCTS,
    HORECA_OFFTAKE_COMPARE_GROUPS,
    HORECA_OWN_PRODUCTS,
    competitor_field,
    product_field,
)
from app.reports.excel_report import (
    _template_for_aggs,
    create_report_workbook,
)
from app.services.report_service import (
    _filter_by_report_type,
    parse_report_command_args,
    parse_summary_command_args,
)


PROJECT_ROOT = Path(__file__).resolve().parents[1]


def _metric_row(rows: list[dict], product: str) -> dict:
    return next(row for row in rows if row["product_name"] == product)


def test_requested_report_command_formats():
    assert parse_report_command_args(["CA3", "HORECA", "2026-07-18"]) == (
        "CA3",
        "2026-07-18",
        "HORECA",
    )
    assert parse_report_command_args(["CA3", "GT", "2026-07-18"]) == (
        "CA3",
        "2026-07-18",
        "GT",
    )
    assert parse_report_command_args(["CA3", "2026-07-18"]) == (
        "CA3",
        "2026-07-18",
        "GT",
    )


def test_requested_summary_command_formats():
    assert parse_summary_command_args(["GT", "2026-07-25"]) == (
        "GT",
        "2026-07-25",
    )
    assert parse_summary_command_args(["HORECA", "2026-07-25"]) == (
        "HORECA",
        "2026-07-25",
    )
    assert parse_summary_command_args(["2026-07-25"]) == (
        "GT",
        "2026-07-25",
    )


def test_new_form_report_type_and_local_drink_are_normalized():
    normalized = normalize_submission(
        {
            "_id": 9001,
            "_submission_time": "2026-07-18T09:30:00",
            "outlet_info": {
                "dealer": "ca3",
                "region": "r1",
                "report_date": "2026-07-18",
                "final_summary_report_type": "horeca",
                "outlet_type": "local_drink",
                "outlet_name": "HORECA Test Outlet",
            },
        }
    )
    assert normalized["dealer"] == "CA3"
    assert normalized["region"] == "R1"
    assert normalized["report_type"] == "HORECA"
    assert normalized["outlet_type"] == "Local Drink"


def test_report_type_filter_prefers_explicit_selector_and_supports_old_rows():
    rows = [
        SimpleNamespace(report_type="GT", outlet_type="Wholesale"),
        SimpleNamespace(report_type="HORECA", outlet_type="Local Eat"),
        SimpleNamespace(report_type=None, outlet_type="Local Drink"),
        SimpleNamespace(report_type=None, outlet_type="Drink Shop"),
    ]
    assert len(_filter_by_report_type(rows, "GT")) == 2
    assert len(_filter_by_report_type(rows, "HORECA")) == 2


def test_horeca_form_field_mappings_and_no_pint_prices():
    flat = {
        "fresh_status_cb_pint": "sale",
        "fresh_movement_score_cb_pint": 8,
        "fresh_stock_status_cb_pint": "full",
        "fresh_bbe_cb_pint": "05.2027",
        "comp_status_tiger": "fast_sale",
        "comp_movement_score_tiger": 9,
        "fresh_status_cb_supeeme_pint": "sale",
        "fresh_movement_score_cb_supeeme_pint": 7,
    }

    own_rows = _product_metrics_from_flat(flat)
    competitor_rows = _competitor_metrics_from_flat(flat)

    cb_pint = _metric_row(own_rows, "CB Pint")
    assert cb_pint["movement_score"] == 8
    assert cb_pint["stock_status"] == "full"
    assert cb_pint["bbe_date"] == "05.2027"
    assert cb_pint["buy_in_price"] is None
    assert cb_pint["sell_out_price"] is None
    assert cb_pint["ring_pull_value"] is None

    tiger = _metric_row(competitor_rows, "Tiger Pint")
    assert tiger["movement_score"] == 9

    # CB SUPEEME Pint is an own product but is also displayed in a comparison
    # column in the locked HORECA template.
    supeeme_compare = _metric_row(competitor_rows, "CB SUPEEME Pint")
    assert supeeme_compare["movement_score"] == 7

    assert "fresh_movement_score_cb_pint" in product_field("CB Pint", "mov")
    assert "comp_movement_score_tiger" in competitor_field("Tiger Pint", "mov")
    assert "fresh_movement_score_cb_supeeme_pint" in competitor_field(
        "CB SUPEEME Pint", "mov"
    )


def test_horeca_product_sets_match_locked_template_shape():
    assert len(GT_OWN_PRODUCTS) == 18
    assert len(HORECA_OWN_PRODUCTS) == 18
    assert len(HORECA_OFFTAKE_COMPARE_GROUPS) == 15
    assert "Tiger Crystal Pint" in HORECA_COMPETITOR_PRODUCTS
    assert "EXPREZ Melon" in HORECA_OWN_PRODUCTS
    assert "V Cola 330ml" in HORECA_COMPETITOR_PRODUCTS


def test_horeca_template_is_selected_and_layout_is_not_resized(tmp_path: Path):
    template = _template_for_aggs([{"report_type": "HORECA"}])
    assert template.name == "template_horeca.xlsx"
    assert template.exists()

    agg = {
        "dealer": "CA3",
        "region": "R1",
        "report_date": date(2026, 7, 18),
        "report_type": "HORECA",
        "channel": "HORECA",
        "total_outlets": 1,
        "outlet_types": {"Local Drink": 1},
        "group_no": 2,
        "member_no": 1,
        "location_text": "Phnom Penh",
        "products": {
            product: {
                "bbe": "05.2027" if product == "CB Pint" else None,
                "mov": 8 if product == "CB Pint" else None,
                "stock": "គ្រប់" if product == "CB Pint" else None,
                "buy_in": None,
                "sell_out": None,
                "ring_pull": None,
                "new_purchase": 0,
                "volume": None,
                "availability": {"Local Drink": 1} if product == "CB Pint" else {},
            }
            for product in HORECA_OWN_PRODUCTS
        },
        "competitors": {
            product: {
                "mov": 9 if product == "Tiger Pint" else None,
                "stock": None,
                "buy_in": None,
                "sell_out": None,
            }
            for product in HORECA_COMPETITOR_PRODUCTS
        },
        "ring_pull": {
            "CBL NCP 6 Can": {"total_outlets": 0, "qty": 0},
            "CBL NCP 5 USD": {"total_outlets": 0, "qty": 0},
        },
        "key_issues": ["Issue 1"],
        "suggestions": ["Action 1"],
    }

    output = tmp_path / "horeca.xlsx"
    create_report_workbook([agg], output)
    wb = load_workbook(output)
    ws = wb.active

    assert ws.max_row == 47
    assert ws.max_column == 27
    assert ws["A2"].value == "Market Improvement Report HORECA"
    assert ws["A42"].value == "Ring Pull In Outlets"
    assert str(ws["H6"].value).startswith("Local Eat:")
    assert str(ws["Z6"].value).startswith("Local Drink:")
    assert ws["AA6"].value == "Volume"
    assert ws["B27"].value == "CB Pint"
    assert ws["H27"].value == "Angkor Pint"
    assert ws["M27"].value == "Tiger Pint"
    assert ws["C27"].value == 8
    assert ws["N27"].value == 9
    assert ws["Z7"].value == 1
    assert ws.print_area == "'CA3'!$A$1:$AA$47"


def test_sync_hash_version_forces_one_metric_rebuild():
    raw = {"_id": 1, "outlet_info": {"dealer": "CA3"}}
    assert SYNC_SCHEMA_VERSION == "v84_gt_horeca_reports_1"
    assert _source_hash(raw) == _source_hash(raw)
    assert _source_hash(raw) != _source_hash({**raw, "changed": True})
