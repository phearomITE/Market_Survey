from __future__ import annotations

from datetime import date
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook

from app.data.dealers import ALL_DEALERS
from app.reports.excel_report import SUMMARY_FONT_NAME, _normalize_khmer_cells
from app.services import report_service


def _submission(dealer: str, outlet_name: str) -> SimpleNamespace:
    return SimpleNamespace(dealer=dealer, outlet_name=outlet_name)


def test_official_dealer_list_contains_exactly_65_unique_dealers() -> None:
    assert len(ALL_DEALERS) == 65
    assert len(set(ALL_DEALERS)) == 65


def test_status_export_counts_only_summary_marker(monkeypatch, tmp_path: Path) -> None:
    rows = [
        _submission("CA1", "បូកសរុបរួម"),
        _submission("CA8", "Normal Outlet"),
        _submission("CA3", "prefix បូក\u200bសរុបរួម suffix"),
        _submission("NOT-OFFICIAL", "បូកសរុបរួម"),
    ]
    monkeypatch.setattr(
        report_service,
        "fetch_report_submissions_fast",
        lambda dealer, report_date: rows,
    )
    monkeypatch.setattr(report_service.settings, "export_dir", str(tmp_path))

    path, message = report_service.generate_summary_status_export("2026-08-01")
    workbook = load_workbook(path, data_only=True)
    sheet = workbook["Summary Status"]

    assert sheet.max_row == 66
    assert [cell.value for cell in sheet[1]] == ["Date", "Region", "Dealer", "Status"]
    statuses = {sheet.cell(row, 3).value: sheet.cell(row, 4).value for row in range(2, 67)}
    assert statuses["CA1"] == "Submitted Summary"
    assert statuses["CA3"] == "Submitted Summary"
    assert statuses["CA8"] == "Missing Summary"
    assert set(statuses) == set(ALL_DEALERS)
    assert "2 submitted, 63 missing, 65 dealers checked" in message


def test_status_export_marks_all_65_missing_when_date_has_no_rows(
    monkeypatch, tmp_path: Path
) -> None:
    monkeypatch.setattr(
        report_service,
        "fetch_report_submissions_fast",
        lambda dealer, report_date: [],
    )
    monkeypatch.setattr(report_service.settings, "export_dir", str(tmp_path))

    path, message = report_service.generate_summary_status_export("2026-08-01")
    sheet = load_workbook(path, data_only=True)["Summary Status"]
    assert sheet.max_row == 66
    assert all(sheet.cell(row, 4).value == "Missing Summary" for row in range(2, 67))
    assert "0 submitted, 65 missing, 65 dealers checked" in message


def test_khmer_clusters_are_normalized_without_hidden_separators() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "គ្រ\u200bប់"

    _normalize_khmer_cells(sheet)

    assert sheet["A1"].value == "គ្រប់"
    assert sheet["A1"].font.name == SUMMARY_FONT_NAME
    assert sheet["A1"].font.scheme is None


def test_map_routes_and_cambodia_bounds_are_present() -> None:
    from app.main import app
    from app.web import router as map_router

    paths = {route.path for route in app.routes}
    assert "/map" in paths
    assert "/api/map/data" in paths
    assert map_router.CAMBODIA_LAT_MIN >= 9
    assert map_router.CAMBODIA_LAT_MAX <= 16
    assert map_router.CAMBODIA_LON_MIN >= 101
    assert map_router.CAMBODIA_LON_MAX <= 109

