from datetime import date, datetime
from types import SimpleNamespace

from openpyxl import load_workbook

from app.reports.aggregator import aggregate_submissions
from app.reports.summary_status import (
    STATUS_MISSING,
    STATUS_SUBMITTED,
    create_summary_status_export,
)


def _submission(dealer, outlet_name, *, idx=1, fall="", issue="", idea=""):
    return SimpleNamespace(
        id=idx,
        submission_id=str(idx),
        submission_time=datetime(2026, 8, 22, 16, idx),
        dealer=dealer,
        region="R1",
        report_date=date(2026, 8, 22),
        report_type="GT",
        group_no=3,
        member_no=11,
        location_text="Phnom Penh",
        outlet_name=outlet_name,
        outlet_type="Drink Shop",
        total_outlet_visit_target=29,
        submitter_name=fall,
        key_issue_text=issue,
        suggestion_text=idea,
        product_metrics=[],
        competitor_metrics=[],
        ring_pull_metrics=[],
    )


def test_prefixed_ca3_summary_populates_all_report_narratives():
    rows = [
        _submission("CA3", "Regular Outlet", idx=1),
        _submission(
            "CA3",
            "# ចែ ម៉ៅ , បូកសរុបរួម",
            idx=2,
            fall="1. ចំណុចដួល CA3",
            issue="1. Key issue CA3",
            idea="1. Initiative CA3",
        ),
    ]
    result = aggregate_submissions(rows, wide_map={})
    assert result["total_outlets"] == 1
    assert result["fall_points"][0] == "ចំណុចដួល CA3"
    assert result["key_issues"][0] == "Key issue CA3"
    assert result["suggestions"][0] == "Initiative CA3"


def test_status_export_uses_same_matcher_and_khmer_labels(tmp_path):
    rows = [
        _submission("CA1", "Regular Outlet", idx=1),
        _submission("CA3", "\u200b' បូក សរុប រួម ' ", idx=2),
    ]
    path = create_summary_status_export(
        rows, date(2026, 8, 22), tmp_path / "status.xlsx"
    )
    workbook = load_workbook(path, data_only=True)
    sheet = workbook["Summary Status"]
    status_by_dealer = {
        sheet.cell(row, 3).value: sheet.cell(row, 4).value
        for row in range(2, sheet.max_row + 1)
    }
    assert status_by_dealer["CA3"] == STATUS_SUBMITTED
    assert status_by_dealer["CA1"] == STATUS_MISSING
    assert "Submitted Summary" not in status_by_dealer.values()
    assert "Missing Summary" not in status_by_dealer.values()


def test_general_template_contains_no_compromise_and_prints_it():
    from pathlib import Path
    from app.reports.excel_report import _layout_rows

    path = Path(__file__).resolve().parents[1] / "templates" / "template_general.xlsx"
    workbook = load_workbook(path, data_only=False)
    sheet = workbook[workbook.sheetnames[0]]
    layout = _layout_rows(sheet, {})
    values = "\n".join(
        str(sheet.cell(row, 1).value or "") for row in range(51, 56)
    )
    assert "No Compromise" in values
    assert "Mass Products" in values
    assert "ផលិតផលហួសដឺឡេ" in values
    assert "ការបញ្ចូលរបាយការណ៍លក់មិនត្រឹមត្រូវ" in values
    assert "របាយការណ៍លក់ជូនអតិថិជនមិនពិត" in values
    assert layout["print_end"] >= 55
