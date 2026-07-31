from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from app.reports.excel_report import (
    SUMMARY_FONT_NAME,
    _apply_khmer_safe_fonts,
    _contains_khmer,
)


def test_detects_khmer_text_only():
    assert _contains_khmer("គ្រប់")
    assert _contains_khmer("Stock: គ្រប់")
    assert not _contains_khmer("full")
    assert not _contains_khmer(None)


def test_khmer_font_fix_preserves_existing_style(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "គ្រប់"
    sheet["A1"].font = Font(
        name="Calibri",
        size=18,
        bold=True,
        italic=True,
        color="A00000",
        scheme="minor",
    )
    sheet["A2"] = "English"
    sheet["A2"].font = Font(name="Calibri", size=12)

    _apply_khmer_safe_fonts(sheet)

    output = tmp_path / "khmer-font-check.xlsx"
    workbook.save(output)
    saved_sheet = load_workbook(output).active

    assert saved_sheet["A1"].font.name == SUMMARY_FONT_NAME
    assert saved_sheet["A1"].font.scheme is None
    assert saved_sheet["A1"].font.sz == 18
    assert saved_sheet["A1"].font.bold is True
    assert saved_sheet["A1"].font.italic is True
    assert saved_sheet["A1"].font.color.rgb == "00A00000"
    assert saved_sheet["A2"].font.name == "Calibri"

