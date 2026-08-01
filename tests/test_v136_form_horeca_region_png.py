from __future__ import annotations

import ast
from pathlib import Path
import unittest

from openpyxl import load_workbook

from app.kobo.parser import normalize_submission


ROOT = Path(__file__).resolve().parents[1]


def _literal_assignment(path: Path, variable: str):
    tree = ast.parse(path.read_text(encoding="utf-8"))
    for node in tree.body:
        if isinstance(node, ast.Assign):
            if any(isinstance(target, ast.Name) and target.id == variable for target in node.targets):
                return ast.literal_eval(node.value)
    raise AssertionError(f"Missing assignment: {variable}")


class V136FormHorecaRegionPngTests(unittest.TestCase):
    def test_nested_region_and_dealer_survive_blank_duplicate_leaf(self):
        row = {
            "region": "",
            "dealer": "",
            "outlet_info": {
                "region": "r5",
                "dealer": "bmc2",
                "report_date": "2026-08-01",
            },
            "_id": 101,
        }
        normalized = normalize_submission(row)
        self.assertEqual(normalized["region"], "R5")
        self.assertEqual(normalized["dealer"], "BMC2")

    def test_raw_export_writes_region_and_dealer(self):
        source = (ROOT / "app" / "reports" / "movement_exports.py").read_text(encoding="utf-8")
        start = source.index("def create_raw_movement_long_export")
        end = source.index("def _ordered_products", start)
        body = source[start:end]
        self.assertIn('getattr(submission, "region", None)', body)
        self.assertIn('getattr(submission, "dealer", None)', body)

    def test_horeca_products_include_shared_cb_lite_ncp_group(self):
        path = ROOT / "app" / "reports" / "aggregator.py"
        own = _literal_assignment(path, "HORECA_OWN_PRODUCTS")
        competitors = _literal_assignment(path, "HORECA_COMPETITOR_PRODUCTS")
        groups = _literal_assignment(path, "OFFTAKE_COMPARE_GROUPS")
        self.assertIn("CB LITE NCP", own)
        self.assertIn("GB SNOW NCP", competitors)
        self.assertIn("Hanuman LITE NCP", competitors)
        self.assertIn("Krud LITE NCP", competitors)
        self.assertIn("Greet LITE NCP", competitors)
        self.assertIn(
            ["CB LITE NCP", "GB SNOW NCP", "Hanuman LITE NCP", "Krud LITE NCP", "Greet LITE NCP"],
            groups,
        )

    def test_kobo_required_routing_and_pint_fields(self):
        path = ROOT / "templates" / "KB_Market_Improvement_XLSForm_GT_HORECA.xlsx"
        wb = load_workbook(path, read_only=False, data_only=False)
        ws = wb["survey"]
        rows = {ws.cell(row, 2).value: row for row in range(2, ws.max_row + 1) if ws.cell(row, 2).value}
        for name in ("region", "dealer", "final_summary_report_type", "report_date", "outlet_type"):
            self.assertEqual(ws.cell(rows[name], 4).value, "yes")

        self.assertIsNone(ws.cell(rows["fresh_cb_lite_ncp_group"], 8).value)
        blank_horeca = ws.cell(rows["fresh_cb_pint_group"], 8).value
        blank_gt = ws.cell(rows["fresh_wurkz_ice_group"], 8).value
        self.assertIn("string-length", blank_horeca)
        self.assertIn("= 'horeca'", blank_horeca)
        self.assertIn("string-length", blank_gt)
        self.assertIn("= 'gt'", blank_gt)

        own_pint_prefixes = (
            "fresh_cb_pint", "fresh_cbl_pint", "fresh_cb_supeeme_pint", "fresh_cb_black_pint",
        )
        pint_names = [name for name in rows if name.startswith(own_pint_prefixes)]
        self.assertFalse(any("buy_in" in name or "sell_out" in name or "ring_pull" in name for name in pint_names))
        wb.close()

    def test_horeca_template_contains_cb_lite_and_ed_comparison(self):
        path = ROOT / "templates" / "template_horeca_products.xlsx"
        wb = load_workbook(path, read_only=False, data_only=True)
        ws = wb["Simaple_form"]
        self.assertEqual(ws["A25"].value, 19)
        self.assertEqual(ws["B25"].value, "CB LITE NCP")
        self.assertEqual(ws["H36"].value, "Super Boostrong")
        self.assertEqual(ws["M36"].value, "King Kong")
        self.assertEqual(ws["R36"].value, "AIRA")
        wb.close()


if __name__ == "__main__":
    unittest.main()
