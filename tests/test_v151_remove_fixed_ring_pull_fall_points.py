from datetime import date, datetime
from pathlib import Path
from types import ModuleType, SimpleNamespace
import sys
import unittest

from openpyxl import load_workbook

# Keep this regression test runnable in a lightweight source checkout where the
# production database/settings packages have not been installed yet.
sqlalchemy = ModuleType("sqlalchemy")
sqlalchemy.text = lambda statement: statement
database = ModuleType("app.db.database")
database.SessionLocal = None
config = ModuleType("app.core.config")
config.settings = SimpleNamespace(
    template_file=Path("templates/template_by_dealer.xlsx"),
    horeca_template_file=Path("templates/template_horeca_products.xlsx"),
    export_path=Path("exports"),
)
sys.modules.setdefault("sqlalchemy", sqlalchemy)
sys.modules.setdefault("app.db.database", database)
sys.modules.setdefault("app.core.config", config)

from app.reports.aggregator import aggregate_submissions
from app.reports.excel_report import (
    DONT_ITEMS,
    NO_COMPROMISE_ITEMS,
    SUMMARY_POINT_COUNT,
    _layout_rows,
    fill_template_sheet,
    get_weekly_management_section,
)


PROJECT_ROOT = Path(__file__).resolve().parents[1]
FORM_PATH = PROJECT_ROOT / "templates" / "KB_Market_Improvement_XLSForm_GT_HORECA.xlsx"


def _submission(index: int, outlet_name: str, **overrides):
    values = {
        "id": index,
        "submission_time": datetime(2026, 9, 3, 8, index),
        "report_date": date(2026, 9, 3),
        "region": "R1",
        "dealer": "CA1",
        "group_no": 2,
        "member_no": 4,
        "location_text": "Phnom Penh",
        "outlet_name": outlet_name,
        "outlet_type": "Drink Shop",
        "submitter_name": None,
        "key_issue_text": None,
        "suggestion_text": None,
        "product_metrics": [],
        "competitor_metrics": [],
        "ring_pull_metrics": [],
    }
    values.update(overrides)
    return SimpleNamespace(**values)


class V153WeeklyDontSixSummaryTests(unittest.TestCase):
    def test_xlsform_removes_fall_point_and_supports_six_summary_points(self):
        workbook = load_workbook(FORM_PATH, read_only=True, data_only=False)
        survey = workbook["survey"]
        headers = [cell.value for cell in next(survey.iter_rows(min_row=1, max_row=1))]
        name_col = headers.index("name")
        label_col = headers.index("label")
        required_col = headers.index("required")

        rows_by_name = {}
        labels = []
        for row in survey.iter_rows(min_row=2, values_only=True):
            name = row[name_col]
            if name:
                rows_by_name[name] = row
            if row[label_col]:
                labels.append(str(row[label_col]))

        self.assertNotIn("submitter_name", rows_by_name)
        self.assertFalse(any("ចំណុចដួល" in label for label in labels))
        self.assertEqual(rows_by_name["key_issues_detail"][label_col], "បញ្ហាទីផ្សារ")
        self.assertEqual(
            rows_by_name["initiative_idea_suggestion"][label_col],
            "បញ្ហាត្រូវដោះស្រាយ",
        )
        required = str(rows_by_name["key_issues_detail"][required_col])
        self.assertIn("contains(", required)
        self.assertIn("សរុបរួម", required)
        self.assertIn("សរុបចុងក្រោយ", required)

        instruction = rows_by_name["summary_instruction"]
        self.assertIn("6 ចំណុច", str(instruction[headers.index("hint")]))
        workbook.close()

    def test_summary_parser_keeps_six_market_issues_and_actions(self):
        summary = _submission(
            2,
            "# ចែ ម៉ៅ, បូកសរុបរួម",
            submitter_name="Legacy value must be ignored",
            key_issue_text="\n".join(f"{i}. Issue {i}" for i in range(1, 8)),
            suggestion_text="\n".join(f"{i}. Action {i}" for i in range(1, 8)),
        )
        result = aggregate_submissions(
            [_submission(1, "Outlet A"), summary],
            wide_map={},
            own_product_names=[],
            competitor_product_names=[],
            include_ring_pull=False,
        )
        self.assertEqual(result["total_outlets"], 1)
        self.assertNotIn("fall_points", result)
        self.assertEqual(result["key_issues"], [f"Issue {i}" for i in range(1, 7)])
        self.assertEqual(result["suggestions"], [f"Action {i}" for i in range(1, 7)])

    def test_weekly_rotation_anchor(self):
        last_title, last_items = get_weekly_management_section(date(2026, 8, 24))
        current_title, current_items = get_weekly_management_section(date(2026, 9, 3))
        next_title, next_items = get_weekly_management_section(date(2026, 9, 7))

        self.assertEqual((last_title, last_items), ("No Compromise", NO_COMPROMISE_ITEMS))
        self.assertEqual((current_title, current_items), ("Don't", DONT_ITEMS))
        self.assertEqual((next_title, next_items), ("No Compromise", NO_COMPROMISE_ITEMS))

    def test_generated_reports_show_dont_and_six_issue_action_lines(self):
        cases = [
            ("template_general.xlsx", "GT", {"Drink Shop": 1}),
            ("template_horeca_products.xlsx", "HORECA", {"Local Eat": 1}),
        ]
        for template_name, channel, outlet_types in cases:
            with self.subTest(template=template_name):
                workbook = load_workbook(PROJECT_ROOT / "templates" / template_name)
                sheet = workbook.active
                agg = {
                    "dealer": "CA1",
                    "report_date": date(2026, 9, 3),
                    "total_outlets": 1,
                    "outlet_types": outlet_types,
                    "location_text": "Phnom Penh",
                    "group_no": 2,
                    "member_no": 4,
                    "channel": channel,
                    "products": {},
                    "competitors": {},
                    "ring_pull": {},
                    "key_issues": [f"Issue {i}" for i in range(1, 7)],
                    "suggestions": [f"Action {i}" for i in range(1, 7)],
                }
                fill_template_sheet(sheet, agg)
                layout = _layout_rows(sheet, agg)

                self.assertEqual(sheet.cell(layout["summary_header"], 1).value, "Don't")
                self.assertEqual(sheet.cell(layout["summary_header"], 9).value, "បញ្ហាទីផ្សារ")
                self.assertEqual(
                    sheet.cell(layout["summary_header"], 19).value,
                    "បញ្ហាត្រូវដោះស្រាយ",
                )
                for index in range(SUMMARY_POINT_COUNT):
                    row = layout["issue_start"] + index
                    self.assertEqual(sheet.cell(row, 9).value, f"{index + 1}. Issue {index + 1}")
                    self.assertEqual(sheet.cell(row, 19).value, f"{index + 1}. Action {index + 1}")
                self.assertEqual(sheet.cell(layout["issue_start"] + 6, 1).value, DONT_ITEMS[6])
                self.assertEqual(sheet.cell(layout["issue_start"] + 6, 9).value, "")
                self.assertEqual(sheet.cell(layout["issue_start"] + 6, 19).value, "")
                workbook.close()


if __name__ == "__main__":
    unittest.main()
