from __future__ import annotations

from datetime import date
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from app.reports.data_export import EXPORT_PRODUCTS, SUMMARY_HEADERS, create_data_export


def _submission(
    sid: str,
    outlet_name: str,
    outlet_type: str,
    lat: float,
    lon: float,
    product_mov: int,
    competitor_mov: int,
    member_no: int = 7,
    group_no: int = 2,
    location_text: str = "Area A",
):
    return SimpleNamespace(
        submission_id=sid,
        report_date=date(2026, 7, 18),
        region="R1",
        dealer="CA1",
        group_no=group_no,
        member_no=member_no,
        total_outlet_visit_target=20,
        outlet_name=outlet_name,
        outlet_type=outlet_type,
        is_new_outlet=False,
        submitter_name="Tester",
        phone_number="+588886631198",
        location_text=location_text,
        gps_text=f"{lat} {lon}",
        gps_latitude=lat,
        gps_longitude=lon,
        key_issue_text="",
        suggestion_text="",
        product_metrics=[
            SimpleNamespace(
                product_name="CB LITE ORD",
                status="sale",
                available=True,
                movement_score=product_mov,
                stock_status=None,
                bbe_date=None,
                buy_in_price=None,
                sell_out_price=None,
                ring_pull_value=None,
                new_outlet_purchase=False,
                volume_ctn=None,
            )
        ],
        competitor_metrics=[
            SimpleNamespace(
                product_name="GB SNOW ORD",
                status="sale",
                movement_score=competitor_mov,
                stock_status=None,
                buy_in_price=None,
                sell_out_price=None,
            )
        ],
        ring_pull_metrics=[],
    )


def _column_by_header(ws, header: str) -> int:
    for column in range(1, ws.max_column + 1):
        if ws.cell(1, column).value == header:
            return column
    raise AssertionError(f"Header not found: {header}")


def _find_product_row(ws, product: str) -> int:
    product_column = _column_by_header(ws, "Product")
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, product_column).value == product:
            return row
    raise AssertionError(f"Product row not found: {product}")


def test_data_export_uses_uploaded_template_columns_without_replacing_them(tmp_path: Path):
    project_root = Path(__file__).resolve().parents[1]
    template = project_root / "templates" / "Template_Data_Survey.xlsx"
    output = tmp_path / "export.xlsx"

    rows = [
        _submission("1", "Outlet One", "Wholesale", 12.08, 106.42, 4, 8, member_no=7),
        _submission("2", "Outlet Two", "Drink Shop", 11.56, 102.99, 6, 10, member_no=7),
        _submission("3", "Outlet Three", "Local Eat", 11.57, 103.00, 5, 9, member_no=8),
        # Control row must not be exported or counted as an outlet.
        _submission("4", "បូកសរុបរួម", "Wholesale", 0, 0, 10, 10, member_no=999),
    ]

    path, stats = create_data_export(
        rows,
        date(2026, 7, 18),
        template_path=template,
        output_path=output,
    )

    assert path == output
    assert stats["dealer_groups"] == 1
    assert stats["summary_rows"] == len(EXPORT_PRODUCTS)
    assert stats["location_rows"] == 3

    workbook = load_workbook(output, data_only=True)
    summary = workbook["Summary_Data"]
    location = workbook["Location_Outlet"]

    # The user's approved template headers and order must remain unchanged.
    assert [summary.cell(1, col).value for col in range(1, 16)] == SUMMARY_HEADERS
    assert summary.max_column == 15
    assert summary.max_row == len(EXPORT_PRODUCTS) + 1

    # Dealer values repeat per product, but Member shows only the most frequent value.
    assert summary.cell(2, _column_by_header(summary, "Region")).value == "R1"
    assert summary.cell(2, _column_by_header(summary, "Dealer")).value == "CA1"
    assert summary.cell(2, _column_by_header(summary, "Member")).value == "7"
    assert summary.cell(2, _column_by_header(summary, "Total Outlets")).value == 3
    assert summary.cell(2, _column_by_header(summary, "Wholesale")).value == 1
    assert summary.cell(2, _column_by_header(summary, "Drink Shop")).value == 1
    assert summary.cell(2, _column_by_header(summary, "Local Eat")).value == 1

    own_row = _find_product_row(summary, "CB LITE ORD")
    competitor_row = _find_product_row(summary, "GB SNOW ORD")
    movement_column = _column_by_header(summary, "Movement")
    assert summary.cell(own_row, movement_column).value is not None
    assert summary.cell(competitor_row, movement_column).value is not None

    assert location.max_row == 4
    assert [location.cell(1, col).value for col in range(1, 9)] == [
        "Date", "Region", "Dealer", "Latitude", "Longitude",
        "Outlet Name", "Outlet Type", "Phone Number Outlet",
    ]
    assert location["F2"].value == "Outlet Three" or location["F2"].value == "Outlet Two" or location["F2"].value == "Outlet One"
    assert all(location.cell(row, 8).number_format == "@" for row in range(2, 5))

    # Existing Excel table styles are kept and table ranges are expanded.
    assert summary.tables
    assert location.tables
    summary_table = next(iter(summary.tables.values()))
    location_table = next(iter(location.tables.values()))
    assert summary_table.ref == f"A1:O{len(EXPORT_PRODUCTS) + 1}"
    assert location_table.ref == "A1:H4"
